"""Excel report builder — openpyxl으로 다중 시트 리포트 생성.

view layer only — 비즈니스 로직 없음. analysis 결과를 받아 표시만.
"""

from __future__ import annotations

from pathlib import Path

import duckdb
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from peer_benchmarking.analysis import cross_section, measurement_model, ratios
from peer_benchmarking.analysis.queries import QuerySpec
from peer_benchmarking.domain import peer_groups
from peer_benchmarking.render.formatters import humanize_peer_table, to_trillion

# ─── styling ────────────────────────────────────────────────────────────

SELF_FILL = PatternFill("solid", fgColor="FFF59D")  # 자사 노란색
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
SECTOR_FILL = {
    "life": PatternFill("solid", fgColor="E3F2FD"),
    "non_life": PatternFill("solid", fgColor="FFF3E0"),
    "reinsurance": PatternFill("solid", fgColor="F3E5F5"),
}
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(italic=True, size=10, color="666666")
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _write_table(ws, df: pd.DataFrame, start_row: int, self_col: str = "자사여부") -> int:
    """Write df starting at start_row. Returns row after last data row."""
    if df.empty:
        ws.cell(row=start_row, column=1, value="(데이터 없음)")
        return start_row + 1

    # Header
    for ci, col in enumerate(df.columns, start=1):
        cell = ws.cell(row=start_row, column=ci, value=str(col))
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = BORDER

    # Rows — figure out which column index marks the "self" highlight
    self_col_idx = list(df.columns).index(self_col) if self_col in df.columns else None
    for ri, row in enumerate(df.itertuples(index=False, name=None), start=start_row + 1):
        is_self = bool(row[self_col_idx]) if self_col_idx is not None else False
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = BORDER
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                cell.alignment = Alignment(horizontal="right")
                if isinstance(val, float):
                    cell.number_format = "#,##0.00"
            if is_self:
                cell.fill = SELF_FILL
    return start_row + len(df) + 1


def _autosize(ws, df: pd.DataFrame, min_w: int = 8) -> None:
    for ci, col in enumerate(df.columns, start=1):
        # Korean chars take ~2 cells in width; rough heuristic
        header_w = sum(2 if ord(ch) > 127 else 1 for ch in str(col)) + 2
        max_data = max(
            (sum(2 if ord(ch) > 127 else 1 for ch in str(v)) for v in df[col]),
            default=0,
        )
        w = max(header_w, max_data + 2, min_w)
        ws.column_dimensions[get_column_letter(ci)].width = min(w, 35)


def _write_title(ws, title: str, subtitle: str | None = None) -> int:
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    if subtitle:
        ws.cell(row=2, column=1, value=subtitle).font = SUBTITLE_FONT
        return 4
    return 3


def _write_summary_box(ws, summary: dict, start_row: int) -> int:
    """Write the self_percentile result as a small fact box."""
    rows = [
        ("자사", summary.get("name_ko", "")),
        ("값", f"{summary.get('value', 0)/1e12:,.2f} 조원" if summary.get("value") else "N/A"),
        ("Peer 중위값", f"{summary.get('median', 0)/1e12:,.2f} 조원"),
        ("Peer 평균", f"{summary.get('mean', 0)/1e12:,.2f} 조원"),
        ("순위 (1=가장 작음)", f"{summary.get('rank', '?')} / {summary.get('n_peers', '?')}"),
        ("백분위(%)", f"{summary.get('percentile', 0):.1f}%"),
    ]
    for i, (k, v) in enumerate(rows):
        ws.cell(row=start_row + i, column=1, value=k).font = Font(bold=True)
        ws.cell(row=start_row + i, column=2, value=v)
    return start_row + len(rows) + 1


def _add_sheet(
    wb: Workbook,
    title: str,
    raw_df: pd.DataFrame,
    subtitle: str | None = None,
    summary: dict | None = None,
    sheet_name: str | None = None,
) -> None:
    """Add one analysis sheet: title + summary box + annotated peer table."""
    name = sheet_name or title[:30]
    ws = wb.create_sheet(name)

    if raw_df.empty:
        ws.cell(row=1, column=1, value=f"{title} — 데이터 없음").font = TITLE_FONT
        return

    row = _write_title(ws, title, subtitle)
    if summary:
        row = _write_summary_box(ws, summary, row)
    annotated = ratios.add_peer_relative_columns(raw_df)
    display = humanize_peer_table(annotated)
    _write_table(ws, display, start_row=row)
    _autosize(ws, display)


def _add_model_share_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    """PAA vs Non-PAA 비중 시트 (별도 포맷)."""
    ws = wb.create_sheet("회계모형(PAA·Non-PAA)")
    row = _write_title(
        ws,
        "회계모형 분포: PAA vs Non-PAA",
        "PAA=Premium Allocation Approach(단기·손보), Non-PAA=GMM+VFA(장기·생보)",
    )
    if df.empty:
        ws.cell(row=row, column=1, value="(데이터 없음)")
        return

    self_cik = peer_groups.self_cik()
    display = df.copy()
    display["회사"] = display["name_ko"]
    display["섹터"] = display["sector"].map({"life": "생명", "non_life": "손해", "reinsurance": "재보험"})
    display["PAA(조원)"] = to_trillion(display["paa_amount"])
    display["Non-PAA(조원)"] = to_trillion(display["nonpaa_amount"])
    display["합계(조원)"] = to_trillion(display["total"])
    display["PAA비중(%)"] = (display["paa_share"].fillna(0) * 100).round(1)
    display["Non-PAA비중(%)"] = (display["nonpaa_share"].fillna(0) * 100).round(1)
    display["자사여부"] = display["cik"] == self_cik
    cols = ["회사", "섹터", "PAA(조원)", "Non-PAA(조원)", "합계(조원)",
            "PAA비중(%)", "Non-PAA비중(%)", "자사여부"]
    _write_table(ws, display[cols], start_row=row)
    _autosize(ws, display[cols])


def _add_component_sheet(wb: Workbook, df: pd.DataFrame) -> None:
    """BEL/RA/CSM 구성요소 분해 시트 (pivot)."""
    ws = wb.create_sheet("구성요소(BEL·RA·CSM)")
    row = _write_title(
        ws,
        "보험계약부채 구성요소 분해 (BEL / RA / CSM)",
        "표준 ifrs-full_InsuranceContractsByComponentsAxis 기준. 한화생명만 CSM을 표준 axis로 보고함.",
    )
    if df.empty:
        ws.cell(row=row, column=1, value="(데이터 없음)")
        return

    self_cik = peer_groups.self_cik()
    pivot = df.pivot_table(
        index=["cik", "name_ko", "sector"],
        columns="component",
        values="amount_krw",
        aggfunc="first",
    ).fillna(0).reset_index()
    for c in ("BEL", "RA", "CSM"):
        if c not in pivot.columns:
            pivot[c] = 0.0
    pivot["합계"] = pivot["BEL"] + pivot["RA"] + pivot["CSM"]
    pivot = pivot.sort_values("합계", ascending=False)
    display = pd.DataFrame({
        "회사": pivot["name_ko"],
        "섹터": pivot["sector"].map({"life": "생명", "non_life": "손해", "reinsurance": "재보험"}),
        "BEL(조원)": to_trillion(pivot["BEL"]),
        "RA(조원)": to_trillion(pivot["RA"]),
        "CSM(조원)": to_trillion(pivot["CSM"]),
        "합계(조원)": to_trillion(pivot["합계"]),
        "CSM비중(%)": (pivot["CSM"] / pivot["합계"].where(pivot["합계"] != 0) * 100).round(1).fillna(0),
        "자사여부": pivot["cik"] == self_cik,
    })
    _write_table(ws, display, start_row=row)
    _autosize(ws, display)


def build_report(
    con: duckdb.DuckDBPyConnection,
    output_path: Path,
    report_date: str = "20251231",
    period_instant: str = "2025-12-31",
    period_start: str = "2025-01-01",
    period_end: str = "2025-12-31",
) -> Path:
    """Build the full Excel report. Returns the output path.

    Sheets:
        1. 개요              — 자사 요약 (보험계약부채/자산/보험수익/percentile)
        2. 보험계약부채(생보) — 생보 4개사 횡단
        3. 보험계약부채(11개) — 전체 보험사 횡단
        4. 자산(11개)        — 비교용 자산 cross-section
        5. 보험수익(생보)    — P&L (duration)
        6. 보험서비스결과    — P&L
        7. 회계모형(PAA·Non-PAA) — 11개사
        8. 구성요소(BEL·RA·CSM) — 생보 4개사
    """
    wb = Workbook()
    # Replace default sheet
    wb.remove(wb.active)

    spec_life = QuerySpec(report_date=report_date, consolidation="consolidated", peer_group="life")
    spec_all = QuerySpec(report_date=report_date, consolidation="consolidated", peer_group="all_insurers")

    # --- 1. Overview ---
    overview = wb.create_sheet("개요")
    row = _write_title(
        overview,
        f"DART XBRL 동업사 비교분석 — {report_date[:4]}년 사업연도",
        f"미래에셋생명 (CIK 00112332) 기준 · 작성: {report_date[:4]}-12-31 · KOSPI 상장 보험사 11개사",
    )

    # Pull key metrics
    life_lib = cross_section.liability_balance(con, spec_life, period_instant=period_instant)
    all_lib = cross_section.liability_balance(con, spec_all, period_instant=period_instant)
    all_assets = cross_section.liability_balance(con, spec_all, item_name="total_asset", period_instant=period_instant)

    self_lib = ratios.self_percentile(life_lib)
    self_lib_all = ratios.self_percentile(all_lib)

    overview.cell(row=row, column=1, value="핵심 요약").font = Font(bold=True, size=12)
    row += 1
    fields = [
        ("자사 보험계약부채", f"{self_lib['value']/1e12:,.2f} 조원" if "value" in self_lib else "N/A"),
        ("생보 4개사 중 순위", f"{self_lib.get('rank','?')} / {self_lib.get('n_peers','?')} (백분위 {self_lib.get('percentile',0):.1f}%)"),
        ("생보 4개사 중위 대비", f"{self_lib['value']/self_lib['median']:.2f}x" if "value" in self_lib else "N/A"),
        ("전체 11개사 중 순위", f"{self_lib_all.get('rank','?')} / {self_lib_all.get('n_peers','?')} (백분위 {self_lib_all.get('percentile',0):.1f}%)"),
        ("", ""),
        ("주의사항", "결산 후 정정 가능 · 별도/연결 표시는 연결 기준"),
        ("데이터 출처", "DART 공시 XBRL 주석 (opendart.fss.or.kr)"),
    ]
    for k, v in fields:
        overview.cell(row=row, column=1, value=k).font = Font(bold=True)
        overview.cell(row=row, column=2, value=v)
        row += 1
    overview.column_dimensions["A"].width = 25
    overview.column_dimensions["B"].width = 50

    # --- 2-3. 보험계약부채 ---
    _add_sheet(
        wb, "보험계약부채 (생보 4개사)", life_lib,
        subtitle="자사 + 동업사 3개. ifrs-full_InsuranceContractsThatAreLiabilities",
        summary=self_lib,
        sheet_name="보험계약부채(생보)",
    )
    _add_sheet(
        wb, "보험계약부채 (전체 11개사)", all_lib,
        subtitle="시장 분포 컨텍스트용. 자사 percentile.",
        summary=self_lib_all,
        sheet_name="보험계약부채(11개)",
    )

    # --- 4. Assets ---
    self_asset = ratios.self_percentile(all_assets)
    _add_sheet(
        wb, "자산인 보험계약 (전체 11개사)", all_assets,
        subtitle="ifrs-full_InsuranceContractsThatAreAssets",
        summary=self_asset,
        sheet_name="자산인 보험계약",
    )

    # --- 5. P&L: 보험수익 ---
    rev = cross_section.pl_item(con, spec_life, "insurance_revenue", period_start, period_end)
    self_rev = ratios.self_percentile(rev) if not rev.empty else {}
    _add_sheet(
        wb, "보험수익 (생보 4개사 · 2025 연간)", rev,
        subtitle="ifrs-full_InsuranceRevenue · duration period",
        summary=self_rev,
        sheet_name="보험수익(생보)",
    )

    # --- 6. P&L: 보험서비스결과 ---
    isr = cross_section.pl_item(con, spec_life, "insurance_service_result", period_start, period_end)
    self_isr = ratios.self_percentile(isr) if not isr.empty else {}
    _add_sheet(
        wb, "보험서비스결과 (생보 4개사)", isr,
        subtitle="ifrs-full_InsuranceServiceResult = 보험수익 − 보험서비스비용",
        summary=self_isr,
        sheet_name="보험서비스결과",
    )

    # --- 7. Measurement model ---
    share = measurement_model.model_share(con, spec_all, period_instant=period_instant)
    _add_model_share_sheet(wb, share)

    # --- 8. Components decomposition ---
    comps = cross_section.component_decomposition(con, spec_life, period_instant=period_instant)
    _add_component_sheet(wb, comps)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    return output_path
