"""BEL 변동 disclosure — 13행 cross-tab, 7개 동업사 XBRL 추출.

전략:
- 미래에셋: 사용자 disclosure (PDF) 값 그대로 (XBRL 에 일부 행은 매칭되는 fact 없음)
- 7개사: ifrs-full 표준 element 의 정확한 매칭 (DI817105, Sep × Issued, 비-Component, 비-위험관리)
- 13행 라벨 → ifrs-full element_id 캐노니컬 매핑 (taxonomy 라벨 의미 기반)
- 잔차 검증: 기초 + Σ변동 = 기말, 잔차 표시
"""
from __future__ import annotations
import duckdb
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

PEERS = [
    ("00112332", "미래에셋생명"),
    ("00126256", "삼성생명"),
    ("00113058", "한화생명"),
    ("00117267", "동양생명"),
    ("00139214", "삼성화재"),
    ("00164973", "현대해상"),
    ("00159102", "DB손해보험"),
    ("00135917", "한화손해보험"),
]
SEP = "ifrs-full_SeparateMember"
ISSUED = "ifrs-full_InsuranceContractsIssuedMember"
ROLE = "dart_2024-06-30_role-DI817105"

# 13행 라벨 + 미래에셋 사용자 disclosure 값 (원 단위)
ROWS = [
    ("부채인 보험계약의 기초 장부금액",            "opening", 26_248_411_835_353),
    ("보험수익",                                  "duration",-1_080_411_547_554),
    ("발생한 보험금 및 기타 보험서비스비용",      "duration",  +592_626_787_607),
    ("보험취득현금흐름 상각",                     "duration",  +218_864_437_494),
    ("발생사고요소의 조정",                       "duration",   +19_584_794_969),
    ("손실부담계약 관련 손실(환입)",              "duration",   +66_710_164_165),
    ("수취한 보험료",                             "duration",+4_050_027_788_462),
    ("보험취득현금흐름 지급",                     "duration",  -594_834_366_070),
    ("보험금(투자요소 포함) 및 기타보험서비스비용의 지급", "duration", -4_296_950_166_750),
    ("당기손익인식 보험금융손익",                 "duration",+2_272_778_949_192),
    ("기타포괄손익인식 보험금융손익",             "duration",  -483_472_446_698),
    ("기타증감",                                  "duration",   -16_637_974_642),
    ("부채인 보험계약의 기말 장부금액",           "closing", 26_996_698_255_523),
]

# 라벨 → 캐노니컬 ifrs-full element_id (사업보고서 21-1 표준 분해)
# Sign convention: 부채 증가 = (+), 부채 감소 = (-). 사용자값 부호와 동일.
# 일부 행은 회사별 fallback element 사용 (riskier — comment 표시)
MAPPING = {
    "부채인 보험계약의 기초 장부금액": {
        "primary": "ifrs-full_InsuranceContractsLiabilityAsset",  # net BS
        "fallback": "ifrs-full_InsuranceContractsThatAreLiabilities",  # 부채만 (gross)
        "period": "opening",
    },
    "부채인 보험계약의 기말 장부금액": {
        "primary": "ifrs-full_InsuranceContractsLiabilityAsset",
        "fallback": "ifrs-full_InsuranceContractsThatAreLiabilities",
        "period": "closing",
    },
    "보험수익": {
        "primary": "ifrs-full_IncreaseDecreaseThroughInsuranceRevenueInsuranceContractsLiabilityAsset",
        "period": "duration",
    },
    "발생한 보험금 및 기타 보험서비스비용": {
        # taxonomy: IncurredClaimsAndOther...ExpensesAdjustment (LIC 발생인식, +방향)
        "primary": "ifrs-full_IncreaseDecreaseThroughIncurredClaimsAndOtherIncurredInsuranceServiceExpensesAdjustmentInsuranceContractsLiabilityAsset",
        "period": "duration",
    },
    "보험취득현금흐름 상각": {
        "primary": "ifrs-full_IncreaseDecreaseThroughAmortisationOfInsuranceAcquisitionCashFlowsInsuranceContractsLiabilityAsset",
        "period": "duration",
    },
    "발생사고요소의 조정": {
        # taxonomy: ChangesThatRelateToPastService (LIC 발생사고 추정변경)
        "primary": "ifrs-full_IncreaseDecreaseThroughChangesThatRelateToPastServiceInsuranceContractsLiabilityAsset",
        "period": "duration",
    },
    "손실부담계약 관련 손실(환입)": {
        # taxonomy: EffectsOfGroupsOfOnerousContractsInitiallyRecognised
        "primary": "ifrs-full_IncreaseDecreaseThroughEffectsOfGroupsOfOnerousContractsInitiallyRecognisedInPeriodInsuranceContractsLiabilityAsset",
        "period": "duration",
    },
    "수취한 보험료": {
        "primary": "ifrs-full_IncreaseDecreaseThroughPremiumsReceivedForInsuranceContractsIssuedInsuranceContractsLiabilityAsset",
        "period": "duration",
    },
    "보험취득현금흐름 지급": {
        "primary": "ifrs-full_IncreaseDecreaseThroughInsuranceAcquisitionCashFlowsInsuranceContractsLiabilityAsset",
        "period": "duration",
    },
    "보험금(투자요소 포함) 및 기타보험서비스비용의 지급": {
        "primary": "ifrs-full_IncreaseDecreaseThroughIncurredClaimsPaidAndOtherInsuranceServiceExpensesPaidForInsuranceContractsIssuedExcludingInsuranceAcquisitionCashFlowsInsuranceContractsLiabilityAsset",
        "period": "duration",
    },
    "당기손익인식 보험금융손익": {
        # 회사마다 두 element 중 하나 사용:
        # (a) ifrs-full_InsuranceFinanceIncomeExpensesFromInsuranceContractsIssuedRecognisedInProfitOrLoss (P&L 라인 직접)
        # (b) ifrs-full_IncreaseDecreaseThroughInsuranceFinanceIncomeOrExpensesInsuranceContractsLiabilityAsset (BS 변동 라인, OCI 포함)
        # 21-1 disclosure 는 P&L 만 분리 → (a) 우선, 없으면 (b)
        "primary": "ifrs-full_InsuranceFinanceIncomeExpensesFromInsuranceContractsIssuedRecognisedInProfitOrLoss",
        "period": "duration",
        # sign: 사용자 mirae 값은 +22,728억 (부채증가). 해당 element 는 P&L 비용 (부채 증가 ↔ 비용). 부호 확인 필요.
    },
    "기타포괄손익인식 보험금융손익": {
        "primary": "ifrs-full_OtherComprehensiveIncomeBeforeTaxInsuranceFinanceIncomeExpensesFromInsuranceContractsIssuedExcludedFromProfitOrLoss",
        "period": "duration",
    },
    "기타증감": {
        # 잔여 — taxonomy: AdditionalItemsNecessaryToUnderstandChange
        "primary": "ifrs-full_IncreaseDecreaseThroughAdditionalItemsNecessaryToUnderstandChangeInNetCarryingAmountOfInsuranceContractsInsuranceContractsLiabilityAsset",
        "fallback": "ifrs-full_IncreaseDecreaseThroughTransfersAndOtherChangesEquity",
        "period": "duration",
    },
}

# 미래에셋 사용자 disclosure 값 직접
MIRAE_USER = {(label, pk): v for label, pk, v in ROWS}

con = duckdb.connect("data/db/benchmark.duckdb", read_only=True)

SQL_VAL = """
WITH ctxs AS (
  SELECT CONTEXT_ID, COUNT(*) AS n_ax,
         MAX(PERIOD_INSTANT) AS p_inst,
         MAX(PERIOD_START_DATE) AS p_start,
         MAX(PERIOD_END_DATE) AS p_end
  FROM cntxt_insurers WHERE CIK=? AND REPORT_DATE='20251231'
  GROUP BY CONTEXT_ID
  HAVING BOOL_OR(MEMBER_ELEMENT_ID = ?) AND BOOL_OR(MEMBER_ELEMENT_ID = ?)
     AND NOT BOOL_OR(MEMBER_ELEMENT_ID LIKE '%OfDisclosureOfNatureAndExtentOfRisks%')
     AND NOT BOOL_OR(AXIS_ELEMENT_ID = 'ifrs-full_InsuranceContractsByComponentsAxis')
),
fact AS (
  SELECT v.amount_krw, c.n_ax,
         CASE WHEN c.p_inst='2024-12-31' THEN 'opening'
              WHEN c.p_inst='2025-12-31' THEN 'closing'
              WHEN c.p_start='2025-01-01' AND c.p_end='2025-12-31' THEN 'duration'
              ELSE NULL END AS pk
  FROM val_insurers v JOIN ctxs c USING(CONTEXT_ID)
  WHERE v.CIK=? AND v.REPORT_DATE='20251231' AND v.ELEMENT_ID=? AND v.amount_krw IS NOT NULL
),
top_only AS (
  SELECT MIN(n_ax) AS min_ax FROM fact WHERE pk=?
)
SELECT SUM(f.amount_krw) FROM fact f, top_only t
WHERE f.n_ax = t.min_ax AND f.pk = ?
"""

def get_value(cik, eid, pk):
    r = con.execute(SQL_VAL, [cik, SEP, ISSUED, cik, eid, pk, pk]).fetchone()
    return r[0] if r and r[0] is not None else None

# 회사 × 행 값 + 매핑 추적
peer_values = {}       # (cik, label) -> value
peer_eid_used = {}     # (cik, label) -> element_id 또는 "(미공시)" / "(미래에셋 PDF)"

for cik, name in PEERS:
    for label, pk, _ in ROWS:
        key = (cik, label)
        if cik == "00112332":
            peer_values[key] = MIRAE_USER[(label, pk)]
            peer_eid_used[key] = "(미래에셋 PDF disclosure 직접)"
            continue
        spec = MAPPING.get(label)
        if not spec:
            peer_values[key] = None
            peer_eid_used[key] = "(매핑 정의 없음)"
            continue
        # primary 시도
        v = get_value(cik, spec["primary"], spec["period"])
        used = spec["primary"]
        if v is None and spec.get("fallback"):
            v = get_value(cik, spec["fallback"], spec["period"])
            used = spec["fallback"] + " (fallback)"
        peer_values[key] = v
        peer_eid_used[key] = used.replace("ifrs-full_", "") if v is not None else "(미공시)"

# 검증식 — 회사별 잔차 (원 단위)
def residual(cik):
    op = peer_values.get((cik, "부채인 보험계약의 기초 장부금액")) or 0
    cl = peer_values.get((cik, "부채인 보험계약의 기말 장부금액")) or 0
    dur_sum = 0
    for label, pk, _ in ROWS:
        if pk == "duration":
            v = peer_values.get((cik, label))
            if v is not None:
                dur_sum += v
    return op + dur_sum - cl

# Excel
wb = Workbook()
ws = wb.active
ws.title = "BEL 변동 disclosure"

ws.append(["§4-1A 보험계약부채 변동표 — 동업사 cross-tab (별도, 발행, 억원)"])
ws.cell(row=1, column=1).font = Font(bold=True, size=14)
ws.append(["기준: 미래에셋 사업보고서 「21-1. 보험계약부채 변동분의 차이조정 공시」 13행 구조"])
ws.cell(row=2, column=1).font = Font(italic=True, color="666666", size=10)
ws.append(["미래에셋 = 사업보고서 PDF disclosure · 7개사 = XBRL ifrs-full 표준 element 매칭 (DI817105, Sep×Issued, 비-Component·비-위험관리)"])
ws.cell(row=3, column=1).font = Font(italic=True, color="C62828", size=10)
ws.append([])

hdrs = ["#", "disclosure 행 라벨", "기간"] + [n for _, n in PEERS]
ws.append(hdrs)
hr = ws.max_row
for c in range(1, len(hdrs)+1):
    cell = ws.cell(row=hr, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1E3A5F")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

opening_fill = PatternFill("solid", fgColor="FFF4CE")
closing_fill = PatternFill("solid", fgColor="FFE0B2")
revenue_fill = PatternFill("solid", fgColor="E3F2FD")
expense_fill = PatternFill("solid", fgColor="FFF3E0")
cf_fill      = PatternFill("solid", fgColor="E0F7FA")
fin_fill     = PatternFill("solid", fgColor="F3E5F5")
other_fill   = PatternFill("solid", fgColor="FAFAFA")

border = Border(left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"))

def fill_for(label, pk):
    if pk == "opening": return opening_fill
    if pk == "closing": return closing_fill
    if "보험수익" in label: return revenue_fill
    if any(w in label for w in ["발생한 보험금", "보험취득현금흐름 상각", "발생사고요소", "손실부담"]): return expense_fill
    if any(w in label for w in ["수취한", "지급", "투자요소"]): return cf_fill
    if "보험금융손익" in label: return fin_fill
    return other_fill

PK_LABEL = {"opening": "기초", "duration": "변동", "closing": "기말"}
for i, (label, pk, _) in enumerate(ROWS, 1):
    row = [i, label, PK_LABEL[pk]]
    for cik, _ in PEERS:
        v = peer_values.get((cik, label))
        row.append(round(v / 1e8, 0) if v is not None else None)
    ws.append(row)
    r = ws.max_row
    fill = fill_for(label, pk)
    for c in range(1, len(hdrs)+1):
        ws.cell(row=r, column=c).fill = fill
        ws.cell(row=r, column=c).border = border
    for c in range(4, len(hdrs)+1):
        cell = ws.cell(row=r, column=c)
        if cell.value is None:
            cell.value = "미공시"
            cell.font = Font(italic=True, color="999999", size=9)
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.number_format = '#,##0;-#,##0;"–"'

# 검증식 행
ws.append([])
ws.append(["검증", "잔차 (기초 + Σ변동 − 기말, 억원)", "", ""] + [""]*len(PEERS))
sum_row_idx = ws.max_row
for i, (cik, _) in enumerate(PEERS, start=4):
    res = residual(cik)
    ws.cell(row=sum_row_idx, column=i).value = round(res / 1e8, 0)
    ws.cell(row=sum_row_idx, column=i).number_format = '#,##0;-#,##0;"– (정합)"'
    # 잔차 |x| < 1000억 (≈ 1조 1%) 이면 정합 OK
    if abs(res) < 100_000_000_000:
        ws.cell(row=sum_row_idx, column=i).font = Font(bold=True, color="2E7D32")
    else:
        ws.cell(row=sum_row_idx, column=i).font = Font(bold=True, color="C62828")
for c in range(1, len(hdrs)+1):
    ws.cell(row=sum_row_idx, column=c).fill = PatternFill("solid", fgColor="E8F5E9")
    ws.cell(row=sum_row_idx, column=c).border = border
ws.cell(row=sum_row_idx, column=1).font = Font(bold=True)
ws.cell(row=sum_row_idx, column=2).font = Font(bold=True)

# Caveat 행
ws.append([])
caveat_row = ws.max_row + 1
ws.cell(row=caveat_row, column=1).value = "주석"
ws.cell(row=caveat_row, column=1).font = Font(bold=True, color="C62828")
ws.cell(row=caveat_row, column=2).value = (
    "회사별 disclosure 분해 방식 차이로 일부 행이 정확히 매칭되지 않음. "
    "특히 '발생한 보험금/취득CF상각/발생사고요소조정/손실부담계약' 4개 행은 회사별 분해 단위가 달라 "
    "ifrs-full 표준 element 가 분해 단위와 일치하지 않을 수 있음. 잔차 규모로 정합성 평가."
)
ws.cell(row=caveat_row, column=2).alignment = Alignment(wrap_text=True, vertical="top")
ws.cell(row=caveat_row, column=2).font = Font(italic=True, color="666666", size=9)
ws.merge_cells(start_row=caveat_row, start_column=2, end_row=caveat_row, end_column=len(hdrs))
ws.row_dimensions[caveat_row].height = 45

# 컬럼 폭
widths = [4, 40, 6] + [13]*len(PEERS)
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[hr].height = 28
ws.freeze_panes = f"D{hr+1}"

# 진단 시트: element 매핑
diag = wb.create_sheet(title="element 매핑")
diag.append(["회사별 disclosure 행 라벨 → 사용된 element_id (ifrs-full prefix 생략)"])
diag.cell(row=1, column=1).font = Font(bold=True, size=12)
diag.append(["미공시 = 해당 회사 XBRL 에 매칭되는 element 없음. fallback = primary 없어서 대체 element 사용."])
diag.cell(row=2, column=1).font = Font(italic=True, color="666666", size=9)
diag.append([])
diag.append(["disclosure 행"] + [n for _, n in PEERS])
hr2 = diag.max_row
for c in range(1, 2 + len(PEERS)):
    cell = diag.cell(row=hr2, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1E3A5F")
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

for label, pk, _ in ROWS:
    row = [label]
    for cik, _ in PEERS:
        row.append(peer_eid_used.get((cik, label), ""))
    diag.append(row)
    r = diag.max_row
    for c in range(1, 2 + len(PEERS)):
        diag.cell(row=r, column=c).border = border
        diag.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)
        if "미공시" in str(diag.cell(row=r, column=c).value or ""):
            diag.cell(row=r, column=c).font = Font(color="C62828", size=9, italic=True)

for i, w in enumerate([35] + [30]*len(PEERS), 1):
    diag.column_dimensions[get_column_letter(i)].width = w
diag.freeze_panes = f"B{hr2+1}"

# 매핑 규칙 시트
rule = wb.create_sheet(title="매핑 규칙")
rule.append(["13행 라벨 → 캐노니컬 ifrs-full element"])
rule.cell(row=1, column=1).font = Font(bold=True, size=12)
rule.append([])
rule.append(["#", "disclosure 행 라벨", "기간", "primary element_id", "fallback element_id"])
hr3 = rule.max_row
for c in range(1, 6):
    cell = rule.cell(row=hr3, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1E3A5F")
for i, (label, pk, _) in enumerate(ROWS, 1):
    spec = MAPPING.get(label, {})
    rule.append([i, label, PK_LABEL[pk],
                 spec.get("primary", "").replace("ifrs-full_", ""),
                 spec.get("fallback", "").replace("ifrs-full_", "") if spec.get("fallback") else ""])
for i, w in enumerate([4, 38, 6, 65, 50], 1):
    rule.column_dimensions[get_column_letter(i)].width = w

out = Path("outputs/bel_disclosure_filled.xlsx")
wb.save(out)
print(f"wrote {out}")
print(f"sheets: {wb.sheetnames}")

# 잔차 정합 보고서 stdout
print("\n=== 회사별 잔차 (기초 + Σ변동 − 기말, 억원) ===")
for cik, name in PEERS:
    res = residual(cik)
    flag = "OK" if abs(res) < 100_000_000_000 else "잔차큼"
    print(f"  {name:<12s} {res/1e8:>+12,.0f}억  [{flag}]")
