"""§4-1A 보험계약부채 변동표 — 회사별 충실 재현 + 핵심 라인만 cross-tab.

전략: 매핑 자동화 포기, 회사별 사업보고서 disclosure 그대로 재현(per-company sheet),
       명확한 표준 element만 cross-tab.

회사별 시트 = DI817105 role의 pre_insurers 트리를 walk하여 disclosure 라벨/들여쓰기 유지
핵심 비교 시트 = 기초·기말·보험수익·보험서비스결과 (P&L 명확 element만)
"""
from __future__ import annotations
import duckdb
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

PEERS = [
    ("00112332", "미래에셋생명"),
    ("00126256", "삼성생명"),
    ("00113058", "한화생명"),
    ("00117267", "동양생명"),
    ("00139214", "삼성화재"),
    ("00164973", "현대해상"),
    ("00159102", "DB손해보험"),
    ("00135917", "한화손해보험"),
]
SEP = "ifrs-full_SeparateMember"
ISSUED = "ifrs-full_InsuranceContractsIssuedMember"
ROLE = "dart_2024-06-30_role-DI817105"  # 별도 변동표

con = duckdb.connect("data/db/benchmark.duckdb", read_only=True)

# 회사별 pre tree
SQL_TREE = """
SELECT p.ELEMENT_ID, p.PARENT_ELEMENT_ID, p.ORDER, p.PREFERREDLABEL
FROM pre_insurers p
WHERE p.CIK=? AND p.REPORT_DATE='20251231' AND p.ROLE_ID=?
ORDER BY p."ORDER"
"""

# 라벨 lookup
SQL_LABEL = """
SELECT ELMT_ID,
       MAX(CASE WHEN LABEL_ROLE_URI LIKE '%terseLabel%' THEN LABEL END) AS terse,
       MAX(CASE WHEN LABEL_ROLE_URI='http://www.xbrl.org/2003/role/label' THEN LABEL END) AS dflt
FROM lab_insurers WHERE CIK=? AND REPORT_DATE='20251231' AND LANG='ko'
GROUP BY ELMT_ID
"""

# element 값: 회사별 (element, period_kind)의 최소 axis depth fact만 합산
SQL_VALUES = """
WITH ctxs AS (
  SELECT CONTEXT_ID, COUNT(*) AS n_ax,
         MAX(PERIOD_INSTANT) AS p_inst,
         MAX(PERIOD_START_DATE) AS p_start,
         MAX(PERIOD_END_DATE) AS p_end
  FROM cntxt_insurers WHERE CIK=? AND REPORT_DATE='20251231'
  GROUP BY CONTEXT_ID
  HAVING BOOL_OR(MEMBER_ELEMENT_ID = ?) AND BOOL_OR(MEMBER_ELEMENT_ID = ?)
     AND NOT BOOL_OR(MEMBER_ELEMENT_ID LIKE '%OfDisclosureOfNatureAndExtentOfRisks%')
     AND NOT BOOL_OR(AXIS_ELEMENT_ID = 'ifrs-full_InsuranceContractsByComponentsAxis')
),
fact AS (
  SELECT v.ELEMENT_ID,
         CASE WHEN c.p_inst='2024-12-31' THEN 'opening'
              WHEN c.p_inst='2025-12-31' THEN 'closing'
              WHEN c.p_start='2025-01-01' AND c.p_end='2025-12-31' THEN 'duration'
              ELSE NULL END AS pk,
         c.n_ax, v.amount_krw
  FROM val_insurers v JOIN ctxs c USING(CONTEXT_ID)
  WHERE v.CIK=? AND v.REPORT_DATE='20251231' AND v.amount_krw IS NOT NULL
    AND (c.p_inst IN ('2024-12-31','2025-12-31')
         OR (c.p_start='2025-01-01' AND c.p_end='2025-12-31'))
),
top_only AS (
  SELECT ELEMENT_ID, pk, MIN(n_ax) AS min_ax FROM fact
  WHERE pk IS NOT NULL GROUP BY ELEMENT_ID, pk
)
SELECT f.ELEMENT_ID, f.pk, SUM(f.amount_krw)/1e8
FROM fact f JOIN top_only t USING(ELEMENT_ID, pk)
WHERE f.n_ax = t.min_ax
GROUP BY f.ELEMENT_ID, f.pk
"""

# ─── helpers ────────────────────────────────────────────────────────────────
def fetch_tree_and_vals(cik):
    """회사 1개의 pre tree + 라벨 + 값 모두."""
    tree = con.execute(SQL_TREE, [cik, ROLE]).fetchall()
    lbls = {r[0]: (r[1] or r[2] or "")
            for r in con.execute(SQL_LABEL, [cik]).fetchall()}
    vals = {}  # (eid, pk) -> amount
    for eid, pk, amt in con.execute(SQL_VALUES, [cik, SEP, ISSUED, cik]).fetchall():
        vals[(eid, pk)] = amt
    return tree, lbls, vals


def short_eid(eid):
    s = eid.replace("ifrs-full_", "").replace("dart_", "d:")
    for cik, _ in PEERS:
        s = s.replace(f"entity{cik}_", "[E]")
    return s


def walk_tree(tree, lbls, parent=None, depth=0, visited=None):
    """parent 기준 children 재귀 yield → (depth, eid, label)."""
    if visited is None:
        visited = set()
    # ORDER 순 children
    children = sorted([t for t in tree if t[1] == parent and t[0] not in visited],
                      key=lambda x: (x[2] or 0))
    for eid, par, order, pref in children:
        if eid in visited:
            continue
        visited.add(eid)
        label = lbls.get(eid, "") or short_eid(eid)
        yield depth, eid, label
        yield from walk_tree(tree, lbls, parent=eid, depth=depth + 1, visited=visited)


# ─── Excel ──────────────────────────────────────────────────────────────────
wb = Workbook()
wb.remove(wb.active)

hdr_font = Font(bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="1E3A5F")
abstract_fill = PatternFill("solid", fgColor="DBE7F4")
opening_fill = PatternFill("solid", fgColor="FFF4CE")
closing_fill = PatternFill("solid", fgColor="FFE0B2")
border = Border(left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"))

CAVEAT = {
    "00112332": (
        "⚠ 미래에셋 XBRL 신고 quirk: element 의미가 일반 회사와 반대. "
        "「자산인 보험계약」 (InsuranceContractsThatAreAssets) element 에 부채 BEL 잔액(+26.25조)이 저장됨. "
        "「부채인 보험계약」 (InsuranceContractsThatAreLiabilities) element 는 위험관리·상품군 partition 분해값들이 들어있어 element 단순합은 의미없음. "
        "사업보고서 PDF disclosure 의 「부채인 보험계약 기초/기말」 값은 본 시트의 "
        "「자산인 보험계약」 row(InsuranceContractsThatAreAssets) 를 참고할 것."
    ),
}

for cik, name in PEERS:
    ws = wb.create_sheet(title=name)
    ws.append([f"§4-1A 보험계약부채 변동표 (충실 재현) — {name}"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.append([f"role: {ROLE} (별도 변동표) · 필터: 별도 × 발행, 컴포넌트축·위험관리 sub-table 제외, 최소 axis depth fact 합산"])
    ws.cell(row=2, column=1).font = Font(italic=True, color="666666", size=10)
    if cik in CAVEAT:
        ws.append([CAVEAT[cik]])
        ws.cell(row=3, column=1).font = Font(bold=True, color="C62828", size=10)
        ws.cell(row=3, column=1).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=3, column=1).fill = PatternFill("solid", fgColor="FFEBEE")
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=6)
        ws.row_dimensions[3].height = 60
    ws.append([])

    ws.append(["#", "disclosure 라벨 (회사 표기 그대로)", "element_id", "기초", "변동", "기말"])
    hr = ws.max_row
    for c in range(1, 7):
        cell = ws.cell(row=hr, column=c)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    tree, lbls, vals = fetch_tree_and_vals(cik)
    # 값 있는 element만 (axis/abstract 노드 제거)
    visible = sorted({eid for (eid, pk) in vals.keys()})
    # 정렬: 자사 기초 절대값 큰 순 (기초 → 변동 → 기말)
    def sort_key(eid):
        op = vals.get((eid, "opening")) or 0
        du = vals.get((eid, "duration")) or 0
        cl = vals.get((eid, "closing")) or 0
        # period kind 우선순위
        kind = 0 if op != 0 else (1 if du != 0 else 2)
        magnitude = -abs(op or du or cl or 0)
        return (kind, magnitude)
    for eid in sorted(visible, key=sort_key):
        op = vals.get((eid, "opening"))
        du = vals.get((eid, "duration"))
        cl = vals.get((eid, "closing"))
        label = lbls.get(eid, "") or short_eid(eid)
        ws.append(["", label, short_eid(eid),
                   round(op, 0) if op is not None else None,
                   round(du, 0) if du is not None else None,
                   round(cl, 0) if cl is not None else None])
        r = ws.max_row
        for c in range(1, 7):
            ws.cell(row=r, column=c).border = border
        for c in range(4, 7):
            ws.cell(row=r, column=c).number_format = '#,##0;-#,##0;"–"'
        # 잔액 vs 변동 색
        if op is not None and du is None and cl is None:
            for c in range(1, 7): ws.cell(row=r, column=c).fill = opening_fill
        elif cl is not None and du is None and op is None:
            for c in range(1, 7): ws.cell(row=r, column=c).fill = closing_fill

    for i, w in enumerate([10, 50, 50, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = f"D{hr+1}"

# ─── 핵심 비교 시트 (P&L 명확 element만) ─────────────────────────────────────
core = wb.create_sheet(title="핵심 라인 비교", index=0)
core.append(["핵심 라인 동업사 비교 — 표준 ifrs-full element 의미 명확한 것만 (별도, FY2025, 억원)"])
core.cell(row=1, column=1).font = Font(bold=True, size=14)
core.append(["기초·기말 잔액 + 손익계산서 표시 항목 (보험수익·서비스결과). 변동 라인은 회사별 element 매핑 차이로 비교 보류"])
core.cell(row=2, column=1).font = Font(italic=True, color="666666", size=10)
core.append([])

CORE_LINES = [
    # (기간, 라벨, element_id)
    ("opening", "부채인 보험계약 기초",       "ifrs-full_InsuranceContractsThatAreLiabilities"),
    ("opening", "자산인 보험계약 기초",       "ifrs-full_InsuranceContractsThatAreAssets"),
    ("closing", "부채인 보험계약 기말",       "ifrs-full_InsuranceContractsThatAreLiabilities"),
    ("closing", "자산인 보험계약 기말",       "ifrs-full_InsuranceContractsThatAreAssets"),
    ("duration", "보험수익 (P&L)",            "ifrs-full_InsuranceRevenue"),
    ("duration", "보험서비스결과 변동 (그룹)", "ifrs-full_IncreaseDecreaseThroughInsuranceServiceResultInsuranceContractsLiabilityAsset"),
    ("duration", "보험계약마진 당기인식",     "ifrs-full_InsuranceRevenueContractualServiceMarginRecognisedInProfitOrLoss"),
    ("duration", "RA 변동 (별도 element)",   "ifrs-full_IncreaseDecreaseThroughChangeInRiskAdjustmentForNonfinancialRiskThatDoesNotRelateToFutureOrPastServiceInsuranceContractsLiabilityAsset"),
    ("duration", "CSM 미조정 추정변동",       "ifrs-full_IncreaseDecreaseThroughChangesInEstimatesThatDoNotAdjustContractualServiceMarginInsuranceContractsLiabilityAsset"),
    ("duration", "신계약 인식 효과",          "ifrs-full_IncreaseDecreaseThroughEffectsOfContractsInitiallyRecognisedInPeriodInsuranceContractsLiabilityAsset"),
    ("duration", "손실부담계약 손실/환입",    "ifrs-full_IncreaseDecreaseThroughEffectsOfGroupsOfOnerousContractsInitiallyRecognisedInPeriodInsuranceContractsLiabilityAsset"),
]

core.append(["기간", "라벨", "element_id"] + [n for _, n in PEERS])
hr2 = core.max_row
for c in range(1, 4 + len(PEERS)):
    cell = core.cell(row=hr2, column=c)
    cell.font = hdr_font; cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = border

# 8개사 vals 미리 로드
peer_vals = {}
for cik, _ in PEERS:
    peer_vals[cik] = {(eid, pk): amt
                      for eid, pk, amt in con.execute(SQL_VALUES, [cik, SEP, ISSUED, cik]).fetchall()}

PK_LABEL = {"opening": "기초", "duration": "변동", "closing": "기말"}
for pk, label, eid in CORE_LINES:
    row = [PK_LABEL[pk], label, eid.replace("ifrs-full_", "")[:55]]
    for cik, _ in PEERS:
        v = peer_vals[cik].get((eid, pk))
        row.append(round(v, 0) if v is not None else None)
    core.append(row)
    r = core.max_row
    for c in range(1, 4 + len(PEERS)):
        core.cell(row=r, column=c).border = border
        if pk == "opening": core.cell(row=r, column=c).fill = opening_fill
        elif pk == "closing": core.cell(row=r, column=c).fill = closing_fill
    for c in range(4, 4 + len(PEERS)):
        core.cell(row=r, column=c).number_format = '#,##0;-#,##0;"–"'

for i, w in enumerate([8, 32, 55] + [13]*len(PEERS), 1):
    core.column_dimensions[get_column_letter(i)].width = w
core.row_dimensions[hr2].height = 28
core.freeze_panes = f"D{hr2+1}"

out = Path("outputs/bel_faithful_per_peer.xlsx")
wb.save(out)
print(f"wrote {out}")
print(f"sheets: {wb.sheetnames}")
