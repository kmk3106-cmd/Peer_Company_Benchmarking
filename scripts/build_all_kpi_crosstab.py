"""8개사 KPI cross-tab Excel 종합 빌더.

기존 산출 JSON과 DB에서 직접 추출한 별도 기준 데이터를 통합하여
사업비/위험률 마진(§5-B), 이익잉여금 처분(§5-D), 추가 KPI(ROE·CSM 상각률·
보험서비스결과율 등)를 cross-tab 형식 다중 시트 Excel로 산출.

원칙
- 정확 element search (entity 확장만 사용). 추정·유추 금지.
- 별도(Separate) 기준만.
- 미공시는 "미공시"로 명시.
- 자사(미래에셋) 시그널: 1위/평균/최저/순위 자동 산출.
"""
from __future__ import annotations
import json
from pathlib import Path
import duckdb
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
DB = ROOT / "data" / "db" / "benchmark.duckdb"
OUT = ROOT / "outputs" / "all_kpi_crosstab.xlsx"
REPORT_DIR = ROOT / "report"

PEERS = [
    ("00112332", "미래에셋생명"),
    ("00126256", "삼성생명"),
    ("00113058", "한화생명"),
    ("00117267", "동양생명"),
    ("00139214", "삼성화재"),
    ("00164973", "현대해상"),
    ("00159102", "DB손해보험"),
    ("00135917", "한화손해보험"),
]
SELF_CIK = "00112332"
SEP = "ifrs-full_SeparateMember"

con = duckdb.connect(str(DB), read_only=True)

# ─────────────────────────── DB 헬퍼 ────────────────────────────
def get_separate_instant(cik: str, element_id: str, period: str = "2025-12-31") -> float | None:
    """별도 기준 axis만 갖는 (다른 dimension 없음) instant value."""
    q = """
    SELECT v.amount_krw FROM val_norm v JOIN cntxt_insurers c USING(CIK,REPORT_DATE,CONTEXT_ID)
    WHERE v.CIK=? AND v.ELEMENT_ID=? AND v.amount_krw IS NOT NULL
      AND c.MEMBER_ELEMENT_ID=?
      AND c.PERIOD_INSTANT=?
      AND NOT EXISTS(
        SELECT 1 FROM cntxt_insurers c2
        WHERE c2.CIK=v.CIK AND c2.REPORT_DATE=v.REPORT_DATE AND c2.CONTEXT_ID=v.CONTEXT_ID
          AND c2.MEMBER_ELEMENT_ID <> ?)
    ORDER BY v.amount_krw DESC LIMIT 1
    """
    r = con.execute(q, [cik, element_id, SEP, period, SEP]).fetchone()
    return r[0] if r else None


def get_separate_duration(cik: str, element_id: str, start: str = "2025-01-01") -> float | None:
    q = """
    SELECT v.amount_krw FROM val_norm v JOIN cntxt_insurers c USING(CIK,REPORT_DATE,CONTEXT_ID)
    WHERE v.CIK=? AND v.ELEMENT_ID=? AND v.amount_krw IS NOT NULL
      AND c.MEMBER_ELEMENT_ID=?
      AND c.PERIOD_START_DATE=?
      AND NOT EXISTS(
        SELECT 1 FROM cntxt_insurers c2
        WHERE c2.CIK=v.CIK AND c2.REPORT_DATE=v.REPORT_DATE AND c2.CONTEXT_ID=v.CONTEXT_ID
          AND c2.MEMBER_ELEMENT_ID <> ?)
    LIMIT 1
    """
    r = con.execute(q, [cik, element_id, SEP, start, SEP]).fetchone()
    return r[0] if r else None


# ──────────────────── 1) 기 산출 JSON 로드 ────────────────────
def load_json(name: str) -> dict:
    p = REPORT_DIR / name
    if not p.exists():
        return {}
    return json.loads(p.read_text(encoding="utf-8"))


PRICING = load_json("pricing_margins.json")
RESERVES = load_json("surrender_guarantee_reserves.json")
CSM_AMORT = load_json("csm_amortization.json")
CSM_MOVE = load_json("csm_movement.json")
RA_REL = load_json("ra_release.json")
OPEX = load_json("operating_expense_results.json")


# ──────────────── 2) 추가 BS·P/L 별도 기준 추출 ────────────────
def fetch_core_bs_pl(cik: str) -> dict:
    eq = get_separate_instant(cik, "ifrs-full_Equity")
    ret = get_separate_instant(cik, "ifrs-full_RetainedEarnings")
    ni = get_separate_duration(cik, "ifrs-full_ProfitLoss")
    rev = get_separate_duration(cik, "ifrs-full_InsuranceRevenue")
    res = get_separate_duration(cik, "ifrs-full_InsuranceServiceResult")
    return {
        "equity": eq, "retained": ret, "ni": ni,
        "insurance_revenue": rev, "insurance_service_result": res,
    }


def fetch_carry_forward(cik: str) -> dict:
    """전기이월 / 차기이월 미처분이익잉여금 (이익잉여금처분 공시)."""
    def _max(elm: str) -> float | None:
        q = """
        SELECT amount_krw FROM val_norm
        WHERE CIK=? AND ELEMENT_ID=? AND amount_krw IS NOT NULL
        ORDER BY amount_krw DESC LIMIT 1
        """
        r = con.execute(q, [cik, elm]).fetchone()
        return r[0] if r else None
    return {
        "prior_carry": _max("dart_UnappropriatedRetainedEarningsCarriedOverFromPriorYear"),
        "next_carry": _max("dart_UnappropriatedRetainedEarningsToBeCarriedForward"),
    }


CORE = {cik: fetch_core_bs_pl(cik) for cik, _ in PEERS}
CARRY = {cik: fetch_carry_forward(cik) for cik, _ in PEERS}


# ──────────── 3) Mirae 위험률 마진 잔여기간대별 분해 ────────────
# 자사 entity 확장 element + MaturityAxis member 별 합산.
MIRAE_ELEMS = {
    "risk_premium":   "entity00112332_RiskInsurancePremiumOfExpectedInsurancePayoutComparedToRiskPremiumOfExpectedInsurancePayoutComparedToRiskPremiumTableOfItems",
    "expected_claim": "entity00112332_ExpectedInsuranceAmountOfExpectedInsurancePayoutComparedToRiskPremiumOfExpectedInsurancePayoutComparedToRiskPremiumTableOfItems",
    "planned_maint":  "entity00112332_ScheduledMaintenanceCostOfExpectedMaintenanceCostComparedToPlannedMaintenanceCostEtcOfExpectedMaintenanceCostComparedToPlannedMaintenanceCostEtcTableOfItems",
    "expected_maint": "entity00112332_ExpectedMaintenanceCostsEtcOfExpectedMaintenanceCostComparedToPlannedMaintenanceCostEtcOfExpectedMaintenanceCostComparedToPlannedMaintenanceCostEtcTableOfItems",
}

MATURITY_BUCKETS = [
    ("entity00112332_Within10YearsOfAggregatedTimeBandsMemberOfOne63ExpectedRevenueRecognitionAmountByPeriodOfContractualServiceMarginTableOfMember", "10년 이내"),
    ("ifrs-full_LaterThanTenYearsAndNotLaterThanFifteenYearsMember", "10~15년"),
    ("ifrs-full_LaterThanFifteenYearsAndNotLaterThanTwentyYearsMember", "15~20년"),
    ("ifrs-full_LaterThanTwentyYearsAndNotLaterThanTwentyfiveYearsMember", "20~25년"),
    ("entity00112332_Over25YearsButWithin30YearsOfAggregatedTimeBandsMemberOfOne63ExpectedRevenueRecognitionAmountByPeriodOfContractualServiceMarginTableOfMember", "25~30년"),
    ("entity00112332_Over30YearsOfAggregatedTimeBandsMemberOfOne63ExpectedRevenueRecognitionAmountByPeriodOfContractualServiceMarginTableOfMember", "30년 초과"),
]


def fetch_mirae_by_maturity(element_id: str) -> dict[str, float]:
    """MaturityAxis member 별 합 (중복 axis 조합 제거)."""
    q = """
    WITH ax AS (
      SELECT CONTEXT_ID, STRING_AGG(MEMBER_ELEMENT_ID, '||' ORDER BY MEMBER_ELEMENT_ID) AS mems
      FROM cntxt_insurers WHERE CIK='00112332' GROUP BY 1
    ),
    matur AS (
      SELECT CONTEXT_ID, MEMBER_ELEMENT_ID FROM cntxt_insurers
      WHERE CIK='00112332' AND AXIS_ELEMENT_ID='ifrs-full_MaturityAxis'
    ),
    sep AS (SELECT DISTINCT CONTEXT_ID FROM cntxt_insurers
            WHERE CIK='00112332' AND MEMBER_ELEMENT_ID='ifrs-full_SeparateMember')
    SELECT m.MEMBER_ELEMENT_ID, SUM(amt) FROM (
      SELECT v.CONTEXT_ID, ax.mems, ANY_VALUE(v.amount_krw) AS amt
      FROM val_norm v JOIN ax USING(CONTEXT_ID) JOIN sep USING(CONTEXT_ID)
      WHERE v.CIK='00112332' AND v.ELEMENT_ID=? AND v.amount_krw IS NOT NULL
      GROUP BY v.CONTEXT_ID, ax.mems
    ) x JOIN matur m USING(CONTEXT_ID)
    GROUP BY 1
    """
    return {mem: val for mem, val in con.execute(q, [element_id]).fetchall()}


mirae_maturity = {k: fetch_mirae_by_maturity(e) for k, e in MIRAE_ELEMS.items()}


# ───────────────────── 4) Excel 작성 헬퍼 ─────────────────────
HDR_FILL = PatternFill("solid", fgColor="1F4E78")
HDR_FONT = Font(bold=True, color="FFFFFF", size=10)
SELF_FILL = PatternFill("solid", fgColor="FFF3CD")
NUM_FILL = PatternFill("solid", fgColor="F8F9FA")
BORDER = Border(*[Side(style="thin", color="CCCCCC")] * 4)
CENTER = Alignment(horizontal="center", vertical="center")
LEFT = Alignment(horizontal="left", vertical="center")
RIGHT = Alignment(horizontal="right", vertical="center")


def to_eok(v):
    if v is None:
        return "미공시"
    return v / 1e8


def fmt_pct(v):
    if v is None:
        return "미공시"
    return v


def add_signals(ws, start_row: int, header: list[str], values: list[list], peer_names: list[str]):
    """자사 시그널 행 추가: 1위·평균·최저·자사 순위."""
    # values: rows aligned with peer_names; columns aligned with header[1:]
    n_cols = len(header) - 1
    ws.append(["[자사 시그널]"] + [""] * n_cols)
    sig_row = ws.max_row
    ws.cell(row=sig_row, column=1).font = Font(bold=True)

    # For each numeric column, compute max/min/avg + Mirae rank
    for c in range(1, n_cols + 1):
        col_vals = [r[c - 1] for r in values if isinstance(r[c - 1], (int, float))]
        if not col_vals:
            continue
        mx, mn, avg = max(col_vals), min(col_vals), sum(col_vals) / len(col_vals)
        # find Mirae value
        mirae_idx = 0  # 미래에셋 is first peer
        mirae_val = values[mirae_idx][c - 1] if isinstance(values[mirae_idx][c - 1], (int, float)) else None
        # rank (descending)
        sorted_vals = sorted(col_vals, reverse=True)
        rank = sorted_vals.index(mirae_val) + 1 if mirae_val is not None and mirae_val in sorted_vals else None
        info = f"최고={mx:,.0f} / 평균={avg:,.0f} / 최저={mn:,.0f}"
        if rank:
            info += f" / 자사 {rank}위/{len(col_vals)}"
        ws.cell(row=sig_row + c - 1, column=1, value=info)


def write_sheet(wb: Workbook, name: str, header: list[str], rows: list[list], note: str | None = None,
                value_format: str | None = None):
    ws = wb.create_sheet(name)
    cur = 1
    if note:
        ws.cell(row=cur, column=1, value=note).font = Font(italic=True, color="555555", size=9)
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=len(header))
        cur += 2
    # header
    for ci, h in enumerate(header, 1):
        c = ws.cell(row=cur, column=ci, value=h)
        c.fill, c.font, c.alignment, c.border = HDR_FILL, HDR_FONT, CENTER, BORDER
    cur += 1
    # body
    for r in rows:
        is_self = "미래에셋" in str(r[0])
        for ci, val in enumerate(r, 1):
            c = ws.cell(row=cur, column=ci, value=val)
            c.border = BORDER
            if ci == 1:
                c.alignment = LEFT
            else:
                c.alignment = RIGHT
                if value_format and isinstance(val, (int, float)):
                    c.number_format = value_format
            if is_self:
                c.fill = SELF_FILL
                c.font = Font(bold=True)
        cur += 1
    # signals
    cur += 1
    ws.cell(row=cur, column=1, value="── 자사(미래에셋생명) 시그널 ──").font = Font(bold=True, color="1F4E78")
    cur += 1
    # compute per numeric column
    for ci in range(2, len(header) + 1):
        nums = [r[ci - 1] for r in rows if isinstance(r[ci - 1], (int, float))]
        if not nums:
            continue
        mirae_val = rows[0][ci - 1] if isinstance(rows[0][ci - 1], (int, float)) else None
        mx, mn, avg = max(nums), min(nums), sum(nums) / len(nums)
        sorted_desc = sorted(nums, reverse=True)
        rank = sorted_desc.index(mirae_val) + 1 if mirae_val is not None else None
        line = f"{header[ci-1]}: 최고 {mx:,.1f} / 평균 {avg:,.1f} / 최저 {mn:,.1f}"
        if rank and mirae_val is not None:
            line += f" | 자사 {mirae_val:,.1f} (순위 {rank}/{len(nums)})"
        ws.cell(row=cur, column=1, value=line)
        ws.merge_cells(start_row=cur, start_column=1, end_row=cur, end_column=len(header))
        cur += 1
    # column widths
    ws.column_dimensions["A"].width = 20
    for ci in range(2, len(header) + 1):
        ws.column_dimensions[get_column_letter(ci)].width = 16


# ───────────────────────── 시트 작성 ─────────────────────────
wb = Workbook()
wb.remove(wb.active)

# (0) 표지
ws0 = wb.create_sheet("표지", 0)
ws0["A1"] = "8개사 KPI Cross-Tab (FY2025 별도)"
ws0["A1"].font = Font(size=16, bold=True)
ws0["A3"] = "출처: 각사 별도 FY2025 사업보고서 / 보험계약 정보공시 XBRL"
ws0["A4"] = "원칙: entity 확장 element 정확 search. 미공시 = 데이터 없음."
ws0["A5"] = "단위: 별도 표시 없으면 억원."
ws0["A6"] = "자사 = 미래에셋생명 (CIK 00112332). 자사 행은 노란색 배경."
ws0["A8"] = "시트 목록:"
sheets = [
    ("§5-B-1. 위험률 마진", "잔여 예상보험금 ÷ 잔여 위험보험료. <100% 보수적."),
    ("§5-B-2. 사업비 마진", "잔여 예상유지비 ÷ 잔여 예정유지비. <100% 보수적."),
    ("§5-B-3. 자사 잔여기간대 분해", "미래에셋 MaturityAxis 6 버킷별 마진 (잔여기간대별)."),
    ("§5-C. 사업비 항목", "P/L 사업비·핵심사업비·사업비율."),
    ("§5-D-1. 이익잉여금·NI·ROE", "별도 기준 자본·이익잉여금·당기순이익·ROE."),
    ("§5-D-2. 미처분이익잉여금 흐름", "전기이월 → 차기이월 + 변동."),
    ("§5-D-3. 해약환급금·보증준비금", "기적립·적립예정·누계."),
    ("§Extra-1. 보험서비스결과율", "보험서비스결과 ÷ 보험수익."),
    ("§Extra-2. CSM 상각률", "당기상각 ÷ 평균CSM."),
    ("§Extra-3. RA 해소", "당기 RA 해소액 (보험수익 인식분)."),
    ("§Extra-4. 손익 지표 종합", "보험수익·서비스결과·NI·NI/Equity."),
]
for i, (n, d) in enumerate(sheets, 9):
    ws0.cell(row=i, column=1, value=n).font = Font(bold=True)
    ws0.cell(row=i, column=2, value=d)
ws0.column_dimensions["A"].width = 36
ws0.column_dimensions["B"].width = 70

# (1) §5-B-1 위험률 마진
hdr = ["회사", "위험보험료(억)", "예상보험금(억)", "예상/위험(%)", "마진 절대액(억)"]
rows = []
for cik, name in PEERS:
    p = PRICING.get(cik, {})
    rp = to_eok(p.get("risk_premium"))
    ep = to_eok(p.get("expected_claim"))
    pct = p.get("claim_ratio_pct")
    if isinstance(rp, float) and isinstance(ep, float):
        margin = rp - ep
    else:
        margin = "미공시"
    rows.append([name, rp, ep, pct if pct is not None else "미공시", margin])
write_sheet(wb, "5-B-1_위험률마진", hdr, rows,
            note="출처: 각사 entity 확장 (위험보험료·예상보험금). 잔여기간 합계, undiscounted. 비율 <100% = 보수적 가격책정.",
            value_format="#,##0.0")

# (2) §5-B-2 사업비 마진
hdr = ["회사", "예정유지비(억)", "예상유지비(억)", "예상/예정(%)", "마진 절대액(억)"]
rows = []
for cik, name in PEERS:
    p = PRICING.get(cik, {})
    pm = to_eok(p.get("planned_maint"))
    em = to_eok(p.get("expected_maint"))
    pct = p.get("maint_ratio_pct")
    if isinstance(pm, float) and isinstance(em, float):
        margin = pm - em
    else:
        margin = "미공시"
    rows.append([name, pm, em, pct if pct is not None else "미공시", margin])
write_sheet(wb, "5-B-2_사업비마진", hdr, rows,
            note="출처: 각사 entity 확장 (예정유지비·예상유지비). 잔여기간 합계. <100% = 사업비 보수적.",
            value_format="#,##0.0")

# (3) §5-B-3 자사 잔여기간대 분해
hdr = ["잔여기간 버킷", "위험보험료(억)", "예상보험금(억)", "예상/위험(%)",
       "예정유지비(억)", "예상유지비(억)", "예상/예정(%)"]
rows = []
for mem, label in MATURITY_BUCKETS:
    rp = mirae_maturity["risk_premium"].get(mem)
    ec = mirae_maturity["expected_claim"].get(mem)
    pm = mirae_maturity["planned_maint"].get(mem)
    em = mirae_maturity["expected_maint"].get(mem)
    rp_e = to_eok(rp); ec_e = to_eok(ec); pm_e = to_eok(pm); em_e = to_eok(em)
    cr = (ec / rp * 100) if rp and ec else "미공시"
    mr = (em / pm * 100) if pm and em else "미공시"
    rows.append([label, rp_e, ec_e, cr, pm_e, em_e, mr])
ws3 = wb.create_sheet("5-B-3_자사_잔여기간대분해")
ws3.cell(row=1, column=1, value="미래에셋생명 잔여기간 버킷별 마진 (FY2025 별도, 억원)").font = Font(bold=True, size=12)
ws3.cell(row=2, column=1, value="출처: entity00112332 위험보험료·예상보험금·예정유지비·예상유지비 × MaturityAxis").font = Font(italic=True, color="555555", size=9)
for ci, h in enumerate(hdr, 1):
    c = ws3.cell(row=4, column=ci, value=h)
    c.fill, c.font, c.alignment, c.border = HDR_FILL, HDR_FONT, CENTER, BORDER
for ri, r in enumerate(rows, 5):
    for ci, v in enumerate(r, 1):
        c = ws3.cell(row=ri, column=ci, value=v)
        c.border = BORDER
        if ci > 1 and isinstance(v, (int, float)):
            c.number_format = "#,##0.0"
            c.alignment = RIGHT
ws3.column_dimensions["A"].width = 18
for ci in range(2, len(hdr) + 1):
    ws3.column_dimensions[get_column_letter(ci)].width = 16

# (4) §5-C 사업비 항목
hdr = ["회사", "보험수익(억)", "판매비와관리비(억)", "보험영업비용(억)",
       "기타보험영업비용(억)", "핵심사업비(억)", "사업비율(%)"]
rows = []
for cik, name in PEERS:
    d = OPEX.get(cik, {})
    rows.append([
        name,
        to_eok(d.get("보험수익")),
        to_eok(d.get("판매비와관리비")),
        to_eok(d.get("보험영업비용")),
        to_eok(d.get("기타보험영업비용")),
        to_eok(d.get("핵심사업비")),
        d.get("사업비율") if d.get("사업비율") is not None else "미공시",
    ])
write_sheet(wb, "5-C_사업비항목", hdr, rows,
            note="출처: 각사 별도 P/L. 핵심사업비 = 판관비 + 보험영업비용(투자영업 제외). 사업비율 = 핵심사업비/보험수익.",
            value_format="#,##0.0")

# (5) §5-D-1 이익잉여금·NI·ROE
hdr = ["회사", "자본총계(억)", "이익잉여금(억)", "당기순이익(억)",
       "ROE(NI/자본,%)", "이익잉여금증가율(NI/이익잉여금,%)"]
rows = []
for cik, name in PEERS:
    d = CORE.get(cik, {})
    eq, ret, ni = d.get("equity"), d.get("retained"), d.get("ni")
    roe = (ni / eq * 100) if ni and eq else "미공시"
    ret_pct = (ni / ret * 100) if ni and ret else "미공시"
    rows.append([name, to_eok(eq), to_eok(ret), to_eok(ni), roe, ret_pct])
write_sheet(wb, "5-D-1_이익잉여금_NI_ROE", hdr, rows,
            note="출처: 각사 별도 BS·P/L (ifrs-full_Equity / RetainedEarnings / ProfitLoss). ROE = NI/평균자본 대신 NI/기말자본.",
            value_format="#,##0.0")

# (6) §5-D-2 미처분이익잉여금 흐름
hdr = ["회사", "전기이월(억)", "차기이월(억)", "변동(차기-전기,억)"]
rows = []
for cik, name in PEERS:
    d = CARRY.get(cik, {})
    pc, nc = d.get("prior_carry"), d.get("next_carry")
    diff = (nc - pc) if (pc is not None and nc is not None) else "미공시"
    rows.append([name, to_eok(pc), to_eok(nc), to_eok(diff) if isinstance(diff, (int, float)) else "미공시"])
write_sheet(wb, "5-D-2_미처분이익잉여금", hdr, rows,
            note="출처: 각사 이익잉여금처분 공시 (dart_UnappropriatedRetainedEarnings*). 음수 변동 = 적립금 전입·배당 우세.",
            value_format="#,##0.0")

# (7) §5-D-3 해약환급금·보증준비금
hdr = ["회사", "해약 기적립(억)", "해약 적립예정(억)", "해약 누계(억)",
       "보증 기적립(억)", "보증 적립예정(억)", "보증 누계(억)"]
rows = []
for cik, name in PEERS:
    r = RESERVES.get(cik, {})
    sb = r.get("sur_bal"); sa = r.get("sur_add"); gb = r.get("gua_bal"); ga = r.get("gua_add")
    sum_s = (sb + sa) if isinstance(sb, (int, float)) and isinstance(sa, (int, float)) else "미공시"
    sum_g = (gb + ga) if isinstance(gb, (int, float)) and isinstance(ga, (int, float)) else "미공시"
    rows.append([
        name,
        sb if sb is not None else "미공시",
        sa if sa is not None else "미공시",
        sum_s,
        gb if gb is not None else "미공시",
        ga if ga is not None else "미공시",
        sum_g,
    ])
write_sheet(wb, "5-D-3_해약_보증준비금", hdr, rows,
            note="출처: dart_SurrenderValueReserve(ToBeAdded) / dart_GuranteeReserve(ToBeAdded). 손보 보증준비금은 대부분 미공시.",
            value_format="#,##0.0")

# (8) §Extra-1 보험서비스결과율
hdr = ["회사", "보험수익(억)", "보험서비스결과(억)", "서비스결과율(%)"]
rows = []
for cik, name in PEERS:
    d = CORE.get(cik, {})
    rev, res = d.get("insurance_revenue"), d.get("insurance_service_result")
    pct = (res / rev * 100) if rev and res else "미공시"
    rows.append([name, to_eok(rev), to_eok(res), pct])
write_sheet(wb, "Extra-1_서비스결과율", hdr, rows,
            note="출처: ifrs-full_InsuranceRevenue / InsuranceServiceResult (별도). 결과율 = 보험본업 수익성.",
            value_format="#,##0.0")

# (9) §Extra-2 CSM 상각률
hdr = ["회사", "기시 CSM(억)", "기말 CSM(억)", "평균 CSM(억)", "당기상각(억)", "상각률(%)"]
rows = []
for cik, name in PEERS:
    d = CSM_AMORT.get(cik, {})
    beg = d.get("beg"); end = d.get("end"); avg = d.get("avg"); amort = d.get("amort"); rate = d.get("rate")
    rows.append([
        name,
        to_eok(beg), to_eok(end), to_eok(avg), to_eok(amort),
        rate if rate is not None else "미공시",
    ])
write_sheet(wb, "Extra-2_CSM상각률", hdr, rows,
            note="출처: 보험계약 CSM 변동표 (v2-clean: CurrentService × CSM-component). 상각률 = 당기상각/평균CSM.",
            value_format="#,##0.0")

# (10) §Extra-3 RA 해소
hdr = ["회사", "RA 해소(억)", "보험수익(억)", "RA해소/보험수익(%)", "출처"]
rows = []
for cik, name in PEERS:
    d = RA_REL.get(cik, {})
    ra = d.get("ra_release"); src = d.get("source", "미공시")
    rev = CORE.get(cik, {}).get("insurance_revenue")
    pct = (ra / rev * 100) if ra and rev else "미공시"
    rows.append([name, to_eok(ra), to_eok(rev), pct, src])
write_sheet(wb, "Extra-3_RA해소", hdr, rows,
            note="출처: 보험수익 중 위험조정 해소 기여분 (dart 표준 또는 entity 확장).",
            value_format="#,##0.0")

# (11) §Extra-4 손익 지표 종합
hdr = ["회사", "보험수익(억)", "서비스결과(억)", "당기순이익(억)", "자본총계(억)", "NI/자본(%)", "서비스결과/보험수익(%)"]
rows = []
for cik, name in PEERS:
    d = CORE.get(cik, {})
    rev, res, ni, eq = d.get("insurance_revenue"), d.get("insurance_service_result"), d.get("ni"), d.get("equity")
    roe = (ni / eq * 100) if ni and eq else "미공시"
    srate = (res / rev * 100) if rev and res else "미공시"
    rows.append([name, to_eok(rev), to_eok(res), to_eok(ni), to_eok(eq), roe, srate])
write_sheet(wb, "Extra-4_손익지표종합", hdr, rows,
            note="별도 기준 핵심 손익·자본 지표 종합표.",
            value_format="#,##0.0")

# 저장
OUT.parent.mkdir(parents=True, exist_ok=True)
wb.save(OUT)
print(f"saved: {OUT}")
print(f"sheets: {wb.sheetnames}")
