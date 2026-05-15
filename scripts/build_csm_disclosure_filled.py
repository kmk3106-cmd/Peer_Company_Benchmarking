"""CSM(보험계약마진) 변동 disclosure — BEL과 동일한 13행 구조로 8개사 cross-tab.

전략:
- BEL 스크립트와 동일한 13행 라벨·동일한 ifrs-full element 매핑 사용
- 단, context filter 에 ComponentsAxis=ContractualServiceMarginMember 를 추가 → CSM slice 만 추출
- CSM 표준 분해 (transition/non-transition) 도 합산 후보로 활용 (자기상쇄 방지 위해 합계 member 우선)
- Sep × Issued × CSM Component, 비-위험관리

CSM 의미적으로:
- 기초/기말: CSM 잔액 (InsuranceContractsLiabilityAsset 의 CSM slice = "부채인 보험계약" 라벨 그대로 사용하되 의미는 CSM 잔액)
- 보험수익 행: CSM 인식분 (CSM→PL transfer) 으로 자연 대응
- 현금흐름 (수취보험료/취득CF/보험금지급): CSM 에는 영향 없음 → 보통 0 또는 미공시
- 금융손익: CSM 누적이자 (InsuranceFinanceIncomeOrExpenses 의 CSM 부분)
- 추정변경: ChangesInEstimatesThatAdjustCSM
"""
from __future__ import annotations
import duckdb
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

PEERS = [
    ("00112332", "미래에셋생명"),
    ("00126256", "삼성생명"),
    ("00113058", "한화생명"),
    ("00117267", "동양생명"),
    ("00139214", "삼성화재"),
    ("00164973", "현대해상"),
    ("00159102", "DB손해보험"),
    ("00135917", "한화손해보험"),
]
SEP = "ifrs-full_SeparateMember"
ISSUED = "ifrs-full_InsuranceContractsIssuedMember"
COMP_AXIS = "ifrs-full_InsuranceContractsByComponentsAxis"
CSM_TOTAL = "ifrs-full_ContractualServiceMarginMember"
# CSM 세부 분해 멤버 (transition / non-transition / FV / MRA)
CSM_SUB_MEMBERS = [
    "ifrs-full_ContractualServiceMarginRelatedToContractsThatExistedAtTransitionDateToWhichModifiedRetrospectiveApproachHasBeenAppliedMember",
    "ifrs-full_ContractualServiceMarginRelatedToContractsThatExistedAtTransitionDateToWhichFairValueApproachHasBeenAppliedMember",
    "ifrs-full_ContractualServiceMarginNotRelatedToContractsThatExistedAtTransitionDateToWhichModifiedRetrospectiveApproachOrFairValueApproachHasBeenAppliedMember",
]
ROLE = "dart_2024-06-30_role-DI817105"

# 13행 (라벨, 기간) — BEL 과 동일
ROWS = [
    ("부채인 보험계약의 기초 장부금액(CSM)",        "opening"),
    ("보험수익",                                    "duration"),
    ("발생한 보험금 및 기타 보험서비스비용",        "duration"),
    ("보험취득현금흐름 상각",                       "duration"),
    ("발생사고요소의 조정",                         "duration"),
    ("손실부담계약 관련 손실(환입)",                "duration"),
    ("수취한 보험료",                               "duration"),
    ("보험취득현금흐름 지급",                       "duration"),
    ("보험금(투자요소 포함) 및 기타보험서비스비용의 지급", "duration"),
    ("당기손익인식 보험금융손익",                   "duration"),
    ("기타포괄손익인식 보험금융손익",               "duration"),
    ("기타증감",                                    "duration"),
    ("부채인 보험계약의 기말 장부금액(CSM)",        "closing"),
]

# BEL 와 동일한 element 매핑. CSM 행에서는 ComponentsAxis 필터로 CSM slice 만 추출됨
ELEMENT_PATTERNS = {
    "부채인 보험계약의 기초 장부금액(CSM)": {
        "period": "opening",
        "primary_eids": [
            "ifrs-full_InsuranceContractsLiabilityAsset",
            "ifrs-full_InsuranceContractsIssuedThatAreLiabilities",
            "ifrs-full_InsuranceContractsThatAreLiabilities",
        ],
        "fallback_eids": [],
    },
    "부채인 보험계약의 기말 장부금액(CSM)": {
        "period": "closing",
        "primary_eids": [
            "ifrs-full_InsuranceContractsLiabilityAsset",
            "ifrs-full_InsuranceContractsIssuedThatAreLiabilities",
            "ifrs-full_InsuranceContractsThatAreLiabilities",
        ],
        "fallback_eids": [],
    },
    "보험수익": {
        "period": "duration",
        # CSM 관점에서 "보험수익" 행 = CSM이 PL로 transfer 되어 인식된 분 (음수)
        "primary_eids": [
            "ifrs-full_IncreaseDecreaseThroughRecognitionOfContractualServiceMarginInProfitOrLossToReflectTransferOfServicesInsuranceContractsLiabilityAsset",
            "ifrs-full_IncreaseDecreaseThroughInsuranceRevenueInsuranceContractsLiabilityAsset",
        ],
        "fallback_eids": [],
    },
    "발생한 보험금 및 기타 보험서비스비용": {
        "period": "duration",
        "primary_eids": ["ifrs-full_IncreaseDecreaseThroughIncurredClaimsAndOtherIncurredInsuranceServiceExpensesInsuranceContractsLiabilityAsset"],
        "fallback_eids": ["ifrs-full_IncreaseDecreaseThroughIncurredClaimsAndOtherIncurredInsuranceServiceExpensesAdjustmentInsuranceContractsLiabilityAsset"],
    },
    "보험취득현금흐름 상각": {
        "period": "duration",
        "primary_eids": ["ifrs-full_IncreaseDecreaseThroughAmortisationOfInsuranceAcquisitionCashFlowsInsuranceContractsLiabilityAsset"],
        "fallback_eids": [],
    },
    "발생사고요소의 조정": {
        "period": "duration",
        "primary_eids": ["ifrs-full_IncreaseDecreaseThroughChangesThatRelateToPastServiceInsuranceContractsLiabilityAsset"],
        "fallback_eids": [],
    },
    "손실부담계약 관련 손실(환입)": {
        "period": "duration",
        "sum_mode": "sum_all",
        "primary_eids": [
            "ifrs-full_IncreaseDecreaseThroughEffectsOfGroupsOfOnerousContractsInitiallyRecognisedInPeriodInsuranceContractsLiabilityAsset",
            "ifrs-full_IncreaseDecreaseThroughChangesInEstimatesThatDoNotAdjustContractualServiceMarginInsuranceContractsLiabilityAsset",
        ],
        "fallback_eids": [],
    },
    "수취한 보험료": {
        "period": "duration",
        "primary_eids": ["ifrs-full_IncreaseDecreaseThroughPremiumsReceivedForInsuranceContractsIssuedInsuranceContractsLiabilityAsset"],
        "fallback_eids": [],
    },
    "보험취득현금흐름 지급": {
        "period": "duration",
        "primary_eids": ["ifrs-full_IncreaseDecreaseThroughInsuranceAcquisitionCashFlowsInsuranceContractsLiabilityAsset"],
        "fallback_eids": [],
    },
    "보험금(투자요소 포함) 및 기타보험서비스비용의 지급": {
        "period": "duration",
        "primary_eids": ["ifrs-full_IncreaseDecreaseThroughIncurredClaimsPaidAndOtherInsuranceServiceExpensesPaidForInsuranceContractsIssuedExcludingInsuranceAcquisitionCashFlowsInsuranceContractsLiabilityAsset"],
        "fallback_eids": [],
    },
    "당기손익인식 보험금융손익": {
        "period": "duration",
        "primary_eids": ["ifrs-full_IncreaseDecreaseThroughInsuranceFinanceIncomeOrExpensesInsuranceContractsLiabilityAsset"],
        "fallback_eids": ["ifrs-full_InsuranceFinanceIncomeExpensesFromInsuranceContractsIssuedRecognisedInProfitOrLoss"],
    },
    "기타포괄손익인식 보험금융손익": {
        "period": "duration",
        "primary_eids": ["ifrs-full_OtherComprehensiveIncomeBeforeTaxInsuranceFinanceIncomeExpensesFromInsuranceContractsIssuedExcludedFromProfitOrLossThatWillBeReclassifiedToProfitOrLoss"],
        "fallback_eids": ["ifrs-full_OtherComprehensiveIncomeBeforeTaxInsuranceFinanceIncomeExpensesFromInsuranceContractsIssuedExcludedFromProfitOrLoss"],
    },
    "기타증감": {
        "period": "duration",
        "primary_eids": [
            "ifrs-full_IncreaseDecreaseThroughAdditionalItemsNecessaryToUnderstandChangeInsuranceContractsLiabilityAsset",
        ],
        "fallback_eids": [
            "ifrs-full_IncreaseDecreaseThroughAdditionalItemsNecessaryToUnderstandChangeInNetCarryingAmountOfInsuranceContractsInsuranceContractsLiabilityAsset",
        ],
    },
}

con = duckdb.connect("data/db/benchmark.duckdb", read_only=True)

# CSM-filtered value 추출 SQL
# - ComponentsAxis 가 반드시 존재해야 함
# - ComponentsAxis 멤버 ∈ {CSM_TOTAL} 우선; 없으면 CSM_SUB_MEMBERS 합산 (n_extra 동일 그룹 내)
# - 그 외 axis (PAA, 상품, 측정모형 등) 는 extra_axis 로 카운트하여 최상위 우선
SQL_VAL_CSM = """
WITH ctx_ax AS (
  SELECT CONTEXT_ID,
         MAX(PERIOD_INSTANT) AS p_inst,
         MAX(PERIOD_START_DATE) AS p_start,
         MAX(PERIOD_END_DATE) AS p_end,
         -- extra axis: Sep, Disagg, Components 외
         COUNT(DISTINCT CASE
           WHEN AXIS_ELEMENT_ID IN (
             'ifrs-full_ConsolidatedAndSeparateFinancialStatementsAxis',
             'ifrs-full_DisaggregationOfInsuranceContractsAxis',
             'ifrs-full_InsuranceContractsByComponentsAxis') THEN NULL
           ELSE AXIS_ELEMENT_ID END) AS n_extra,
         MAX(CASE
           WHEN AXIS_ELEMENT_ID IN (
             'ifrs-full_ConsolidatedAndSeparateFinancialStatementsAxis',
             'ifrs-full_DisaggregationOfInsuranceContractsAxis',
             'ifrs-full_InsuranceContractsByComponentsAxis') THEN NULL
           ELSE AXIS_ELEMENT_ID END) AS extra_axis,
         -- ComponentsAxis 멤버 추출
         MAX(CASE WHEN AXIS_ELEMENT_ID='ifrs-full_InsuranceContractsByComponentsAxis'
                  THEN MEMBER_ELEMENT_ID END) AS comp_member
  FROM cntxt_insurers
  WHERE CIK=? AND REPORT_DATE='20251231'
  GROUP BY CONTEXT_ID
  HAVING BOOL_OR(MEMBER_ELEMENT_ID=?)      -- Sep
     AND BOOL_OR(MEMBER_ELEMENT_ID=?)      -- Issued
     AND BOOL_OR(AXIS_ELEMENT_ID='ifrs-full_InsuranceContractsByComponentsAxis')
     AND NOT BOOL_OR(MEMBER_ELEMENT_ID LIKE '%OfDisclosureOfNatureAndExtentOfRisks%')
),
fact AS (
  SELECT v.CONTEXT_ID, c.n_extra, c.extra_axis, c.comp_member,
         CASE WHEN c.p_inst='2024-12-31' THEN 'opening'
              WHEN c.p_inst='2025-12-31' THEN 'closing'
              WHEN c.p_start='2025-01-01' AND c.p_end='2025-12-31' THEN 'duration'
              ELSE NULL END AS pk,
         MAX(v.amount_krw) AS amount_krw
  FROM val_insurers v JOIN ctx_ax c USING(CONTEXT_ID)
  WHERE v.CIK=? AND v.REPORT_DATE='20251231' AND v.ELEMENT_ID=? AND v.amount_krw IS NOT NULL
  GROUP BY v.CONTEXT_ID, c.n_extra, c.extra_axis, c.comp_member, c.p_inst, c.p_start, c.p_end
),
-- 우선순위 a: comp_member = CSM_TOTAL, n_extra=0
a0 AS (SELECT SUM(amount_krw) AS amt FROM fact WHERE pk=? AND n_extra=0 AND comp_member=?),
-- 우선순위 b: comp_member ∈ CSM 세부 분해 멤버 (sub), n_extra=0 → 합산
b0 AS (SELECT SUM(amount_krw) AS amt FROM fact WHERE pk=? AND n_extra=0 AND comp_member = ANY(?)),
-- 우선순위 c: comp_member = CSM_TOTAL, n_extra=1 (단일 추가 axis 의 합)
a1 AS (
  SELECT SUM(amount_krw) AS amt, extra_axis
  FROM fact WHERE pk=? AND n_extra=1 AND comp_member=?
  GROUP BY extra_axis
  ORDER BY ABS(SUM(amount_krw)) DESC LIMIT 1
),
-- 우선순위 d: CSM sub members, n_extra=1
b1 AS (
  SELECT SUM(amount_krw) AS amt, extra_axis
  FROM fact WHERE pk=? AND n_extra=1 AND comp_member = ANY(?)
  GROUP BY extra_axis
  ORDER BY ABS(SUM(amount_krw)) DESC LIMIT 1
)
SELECT amt FROM (
  SELECT amt, 0 AS pri FROM a0 WHERE amt IS NOT NULL
  UNION ALL SELECT amt, 1 FROM b0 WHERE amt IS NOT NULL
  UNION ALL SELECT amt, 2 FROM a1 WHERE amt IS NOT NULL
  UNION ALL SELECT amt, 3 FROM b1 WHERE amt IS NOT NULL
) ORDER BY pri LIMIT 1
"""

def get_csm_value(cik, eid, pk):
    r = con.execute(SQL_VAL_CSM, [
        cik, SEP, ISSUED,           # ctx_ax filters
        cik, eid,                   # fact filter
        pk, CSM_TOTAL,              # a0
        pk, CSM_SUB_MEMBERS,        # b0
        pk, CSM_TOTAL,              # a1
        pk, CSM_SUB_MEMBERS,        # b1
    ]).fetchone()
    return r[0] if r and r[0] is not None else None


peer_values = {}      # (cik, label) -> value
peer_eids_used = {}   # (cik, label) -> list of eids actually used

for cik, name in PEERS:
    for label, pk in ROWS:
        key = (cik, label)
        spec = ELEMENT_PATTERNS[label]
        period = spec["period"]
        total = None
        used = []
        sum_mode = spec.get("sum_mode", "first_nonnull")
        if sum_mode == "sum_all":
            for eid in spec["primary_eids"]:
                v = get_csm_value(cik, eid, period)
                if v is not None:
                    total = (total or 0) + v
                    used.append(eid.replace("ifrs-full_", ""))
        else:
            for eid in spec["primary_eids"]:
                v = get_csm_value(cik, eid, period)
                if v is not None:
                    total = v
                    used.append(eid.replace("ifrs-full_", ""))
                    break
        if total is None:
            for eid in spec.get("fallback_eids", []):
                v = get_csm_value(cik, eid, period)
                if v is not None:
                    total = v
                    used.append(eid.replace("ifrs-full_", "") + " (fallback)")
                    break
        peer_values[key] = total
        peer_eids_used[key] = used if used else ["(미공시)"]

def residual(cik):
    op = peer_values.get((cik, "부채인 보험계약의 기초 장부금액(CSM)")) or 0
    cl = peer_values.get((cik, "부채인 보험계약의 기말 장부금액(CSM)")) or 0
    dur = 0
    for label, pk in ROWS:
        if pk == "duration":
            v = peer_values.get((cik, label))
            if v is not None: dur += v
    return op + dur - cl

# === Excel 출력 ===
wb = Workbook()
ws = wb.active
ws.title = "CSM 변동 disclosure"

ws.append(["§4-1B 보험계약마진(CSM) 변동표 — 동업사 cross-tab (별도, 발행, CSM Component, 억원)"])
ws.cell(row=1, column=1).font = Font(bold=True, size=14)
ws.append(["기준: 보험계약부채 변동표 13행 구조를 ComponentsAxis=CSM slice 로 재현"])
ws.cell(row=2, column=1).font = Font(italic=True, color="666666", size=10)
ws.append(["8개사 모두 XBRL DI817105 ifrs-full 표준 element 매칭 (Sep×Issued×CSM Component, top-level, 비-위험관리). 미래에셋도 동일 추출."])
ws.cell(row=3, column=1).font = Font(italic=True, color="C62828", size=10)
ws.append([])

hdrs = ["#", "disclosure 행 라벨", "기간"] + [n for _, n in PEERS]
ws.append(hdrs)
hr = ws.max_row
for c in range(1, len(hdrs)+1):
    cell = ws.cell(row=hr, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1E3A5F")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

opening_fill = PatternFill("solid", fgColor="FFF4CE")
closing_fill = PatternFill("solid", fgColor="FFE0B2")
revenue_fill = PatternFill("solid", fgColor="E3F2FD")
expense_fill = PatternFill("solid", fgColor="FFF3E0")
cf_fill      = PatternFill("solid", fgColor="E0F7FA")
fin_fill     = PatternFill("solid", fgColor="F3E5F5")
other_fill   = PatternFill("solid", fgColor="FAFAFA")
border = Border(left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"))

def fill_for(label, pk):
    if pk == "opening": return opening_fill
    if pk == "closing": return closing_fill
    if "보험수익" in label: return revenue_fill
    if any(w in label for w in ["발생한 보험금", "보험취득현금흐름 상각", "발생사고요소", "손실부담"]): return expense_fill
    if any(w in label for w in ["수취한", "지급", "투자요소"]): return cf_fill
    if "보험금융손익" in label: return fin_fill
    return other_fill

PK_LABEL = {"opening": "기초", "duration": "변동", "closing": "기말"}
for i, (label, pk) in enumerate(ROWS, 1):
    row = [i, label, PK_LABEL[pk]]
    for cik, _ in PEERS:
        v = peer_values.get((cik, label))
        row.append(round(v / 1e8, 0) if v is not None else None)
    ws.append(row)
    r = ws.max_row
    f = fill_for(label, pk)
    for c in range(1, len(hdrs)+1):
        ws.cell(row=r, column=c).fill = f
        ws.cell(row=r, column=c).border = border
    for c in range(4, len(hdrs)+1):
        cell = ws.cell(row=r, column=c)
        if cell.value is None:
            cell.value = "미공시"
            cell.font = Font(italic=True, color="999999", size=9)
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.number_format = '#,##0;-#,##0;"–"'

# 잔차 행
ws.append([])
ws.append(["검증", "잔차 (기초 + Σ변동 − 기말, 억원)", ""] + [""]*len(PEERS))
sr = ws.max_row
for i, (cik, _) in enumerate(PEERS, start=4):
    res = residual(cik)
    ws.cell(row=sr, column=i).value = round(res / 1e8, 0)
    ws.cell(row=sr, column=i).number_format = '#,##0;-#,##0;"– (정합)"'
    op = peer_values.get((cik, "부채인 보험계약의 기초 장부금액(CSM)")) or 0
    if op and abs(res) < max(abs(op) * 0.01, 100_000_000_000):
        ws.cell(row=sr, column=i).font = Font(bold=True, color="2E7D32")
    else:
        ws.cell(row=sr, column=i).font = Font(bold=True, color="C62828")
for c in range(1, len(hdrs)+1):
    ws.cell(row=sr, column=c).fill = PatternFill("solid", fgColor="E8F5E9")
    ws.cell(row=sr, column=c).border = border
ws.cell(row=sr, column=1).font = Font(bold=True)
ws.cell(row=sr, column=2).font = Font(bold=True)

# Caveat
ws.append([])
cr = ws.max_row + 1
ws.cell(row=cr, column=1).value = "주석"
ws.cell(row=cr, column=1).font = Font(bold=True, color="C62828")
ws.cell(row=cr, column=2).value = (
    "CSM slice 추출: ComponentsAxis=ContractualServiceMarginMember 컨텍스트 우선 사용, "
    "없을 시 transition·non-transition·MRA·FV 세부 분해 멤버 합산 fallback. "
    "BEL 13행 구조 그대로 적용했으나 CSM 의 회계적 특성상 다음 행들은 의미적으로 0 또는 미공시가 정상: "
    "수취 보험료·보험취득현금흐름 지급·보험금 지급 (CSM 은 현금흐름으로 직접 변동하지 않음), "
    "발생한 보험금·보험취득현금흐름 상각 (보험서비스비용 발생은 CSM 항목 아님). "
    "'보험수익' 행은 CSM→PL 인식분 (RecognitionOfContractualServiceMarginInProfitOrLoss) 으로 매핑됨. "
    "회사별 CSM disclosure 분해단위가 매우 다양해 미공시 행 다수 발생 — caveat 注意. "
    "잔차 |%| < 1% 면 정합 OK 로 표시."
)
ws.cell(row=cr, column=2).alignment = Alignment(wrap_text=True, vertical="top")
ws.cell(row=cr, column=2).font = Font(italic=True, color="666666", size=9)
ws.merge_cells(start_row=cr, start_column=2, end_row=cr, end_column=len(hdrs))
ws.row_dimensions[cr].height = 75

widths = [4, 40, 6] + [13]*len(PEERS)
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[hr].height = 28
ws.freeze_panes = f"D{hr+1}"

# 진단 시트
diag = wb.create_sheet(title="element 매핑")
diag.append(["회사별 disclosure 행 → 사용된 element_id (CSM Component slice)"])
diag.cell(row=1, column=1).font = Font(bold=True, size=12)
diag.append(["미공시 = primary·fallback 모두 매칭 안됨 (회사가 해당 항목을 CSM Component 컨텍스트로 보고하지 않음)"])
diag.cell(row=2, column=1).font = Font(italic=True, color="666666", size=9)
diag.append([])
diag.append(["disclosure 행"] + [n for _, n in PEERS])
hr2 = diag.max_row
for c in range(1, 2 + len(PEERS)):
    cell = diag.cell(row=hr2, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1E3A5F")
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
for label, pk in ROWS:
    row = [label]
    for cik, _ in PEERS:
        eids = peer_eids_used.get((cik, label), [])
        row.append("\n".join(eids))
    diag.append(row)
    r = diag.max_row
    for c in range(1, 2 + len(PEERS)):
        diag.cell(row=r, column=c).border = border
        diag.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)
for i, w in enumerate([35] + [40]*len(PEERS), 1):
    diag.column_dimensions[get_column_letter(i)].width = w
diag.freeze_panes = f"B{hr2+1}"

# 매핑 규칙 시트
rule = wb.create_sheet(title="매핑 규칙")
rule.append(["13행 라벨 → 캐노니컬 ifrs-full element_id (primary / fallback) — CSM Component slice"])
rule.cell(row=1, column=1).font = Font(bold=True, size=12)
rule.append(["context filter: Sep × Issued × ComponentsAxis (ContractualServiceMarginMember 또는 세부분해 멤버 합)"])
rule.cell(row=2, column=1).font = Font(italic=True, color="666666", size=9)
rule.append([])
rule.append(["#", "disclosure 행 라벨", "기간", "primary element_ids", "fallback element_ids"])
hr3 = rule.max_row
for c in range(1, 6):
    cell = rule.cell(row=hr3, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1E3A5F")
for i, (label, pk) in enumerate(ROWS, 1):
    spec = ELEMENT_PATTERNS.get(label, {})
    prim = "\n".join(e.replace("ifrs-full_", "") for e in spec.get("primary_eids", []))
    fb = "\n".join(e.replace("ifrs-full_", "") for e in spec.get("fallback_eids", []))
    rule.append([i, label, PK_LABEL[pk], prim, fb])
    r = rule.max_row
    rule.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="top")
    rule.cell(row=r, column=5).alignment = Alignment(wrap_text=True, vertical="top")
for i, w in enumerate([4, 38, 6, 70, 55], 1):
    rule.column_dimensions[get_column_letter(i)].width = w

out = Path("outputs/csm_disclosure_filled.xlsx")
wb.save(out)
print(f"wrote {out}")

import sys
try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass
print("\n=== 회사별 CSM 잔차 (기초 + Sum변동 - 기말, 억원 / 기초 대비 %) ===")
for cik, name in PEERS:
    res = residual(cik)
    op = peer_values.get((cik, "부채인 보험계약의 기초 장부금액(CSM)")) or 0
    pct = res / op * 100 if op else 0
    flag = "OK" if op and abs(res) < max(abs(op) * 0.01, 100_000_000_000) else "잔차큼"
    print(f"  {name:<12s} 기초={op/1e8:>+12,.0f}억  잔차={res/1e8:>+10,.0f}억 ({pct:+.2f}%)  [{flag}]")

print("\n=== 회사별 미공시 행 (CSM) ===")
for cik, name in PEERS:
    miss = [label for label, pk in ROWS
            if peer_eids_used.get((cik, label), [""])[0] == "(미공시)"]
    if miss:
        print(f"  {name}:")
        for m in miss:
            print(f"    - {m}")
