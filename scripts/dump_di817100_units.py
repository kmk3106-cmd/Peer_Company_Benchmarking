"""DI817100 — 11개 sub-table 단위별 시트 분리 dump (미래에셋생명).

각 [개요] abstract 가 한 sub-table 단위. 그 sub-tree 전체를 시트 하나에.

11 단위:
  1. 회계모형별·포트폴리오별 보험부채 현황
  2. 직접참가 파생 상쇄효과
  3. CSM 기간별 기대수익
  4. 회사 투자서비스수익·보험금융손익
  5. 직접참가 기초항목
  6. 보험수익
  7. 계리적가정 보험부채 변동
  8. 구성요소×배당여부 차이조정 (자사 확장)
  9. LRC/LIC×배당여부 차이조정 (자사 확장)
 10. 구성요소별 차이조정 (표준 dart)
 11. LRC/LIC 차이조정 (표준 dart)

출력: report/di817100_units.xlsx
"""
from __future__ import annotations

from pathlib import Path

import duckdb
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

CIK = "00112332"
COMPANY = "미래에셋생명"
ROLE = "dart_2024-06-30_role-DI817100"

UNITS = [
    # (key, label, root_element_id)
    ("U01_회계모형포트폴리오",
     "1. 회계모형별·포트폴리오별 보험부채 현황",
     "entity00112332_Title20256517565149Abstract"),
    ("U02_파생상쇄효과",
     "2. 직접참가 파생 상쇄효과",
     "entity00112332_Title20257289204290Abstract"),
    ("U03_CSM만기수익",
     "3. CSM 기간별 기대수익 인식금액",
     "entity00112332_Title202516161119540Abstract"),
    ("U04_투자수익_보험금융손익",
     "4. 회사 투자서비스수익·보험금융손익",
     "entity00112332_Title20251616323362Abstract"),
    ("U05_직접참가_기초항목",
     "5. 직접참가 기초항목 공정가치",
     "entity00112332_Title202516171838316Abstract"),
    ("U06_보험수익",
     "6. 보험수익 분해",
     "entity00112332_Title202516172856136Abstract"),
    ("U07_계리적가정_보험부채변동",
     "7. 계리적가정에 의한 보험부채 변동내역",
     "entity00112332_ChangesInInsuranceLiabilitiesBasedOnActuarialAssumptionsAbstract"),
    ("U08_구성요소_배당여부_자사",
     "8. 구성요소(BEL/RA/CSM) × 배당여부 — 자사 확장",
     "entity00112332_DisclosureOfReconciliationOfChangesInInsuranceContractsIssuedByComponentsClassificationByDividendStatusAbstract"),
    ("U09_LRC_LIC_배당여부_자사",
     "9. LRC/LIC × 배당여부 — 자사 확장",
     "entity00112332_DisclosureOfReconciliationOfChangesInInsuranceContractsIssuedByRemainingCoverageAndIncurredClaimsClassificationByDividendStatusAbstract"),
    ("U10_구성요소_표준",
     "10. 구성요소(BEL/RA/CSM) 차이조정 — 표준 dart",
     "dart_DisclosureOfReconciliationOfChangesInInsuranceContractsIssuedByComponentsAbstract"),
    ("U11_LRC_LIC_표준",
     "11. LRC/LIC 차이조정 — 표준 dart",
     "dart_DisclosureOfReconciliationOfChangesInInsuranceContractsIssuedByRemainingCoverageAndIncurredClaimsAbstract"),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(italic=True, size=9, color="666666")
ABSTRACT_FILL = PatternFill("solid", fgColor="EFEFEF")
AXIS_FILL = PatternFill("solid", fgColor="DCE6F1")
TABLE_FILL = PatternFill("solid", fgColor="FCE4D6")
ENTITY_FILL = PatternFill("solid", fgColor="FFF4E0")
HAS_VAL_FILL = PatternFill("solid", fgColor="E2EFDA")
THIN = Side(border_style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _short(eid: str) -> str:
    if not eid:
        return ""
    return (
        eid.replace("ifrs-full_", "")
        .replace("dart-gcd_", "")
        .replace("dart_2024-06-30_", "")
        .replace("dart_", "d:")
        .replace(f"entity{CIK}_", "[자사]")
    )


def fetch_unit_tree(con: duckdb.DuckDBPyConnection, root_elem: str) -> pd.DataFrame:
    """단위 root에서 출발해 cycle-safe 재귀 BFS로 sub-tree 추출 + meta + 값."""
    sql = """
    WITH RECURSIVE walk(elem, parent, ord, depth, path) AS (
      SELECT p.ELEMENT_ID, p.PARENT_ELEMENT_ID, TRY_CAST(p."ORDER" AS DOUBLE), 0,
             [CAST(p.ELEMENT_ID AS VARCHAR)]
      FROM pre_insurers p
      WHERE p.CIK=? AND p.ROLE_ID=? AND p.ELEMENT_ID=?
      UNION ALL
      SELECT p.ELEMENT_ID, p.PARENT_ELEMENT_ID, TRY_CAST(p."ORDER" AS DOUBLE),
             w.depth+1, list_append(w.path, p.ELEMENT_ID)
      FROM pre_insurers p
      JOIN walk w ON p.PARENT_ELEMENT_ID = w.elem
      WHERE p.CIK=? AND p.ROLE_ID=? AND w.depth < 8
        AND NOT list_contains(w.path, p.ELEMENT_ID)
    ),
    nodes AS (
      SELECT elem, parent, ord, MIN(depth) AS depth
      FROM walk GROUP BY elem, parent, ord
    ),
    labels AS (
      SELECT ELMT_ID, MAX(LABEL) AS ko_label FROM lab_insurers
      WHERE CIK=? AND LANG='ko'
        AND LABEL_ROLE_URI='http://www.xbrl.org/2003/role/label'
      GROUP BY ELMT_ID
    ),
    val_summary AS (
      SELECT ELEMENT_ID,
             COUNT(DISTINCT CONTEXT_ID) AS n_ctx,
             MAX(ABS(amount_krw)) AS max_abs,
             MIN(amount_krw) AS min_signed,
             MAX(amount_krw) AS max_signed
      FROM val_insurers
      WHERE CIK=? AND amount_krw IS NOT NULL
      GROUP BY ELEMENT_ID
    )
    SELECT
      n.elem AS ELEMENT_ID, n.parent AS PARENT_ELEMENT_ID, n.ord, n.depth,
      em.ABSTRACT, em.PERIOD_TYPE, em.BALANCE, em.SUBSTITUTION_GROUP,
      l.ko_label,
      vs.n_ctx, vs.max_abs, vs.min_signed, vs.max_signed
    FROM nodes n
    LEFT JOIN elmt_raw em ON em.ELEMENT_ID = n.elem
    LEFT JOIN labels l ON l.ELMT_ID = n.elem
    LEFT JOIN val_summary vs ON vs.ELEMENT_ID = n.elem
    """
    return con.execute(sql, [CIK, ROLE, root_elem, CIK, ROLE, CIK, CIK]).df()


def order_dfs(df: pd.DataFrame, root: str) -> pd.DataFrame:
    """depth-first 순서로 정렬 (parent→ord 사용)."""
    if df.empty:
        return df
    children: dict[str, list] = {}
    for _, row in df.iterrows():
        children.setdefault(row["PARENT_ELEMENT_ID"], []).append(
            (row["ord"] if pd.notna(row["ord"]) else 0, row["ELEMENT_ID"])
        )
    for k in children:
        children[k].sort()

    visited = set()
    ordered = []

    def walk(parent, depth):
        for _, eid in children.get(parent, []):
            if eid in visited:
                continue
            visited.add(eid)
            ordered.append((depth, eid))
            walk(eid, depth + 1)

    # Start from root and parent=None (covers all roots)
    visited.add(root)
    ordered.append((0, root))
    walk(root, 1)
    # Sweep orphans
    for _, row in df.iterrows():
        if row["ELEMENT_ID"] not in visited:
            visited.add(row["ELEMENT_ID"])
            ordered.append((1, row["ELEMENT_ID"]))
            walk(row["ELEMENT_ID"], 2)

    order_df = pd.DataFrame(ordered, columns=["dfs_depth", "ELEMENT_ID"])
    order_df["seq"] = range(len(order_df))
    return order_df.merge(df.drop(columns=["depth"], errors="ignore"),
                          on="ELEMENT_ID", how="left").sort_values("seq").reset_index(drop=True)


def classify(row) -> str:
    sub = (row.get("SUBSTITUTION_GROUP") or "")
    if "hypercubeItem" in sub:
        return "Table"
    if "dimensionItem" in sub:
        return "Axis"
    if str(row.get("ABSTRACT")).lower() == "true":
        return "Abstract"
    return "LineItem"


def write_unit_sheet(ws, label: str, root: str, df: pd.DataFrame) -> None:
    ws.cell(row=1, column=1, value=label).font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"{COMPANY} (CIK {CIK}) · root: {_short(root)} · {len(df)} elements").font = SUBTITLE_FONT

    headers = ["#", "depth", "element", "ko_label", "분류", "period", "bal", "ctx#", "min(억)", "max(억)"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER

    for ri, row in df.iterrows():
        depth = int(row.get("dfs_depth") or 0)
        eid = row["ELEMENT_ID"]
        is_entity = isinstance(eid, str) and eid.startswith("entity")
        cls = classify(row)
        indent = "  " * depth
        cells = [
            ri + 1,
            depth,
            indent + _short(eid),
            row.get("ko_label") or "",
            cls,
            row.get("PERIOD_TYPE") or "",
            row.get("BALANCE") or "",
            int(row["n_ctx"]) if pd.notna(row.get("n_ctx")) else 0,
            round(row["min_signed"] / 1e8, 1) if pd.notna(row.get("min_signed")) else None,
            round(row["max_signed"] / 1e8, 1) if pd.notna(row.get("max_signed")) else None,
        ]
        fill = {
            "Table": TABLE_FILL,
            "Axis": AXIS_FILL,
            "Abstract": ABSTRACT_FILL,
        }.get(cls) or (HAS_VAL_FILL if cells[7] > 0 else None)

        for ci, v in enumerate(cells, start=1):
            c = ws.cell(row=5 + ri, column=ci, value=v)
            c.border = BORDER
            if is_entity and ci == 3:
                c.fill = ENTITY_FILL
            elif fill:
                c.fill = fill
            if ci >= 8 and isinstance(v, (int, float)) and not isinstance(v, bool):
                c.alignment = Alignment(horizontal="right")
                if isinstance(v, float):
                    c.number_format = "#,##0.0;(#,##0.0);-"

    widths = [5, 6, 75, 42, 10, 10, 6, 6, 12, 12]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "C5"


def write_toc(ws, unit_dfs: dict[str, pd.DataFrame]) -> None:
    ws.cell(row=1, column=1, value=f"{COMPANY} DI817100 — 11개 sub-table 단위 목차").font = TITLE_FONT
    ws.cell(row=2, column=1, value="[개요] abstract 헤더별로 단위 분리. 각 시트에서 자세히 보기.").font = SUBTITLE_FONT

    headers = ["#", "단위", "전체 element", "값 보유", "Axis", "Abstract", "LineItem", "Table"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER

    for i, (key, label, root) in enumerate(UNITS, start=1):
        df = unit_dfs[key]
        n_total = len(df)
        if n_total == 0:
            n_val = n_axis = n_abs = n_li = n_tab = 0
        else:
            sg = df["SUBSTITUTION_GROUP"].fillna("")
            n_tab = sg.str.contains("hypercubeItem").sum()
            n_axis = sg.str.contains("dimensionItem").sum()
            is_abs = df["ABSTRACT"].astype(str).str.lower() == "true"
            n_abs = is_abs.sum() - n_axis - n_tab
            n_li = n_total - is_abs.sum()
            n_val = (df["n_ctx"].fillna(0) > 0).sum()
        cells = [i, label, n_total, int(n_val), int(n_axis), int(n_abs), int(n_li), int(n_tab)]
        for ci, v in enumerate(cells, start=1):
            c = ws.cell(row=4 + i, column=ci, value=v)
            c.border = BORDER
            if ci >= 3:
                c.alignment = Alignment(horizontal="right")
            if ci == 2:
                c.font = Font(bold=True)

    widths = [5, 50, 14, 12, 10, 12, 12, 8]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def main() -> int:
    con = duckdb.connect("data/db/benchmark.duckdb", read_only=True)

    unit_dfs: dict[str, pd.DataFrame] = {}
    for key, label, root in UNITS:
        df = fetch_unit_tree(con, root)
        df = order_dfs(df, root)
        unit_dfs[key] = df
        print(f"  {key:30s} → {len(df)} elements")

    con.close()

    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    ws_toc = wb.create_sheet("목차")
    write_toc(ws_toc, unit_dfs)

    for key, label, root in UNITS:
        sheet_name = key[:31]  # Excel 31 char limit
        ws = wb.create_sheet(sheet_name)
        write_unit_sheet(ws, label, root, unit_dfs[key])

    out = Path("report/di817100_units.xlsx")
    wb.save(out)
    print(f"\n✓ wrote {out} ({out.stat().st_size / 1024:.1f} KB, {len(UNITS) + 1} sheets)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
