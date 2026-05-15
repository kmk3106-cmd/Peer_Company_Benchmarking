"""§4-1A BEL 변동 — depth-aligned (정규화) 표.

원칙:
  1. axis depth 합산: 회사가 sub-axis(상품군·LRC/LIC·기타)로 쪼개 보고해도
     같은 element 라면 모두 SUM → 한 라인 한 값.
  2. axis 이름 무시: Sep × Issued × BEL component 셋만 조건. 그 외 추가 sub-axis 무엇이든
     facts를 element 단위로 합산.
  3. element는 raw 그대로 보존 (자사 entity-extension은 별도 표시).
"""
from __future__ import annotations
import duckdb
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

PEERS = [
    ("00112332", "미래에셋생명"),
    ("00126256", "삼성생명"),
    ("00113058", "한화생명"),
    ("00117267", "동양생명"),
    ("00139214", "삼성화재"),
    ("00164973", "현대해상"),
    ("00159102", "DB손해보험"),
    ("00135917", "한화손해보험"),
]
SEP = "ifrs-full_SeparateMember"
ISSUED = "ifrs-full_InsuranceContractsIssuedMember"
COMP_AXIS = "ifrs-full_InsuranceContractsByComponentsAxis"
BEL_MEMBER = "ifrs-full_EstimatesOfPresentValueOfFutureCashFlowsMember"

con = duckdb.connect("data/db/benchmark.duckdb", read_only=True)

# Sep × Issued × BEL_comp 포함 contexts + DI817 role (변동표) element만
SQL = """
WITH ko AS (
  SELECT ELMT_ID, MIN(LABEL) AS LABEL FROM lab_insurers
  WHERE CIK=? AND REPORT_DATE='20251231' AND LANG='ko'
  GROUP BY ELMT_ID
),
movement_elems AS (
  SELECT DISTINCT ELEMENT_ID FROM pre_insurers
  WHERE CIK=? AND REPORT_DATE='20251231'
    AND ROLE_ID LIKE 'dart_2024-06-30_role-DI817%'
),
ctxs AS (
  SELECT CONTEXT_ID,
         MAX(PERIOD_INSTANT) AS p_inst,
         MAX(PERIOD_START_DATE) AS p_start,
         MAX(PERIOD_END_DATE) AS p_end
  FROM cntxt_insurers
  WHERE CIK=? AND REPORT_DATE='20251231'
  GROUP BY CONTEXT_ID
  HAVING BOOL_OR(MEMBER_ELEMENT_ID = ?)
     AND BOOL_OR(MEMBER_ELEMENT_ID = ?)
     AND BOOL_OR(AXIS_ELEMENT_ID = ? AND MEMBER_ELEMENT_ID = ?)
     -- 「위험관리 21-2」 sub-table의 배당여부 partition 멤버 제외 (변동표 partition과 중복)
     AND NOT BOOL_OR(MEMBER_ELEMENT_ID LIKE '%OfDisclosureOfNatureAndExtentOfRisks%')
)
SELECT
  CASE WHEN c.p_inst = '2024-12-31' THEN 'opening'
       WHEN c.p_inst = '2025-12-31' THEN 'closing'
       WHEN c.p_start = '2025-01-01' AND c.p_end = '2025-12-31' THEN 'duration'
       ELSE NULL END AS period_kind,
  v.ELEMENT_ID,
  COALESCE(ko.LABEL, '') AS lbl,
  SUM(v.amount_krw)/1e8 AS amt_eok,
  COUNT(*) AS n_facts
FROM val_insurers v
JOIN ctxs c USING(CONTEXT_ID)
JOIN movement_elems me ON me.ELEMENT_ID = v.ELEMENT_ID
LEFT JOIN ko ON ko.ELMT_ID = v.ELEMENT_ID
WHERE v.CIK=? AND v.REPORT_DATE='20251231' AND v.amount_krw IS NOT NULL
  AND (c.p_inst IN ('2024-12-31','2025-12-31')
       OR (c.p_start='2025-01-01' AND c.p_end='2025-12-31'))
GROUP BY 1, v.ELEMENT_ID, ko.LABEL
HAVING SUM(v.amount_krw) IS NOT NULL
"""


def short_eid(eid: str) -> str:
    is_entity = False
    s = eid.replace("ifrs-full_", "").replace("dart_", "d:")
    for cik, _ in PEERS:
        if f"entity{cik}_" in s:
            is_entity = True
            s = s.replace(f"entity{cik}_", "[E]")
    return s[:75], is_entity


wb = Workbook()
wb.remove(wb.active)

hdr_font = Font(bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="1E3A5F")
bal_fill = PatternFill("solid", fgColor="FFF4CE")
sect_fill = PatternFill("solid", fgColor="DBE7F4")
chk_fill = PatternFill("solid", fgColor="E8F5E9")
ent_fill = PatternFill("solid", fgColor="FFE4E1")
sect_font = Font(bold=True, size=11)
border = Border(left=Side(style="thin", color="BBBBBB"),
                right=Side(style="thin", color="BBBBBB"),
                top=Side(style="thin", color="BBBBBB"),
                bottom=Side(style="thin", color="BBBBBB"))

for cik, name in PEERS:
    ws = wb.create_sheet(title=name)
    ws.append([f"§4-1A BEL 변동 (정규화) — {name}"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    ws.append(["원천: DI817100/105 「21-1. 보험계약부채(자산) 변동분의 차이조정 공시」 (별도)"])
    ws.cell(row=2, column=1).font = Font(italic=True, color="666666", size=10)
    ws.append(["축 합산: Sep × Issued × BEL component 셋만 조건, sub-axis(상품군·LRC/LIC 등)는 element 단위로 SUM"])
    ws.cell(row=3, column=1).font = Font(italic=True, color="666666", size=10)
    ws.append([])

    ws.append(["구분", "변동 line (한글 라벨)", "element_id (축약)", "표준?", "금액 (억원)", "fact 수"])
    hr = ws.max_row
    for c in range(1, 7):
        cell = ws.cell(row=hr, column=c)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    rows = con.execute(SQL, [cik, cik, cik, SEP, ISSUED, COMP_AXIS, BEL_MEMBER, cik]).fetchall()
    by_kind = {"opening": [], "duration": [], "closing": []}
    for kind, eid, lbl, amt, n in rows:
        if kind in by_kind:
            by_kind[kind].append((eid, lbl, amt, n))

    def section(title):
        ws.append([title] + [""]*5)
        r = ws.max_row
        ws.cell(row=r, column=1).font = sect_font
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = sect_fill
            ws.cell(row=r, column=c).border = border

    def write_rows(kind, items, is_bal=False):
        if not items:
            ws.append([kind, "(공시 없음)", "", "", "", ""])
            return 0
        ordered = sorted(items, key=lambda x: -abs(x[2] or 0))
        total = 0
        for eid, lbl, amt, n in ordered:
            seid, is_entity = short_eid(eid)
            std_mark = "✗ entity" if is_entity else "✓ 표준"
            ws.append([kind, lbl or "(라벨 없음)", seid, std_mark,
                       round(amt, 0) if amt is not None else None, n])
            r = ws.max_row
            ws.cell(row=r, column=5).number_format = '#,##0;-#,##0'
            for c in range(1, 7):
                ws.cell(row=r, column=c).border = border
                if is_bal:
                    ws.cell(row=r, column=c).fill = bal_fill
                elif is_entity:
                    ws.cell(row=r, column=c).fill = ent_fill
            total += amt or 0
        return total

    section("[1] 기초 잔액 (2024-12-31)")
    op_sum = write_rows("기초", by_kind["opening"], is_bal=True)

    section("[2] 변동 (FY2025)")
    du_sum = write_rows("변동", by_kind["duration"])

    section("[3] 기말 잔액 (2025-12-31)")
    cl_sum = write_rows("기말", by_kind["closing"], is_bal=True)

    ws.append([""]*6)
    section("[4] 검증: 기초 + Σ변동 = 기말 ?")
    diff = cl_sum - (op_sum + du_sum)
    for lbl, v in [
        ("기초 합계", op_sum),
        ("Σ변동", du_sum),
        ("기초 + Σ변동 (계산)", op_sum + du_sum),
        ("기말 합계 (실제)", cl_sum),
        ("잔차 (실제 − 계산)", diff),
    ]:
        ws.append(["", lbl, "", "", round(v, 0), ""])
        r = ws.max_row
        ws.cell(row=r, column=5).number_format = '#,##0;-#,##0'
        for c in range(1, 7):
            ws.cell(row=r, column=c).fill = chk_fill
            ws.cell(row=r, column=c).border = border
        ws.cell(row=r, column=2).font = Font(bold=(lbl.startswith("잔차")))

    for i, w in enumerate([10, 50, 60, 10, 16, 8], 1):
        ws.column_dimensions[chr(64+i)].width = w
    ws.freeze_panes = f"A{hr+1}"

# 변동 line cross-tab 시트
ct = wb.create_sheet(title="변동 비교", index=0)
ct.append(["BEL 변동 — 8개사 변동 line 정규화 비교 (FY2025, 별도, 억원)"])
ct.cell(row=1, column=1).font = Font(bold=True, size=14)
ct.append([])
hdrs = ["element_id (표준만)", "한글 라벨"] + [n for _, n in PEERS]
ct.append(hdrs)
hr = ct.max_row
for c in range(1, len(hdrs)+1):
    cell = ct.cell(row=hr, column=c)
    cell.font = hdr_font; cell.fill = hdr_fill
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = border

# duration facts만, 표준 ifrs-full element만, 회사 cross
peer_data = {}
for cik, _ in PEERS:
    rows = con.execute(SQL, [cik, cik, cik, SEP, ISSUED, COMP_AXIS, BEL_MEMBER, cik]).fetchall()
    peer_data[cik] = {eid: (lbl, amt) for kind, eid, lbl, amt, _ in rows
                      if kind == "duration" and eid.startswith("ifrs-full_")}

# 모든 표준 element 합집합
all_elems = set()
for cik, _ in PEERS:
    all_elems.update(peer_data[cik].keys())

# 자사 절대값 큰 순으로 정렬
mirae = peer_data["00112332"]
def sort_key(eid):
    return -abs(mirae.get(eid, ("", 0))[1] or 0) if eid in mirae else 0

for eid in sorted(all_elems, key=sort_key):
    lbl = ""
    for cik, _ in PEERS:
        if eid in peer_data[cik]:
            lbl = peer_data[cik][eid][0]; break
    row = [eid.replace("ifrs-full_", "")[:70], lbl[:50]]
    for cik, _ in PEERS:
        v = peer_data[cik].get(eid)
        row.append(round(v[1], 0) if v else None)
    ct.append(row)
    r = ct.max_row
    for c in range(3, len(hdrs)+1):
        ct.cell(row=r, column=c).number_format = '#,##0;-#,##0'
    for c in range(1, len(hdrs)+1):
        ct.cell(row=r, column=c).border = border
        ct.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)

for i, w in enumerate([55, 38] + [13]*len(PEERS), 1):
    if i <= 26:
        ct.column_dimensions[chr(64+i)].width = w
    else:
        ct.column_dimensions[f"A{chr(64+i-26)}"].width = w
ct.row_dimensions[hr].height = 36
ct.freeze_panes = f"C{hr+1}"

out = Path("outputs/bel_variation_normalized.xlsx")
wb.save(out)
print(f"wrote {out}")
print(f"sheets: {wb.sheetnames}")
print(f"표준 BEL 변동 element union: {len(all_elems)}개")
