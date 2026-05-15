"""§4-1A 보험계약부채 변동표 — 표준 element 기반 동업사 cross-tab.

설계:
  ROW = canonical 표준 element_id (IFRS17 변동표 표준 구조)
  COLUMN = 8개사
  VALUE = 회사별 FY2025 별도 + 발행 + 모든 컴포넌트 합산값
          (LRC+LIC, BEL+RA+CSM 합산. 사용자 disclosure 좌측표 형태)

회사별 presentation linkbase 차이는 무관 — element_id 만으로 1:1 매칭.
회사 disclosure 행 라벨은 다를 수 있어도, element_id 매핑은 같은 자원.
"""
from __future__ import annotations
import duckdb
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

PEERS = [
    ("00112332", "미래에셋생명"),
    ("00126256", "삼성생명"),
    ("00113058", "한화생명"),
    ("00117267", "동양생명"),
    ("00139214", "삼성화재"),
    ("00164973", "현대해상"),
    ("00159102", "DB손해보험"),
    ("00135917", "한화손해보험"),
]
SEP = "ifrs-full_SeparateMember"
ISSUED = "ifrs-full_InsuranceContractsIssuedMember"

con = duckdb.connect("data/db/benchmark.duckdb", read_only=True)

# 변동표 21-1 (DI817100/105) role 소속 element, Sep × Issued, 위험관리 sub-table 제외,
# 컴포넌트 분해 axis(BEL/RA/CSM) 제외 = 합계 row 만
SQL = """
WITH mv_elems AS (
  SELECT DISTINCT ELEMENT_ID FROM pre_insurers
  WHERE CIK=? AND REPORT_DATE='20251231'
    AND ROLE_ID LIKE 'dart_2024-06-30_role-DI817%'
),
ctxs AS (
  SELECT CONTEXT_ID,
         MAX(PERIOD_INSTANT) AS p_inst,
         MAX(PERIOD_START_DATE) AS p_start,
         MAX(PERIOD_END_DATE) AS p_end
  FROM cntxt_insurers WHERE CIK=? AND REPORT_DATE='20251231'
  GROUP BY CONTEXT_ID
  HAVING BOOL_OR(MEMBER_ELEMENT_ID = ?)   -- Separate
     AND BOOL_OR(MEMBER_ELEMENT_ID = ?)   -- Issued
     AND NOT BOOL_OR(MEMBER_ELEMENT_ID LIKE '%OfDisclosureOfNatureAndExtentOfRisks%')
     AND NOT BOOL_OR(AXIS_ELEMENT_ID = 'ifrs-full_InsuranceContractsByComponentsAxis')
)
SELECT
  CASE WHEN c.p_inst='2024-12-31' THEN 'opening'
       WHEN c.p_inst='2025-12-31' THEN 'closing'
       WHEN c.p_start='2025-01-01' AND c.p_end='2025-12-31' THEN 'duration'
       ELSE NULL END,
  v.ELEMENT_ID,
  SUM(v.amount_krw)/1e8
FROM val_insurers v
JOIN ctxs c USING(CONTEXT_ID)
JOIN mv_elems me ON me.ELEMENT_ID=v.ELEMENT_ID
WHERE v.CIK=? AND v.REPORT_DATE='20251231' AND v.amount_krw IS NOT NULL
  AND (c.p_inst IN ('2024-12-31','2025-12-31')
       OR (c.p_start='2025-01-01' AND c.p_end='2025-12-31'))
GROUP BY 1, v.ELEMENT_ID
HAVING SUM(v.amount_krw) IS NOT NULL
"""

# canonical 행 schema — IFRS17 변동표 표준 element + 한글 라벨 + 그룹
ROWS = [
    # (group, label, element_id, period_kind)
    ("기초",       "부채인 보험계약 기초",      "ifrs-full_InsuranceContractsThatAreLiabilities",                                                                 "opening"),
    ("기초",       "자산인 보험계약 기초",      "ifrs-full_InsuranceContractsThatAreAssets",                                                                      "opening"),
    ("보험수익",   "보험수익",                  "ifrs-full_InsuranceRevenue",                                                                                     "duration"),
    ("보험수익",   "발행 보험계약 보험수익(LRC 변동)", "ifrs-full_IncreaseDecreaseThroughInsuranceRevenueInsuranceContractsLiabilityAsset",                          "duration"),
    ("보험서비스비용", "발생 보험금 및 기타 서비스비용 발생", "ifrs-full_IncreaseDecreaseThroughIncurredClaimsAndOtherInsuranceServiceExpensesExcludingInsuranceAcquisitionCashFlowsInsuranceContractsLiabilityAsset", "duration"),
    ("보험서비스비용", "보험취득CF 상각",       "ifrs-full_IncreaseDecreaseThroughAmortisationOfInsuranceAcquisitionCashFlowsInsuranceContractsLiabilityAsset",   "duration"),
    ("보험서비스비용", "발생사고요소 조정",     "ifrs-full_IncreaseDecreaseThroughAdjustmentsToLiabilityForIncurredClaimsRelatedToPastServiceInsuranceContractsLiabilityAsset", "duration"),
    ("보험서비스비용", "손실부담계약 손실(환입)", "ifrs-full_IncreaseDecreaseThroughEffectsOfGroupsOfOnerousContractsInitiallyRecognisedInPeriodInsuranceContractsLiabilityAsset", "duration"),
    ("보험서비스비용", "보험서비스결과 (그룹)",  "ifrs-full_IncreaseDecreaseThroughInsuranceServiceResultInsuranceContractsLiabilityAsset",                       "duration"),
    ("현금흐름",   "수취한 보험료",             "ifrs-full_IncreaseDecreaseThroughPremiumsReceivedForInsuranceContractsIssuedInsuranceContractsLiabilityAsset",     "duration"),
    ("현금흐름",   "보험취득CF 지급",           "ifrs-full_IncreaseDecreaseThroughInsuranceAcquisitionCashFlowsInsuranceContractsLiabilityAsset",                 "duration"),
    ("현금흐름",   "보험금/서비스비용 지급",    "ifrs-full_IncreaseDecreaseThroughIncurredClaimsPaidAndOtherInsuranceServiceExpensesPaidForInsuranceContractsIssuedExcludingInsuranceAcquisitionCashFlowsInsuranceContractsLiabilityAsset", "duration"),
    ("현금흐름",   "현금흐름 (그룹)",           "ifrs-full_CashFlowsFromUsedInInsuranceContracts",                                                                "duration"),
    ("금융손익",   "보험금융손익 (당기손익)",   "ifrs-full_InsuranceFinanceIncomeExpensesFromInsuranceContractsIssuedRecognisedInProfitOrLoss",                  "duration"),
    ("금융손익",   "보험금융손익 (OCI)",        "ifrs-full_InsuranceFinanceIncomeExpensesFromInsuranceContractsIssuedRecognisedInOtherComprehensiveIncome",       "duration"),
    ("금융손익",   "보험금융손익 (그룹)",       "ifrs-full_IncreaseDecreaseThroughInsuranceFinanceIncomeOrExpensesInsuranceContractsLiabilityAsset",             "duration"),
    ("미래서비스", "처음 인식 계약 효과",       "ifrs-full_IncreaseDecreaseThroughEffectsOfContractsInitiallyRecognisedInPeriodInsuranceContractsLiabilityAsset",  "duration"),
    ("미래서비스", "CSM 조정 추정변동",         "ifrs-full_IncreaseDecreaseThroughChangesInEstimatesThatAdjustContractualServiceMarginInsuranceContractsLiabilityAsset", "duration"),
    ("미래서비스", "CSM 미조정 추정변동",       "ifrs-full_IncreaseDecreaseThroughChangesInEstimatesThatDoNotAdjustContractualServiceMarginInsuranceContractsLiabilityAsset", "duration"),
    ("미래서비스", "현행 서비스 변동",          "ifrs-full_IncreaseDecreaseThroughChangesThatRelateToCurrentServiceInsuranceContractsLiabilityAsset",            "duration"),
    ("과거서비스", "과거 서비스 변동",          "ifrs-full_IncreaseDecreaseThroughChangesThatRelateToPastServiceInsuranceContractsLiabilityAsset",                 "duration"),
    ("위험조정",   "RA 변동(미래/과거 무관)",   "ifrs-full_IncreaseDecreaseThroughChangeInRiskAdjustmentForNonfinancialRiskThatDoesNotRelateToFutureOrPastServiceInsuranceContractsLiabilityAsset", "duration"),
    ("CSM 상각",   "보험계약마진 당기인식",     "ifrs-full_InsuranceRevenueContractualServiceMarginRecognisedInProfitOrLoss",                                    "duration"),
    ("CSM 상각",   "CSM 당기손익인식 (변동표)", "ifrs-full_IncreaseDecreaseThroughRecognitionOfContractualServiceMarginInProfitOrLossToReflectInsuranceContractServicesProvidedInPeriodInsuranceContractsLiabilityAsset", "duration"),
    ("기타",       "경험조정",                  "ifrs-full_IncreaseDecreaseThroughExperienceAdjustmentsInsuranceContractsLiabilityAsset",                         "duration"),
    ("기타",       "투자요소/보험료환급",       "ifrs-full_IncreaseDecreaseThroughInvestmentComponentAndPremiumRefundExcludedFromInsuranceRevenueAndInsuranceServiceExpensesInsuranceContractsLiabilityAsset", "duration"),
    ("기타",       "기타 변동",                 "ifrs-full_IncreaseDecreaseThroughOtherChangesLiabilitiesUnderInsuranceContractsAndReinsuranceContractsIssued", "duration"),
    ("기말",       "부채인 보험계약 기말",      "ifrs-full_InsuranceContractsThatAreLiabilities",                                                                 "closing"),
    ("기말",       "자산인 보험계약 기말",      "ifrs-full_InsuranceContractsThatAreAssets",                                                                      "closing"),
]

# 회사별 데이터 수집
peer_facts = {}
for cik, _ in PEERS:
    rows = con.execute(SQL, [cik, cik, SEP, ISSUED, cik]).fetchall()
    facts = {}
    for kind, eid, amt in rows:
        if kind:
            facts[(kind, eid)] = amt
    peer_facts[cik] = facts

# Excel
wb = Workbook()
ws = wb.active
ws.title = "BEL 변동 cross-tab"

ws.append(["§4-1A 보험계약부채 변동표 — 표준 element 기반 동업사 cross-tab (FY2025 별도, 발행, 억원)"])
ws.cell(row=1, column=1).font = Font(bold=True, size=14)
ws.append(["필터: 별도 × 발행, role=DI817*, 컴포넌트 분해축 제외(=BEL+RA+CSM 합계), 위험관리 sub-table 멤버 제외"])
ws.cell(row=2, column=1).font = Font(italic=True, color="666666", size=10)
ws.append([])

hdrs = ["그룹", "표준 row 라벨", "element_id (축약)"] + [n for _, n in PEERS]
ws.append(hdrs)
hr = ws.max_row
for c in range(1, len(hdrs)+1):
    cell = ws.cell(row=hr, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1E3A5F")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(left=Side(style="thin", color="888888"),
                         right=Side(style="thin", color="888888"),
                         top=Side(style="thin", color="888888"),
                         bottom=Side(style="thin", color="888888"))

GROUP_FILL = {
    "기초":       "FFF4CE",
    "보험수익":   "E3F2FD",
    "보험서비스비용": "FFF3E0",
    "현금흐름":   "E0F7FA",
    "금융손익":   "F3E5F5",
    "미래서비스": "E8F5E9",
    "과거서비스": "E8F5E9",
    "위험조정":   "E8F5E9",
    "CSM 상각":   "E8F5E9",
    "기타":       "FAFAFA",
    "기말":       "FFF4CE",
}
border = Border(left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"))

prev_group = None
for grp, lbl, eid, kind in ROWS:
    short = eid.replace("ifrs-full_", "")[:55]
    row = [grp, lbl, short]
    for cik, _ in PEERS:
        v = peer_facts[cik].get((kind, eid))
        row.append(round(v, 0) if v is not None else None)
    ws.append(row)
    r = ws.max_row
    fill = PatternFill("solid", fgColor=GROUP_FILL.get(grp, "FFFFFF"))
    for c in range(1, len(hdrs)+1):
        ws.cell(row=r, column=c).fill = fill
        ws.cell(row=r, column=c).border = border
    for c in range(4, len(hdrs)+1):
        ws.cell(row=r, column=c).number_format = '#,##0;-#,##0;"–"'
        ws.cell(row=r, column=c).alignment = Alignment(horizontal="right")
    if grp != prev_group:
        ws.cell(row=r, column=1).font = Font(bold=True, color="1E3A5F")
    prev_group = grp

# 합계 row: Σ변동 (duration만)
ws.append([])
ws.append(["검증", "Σ변동 (duration 합)", "", ""] + [""]*len(PEERS))
sum_row_idx = ws.max_row
for i, (cik, _) in enumerate(PEERS, start=4):
    col = get_column_letter(i)
    formula_start = hr + 1
    formula_end = hr + len(ROWS)
    ws.cell(row=sum_row_idx, column=i).value = f"=SUM({col}{formula_start}:{col}{formula_end})-{col}{formula_start}-{col}{formula_start+1}-{col}{formula_end-1}-{col}{formula_end}"
    ws.cell(row=sum_row_idx, column=i).number_format = '#,##0;-#,##0'
ws.cell(row=sum_row_idx, column=1).font = Font(bold=True)
ws.cell(row=sum_row_idx, column=2).font = Font(bold=True)
for c in range(1, len(hdrs)+1):
    ws.cell(row=sum_row_idx, column=c).fill = PatternFill("solid", fgColor="E8F5E9")
    ws.cell(row=sum_row_idx, column=c).border = border

# 컬럼 폭
widths = [10, 32, 60] + [13]*len(PEERS)
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[hr].height = 30
ws.freeze_panes = f"D{hr+1}"

# 참고 시트: per-company 보고 여부 매트릭스
ref = wb.create_sheet(title="회사별 보고 매트릭스")
ref.append(["element 보유 매트릭스 — Y = XBRL에 보고됨 (값 0 포함), – = 미보고"])
ref.cell(row=1, column=1).font = Font(bold=True, size=12)
ref.append([])
ref.append(["표준 row 라벨", "element_id (축약)"] + [n for _, n in PEERS])
hr2 = ref.max_row
for c in range(1, len(PEERS)+3):
    cell = ref.cell(row=hr2, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1E3A5F")
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
    cell.border = border

for grp, lbl, eid, kind in ROWS:
    short = eid.replace("ifrs-full_", "")[:55]
    row = [lbl, short]
    for cik, _ in PEERS:
        v = peer_facts[cik].get((kind, eid))
        row.append("Y" if v is not None else "–")
    ref.append(row)
    r = ref.max_row
    for c in range(1, len(PEERS)+3):
        ref.cell(row=r, column=c).border = border
        ref.cell(row=r, column=c).alignment = Alignment(horizontal="center" if c > 2 else "left")
for i, w in enumerate([28, 60] + [10]*len(PEERS), 1):
    ref.column_dimensions[get_column_letter(i)].width = w
ref.row_dimensions[hr2].height = 28
ref.freeze_panes = f"C{hr2+1}"

out = Path("outputs/bel_canonical_crosstab.xlsx")
wb.save(out)
print(f"wrote {out}")
print(f"sheets: {wb.sheetnames}")
