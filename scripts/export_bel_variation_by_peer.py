"""§4-1A BEL 변동 — 회사별 시트, 주석공시 표 형태.

전략: 회사별 (기간×원형: 기초/변동/기말) 단위로
  - Sep × Issued × BEL component 포함 contexts 중 axis 깊이가 가장 얕은 것만 사용
  - 그 깊이의 contexts에서 모든 fact를 SUM (LRC/LIC·상품군·loss 등 sub-axis는 합산 처리)
  - 결과: element_id 단위로 1 row, 금액은 회사 공시 그대로의 raw 합계

표 컬럼: 구분 | 변동 line (한글) | element_id | 금액(억원)
"""
from __future__ import annotations
import duckdb
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from pathlib import Path

PEERS = [
    ("00112332", "미래에셋생명"),
    ("00126256", "삼성생명"),
    ("00113058", "한화생명"),
    ("00117267", "동양생명"),
    ("00139214", "삼성화재"),
    ("00164973", "현대해상"),
    ("00159102", "DB손해보험"),
    ("00135917", "한화손해보험"),
]
SEP = "ifrs-full_SeparateMember"
ISSUED = "ifrs-full_InsuranceContractsIssuedMember"
COMP_AXIS = "ifrs-full_InsuranceContractsByComponentsAxis"
BEL_MEMBER = "ifrs-full_EstimatesOfPresentValueOfFutureCashFlowsMember"

con = duckdb.connect("data/db/benchmark.duckdb", read_only=True)

# 1) 회사의 Sep×Issued×BEL contexts와 axis 깊이
SQL_CTXS = """
SELECT CONTEXT_ID, COUNT(*) AS n_ax,
       MAX(PERIOD_INSTANT) AS p_inst,
       MAX(PERIOD_START_DATE) AS p_start,
       MAX(PERIOD_END_DATE) AS p_end
FROM cntxt_insurers
WHERE CIK=? AND REPORT_DATE='20251231'
GROUP BY CONTEXT_ID
HAVING BOOL_OR(MEMBER_ELEMENT_ID=?)
   AND BOOL_OR(MEMBER_ELEMENT_ID=?)
   AND BOOL_OR(AXIS_ELEMENT_ID=? AND MEMBER_ELEMENT_ID=?)
"""

SQL_FACTS = """
WITH ko AS (
  SELECT ELMT_ID, MIN(LABEL) AS LABEL FROM lab_insurers
  WHERE CIK=? AND REPORT_DATE='20251231' AND LANG='ko'
  GROUP BY ELMT_ID
)
SELECT v.ELEMENT_ID, COALESCE(ko.LABEL,'') AS lbl, SUM(v.amount_krw)/1e8
FROM val_insurers v
LEFT JOIN ko ON ko.ELMT_ID = v.ELEMENT_ID
WHERE v.CIK=? AND v.REPORT_DATE='20251231' AND v.amount_krw IS NOT NULL
  AND v.CONTEXT_ID = ANY(?)
GROUP BY v.ELEMENT_ID, ko.LABEL
ORDER BY ABS(SUM(v.amount_krw)) DESC
"""


def short_eid(eid: str) -> str:
    s = eid.replace("ifrs-full_", "").replace("dart_", "d:")
    for cik, _ in PEERS:
        s = s.replace(f"entity{cik}_", "[E]")
    return s[:80]


wb = Workbook()
wb.remove(wb.active)

hdr_font = Font(bold=True, color="FFFFFF")
hdr_fill = PatternFill("solid", fgColor="1E3A5F")
bal_fill = PatternFill("solid", fgColor="FFF4CE")
sect_fill = PatternFill("solid", fgColor="DBE7F4")
chk_fill = PatternFill("solid", fgColor="E8F5E9")
sect_font = Font(bold=True, size=11)
border = Border(left=Side(style="thin", color="BBBBBB"),
                right=Side(style="thin", color="BBBBBB"),
                top=Side(style="thin", color="BBBBBB"),
                bottom=Side(style="thin", color="BBBBBB"))

for cik, name in PEERS:
    ws = wb.create_sheet(title=name)
    ws.append([f"§4-1A BEL 변동 — {name}"])
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)

    ctxs = con.execute(SQL_CTXS, [cik, SEP, ISSUED, COMP_AXIS, BEL_MEMBER]).fetchall()

    # 기간별 분류 + min axis depth
    opening = [c for c in ctxs if c[2] == "2024-12-31"]
    closing = [c for c in ctxs if c[2] == "2025-12-31"]
    duration = [c for c in ctxs if c[3] == "2025-01-01" and c[4] == "2025-12-31"]

    op_min = min((c[1] for c in opening), default=None)
    cl_min = min((c[1] for c in closing), default=None)
    du_min = min((c[1] for c in duration), default=None)

    ws.append([f"기초 axis depth = {op_min} · 변동 axis depth = {du_min} · 기말 axis depth = {cl_min}  "
               f"(각 기간 최소 깊이 contexts만 사용 — sub-axis는 합산 처리)"])
    ws.cell(row=2, column=1).font = Font(italic=True, color="666666", size=10)
    ws.append([])

    # 헤더
    ws.append(["구분", "변동 line (한글)", "element_id (축약)", "금액 (억원)"])
    hr = ws.max_row
    for c in range(1, 5):
        cell = ws.cell(row=hr, column=c)
        cell.font = hdr_font; cell.fill = hdr_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = border

    def section(title):
        ws.append([title, "", "", ""])
        r = ws.max_row
        ws.cell(row=r, column=1).font = sect_font
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = sect_fill
            ws.cell(row=r, column=c).border = border

    def fetch_section(ctxs_filtered, min_depth):
        if min_depth is None: return []
        ids = [c[0] for c in ctxs_filtered if c[1] == min_depth]
        if not ids: return []
        return con.execute(SQL_FACTS, [cik, cik, ids]).fetchall()

    def write_rows(kind, rows, is_bal=False):
        if not rows:
            ws.append([kind, "(공시 없음)", "", ""])
            r = ws.max_row
            for c in range(1, 5): ws.cell(row=r, column=c).border = border
            return 0
        total = 0
        for eid, lbl, amt in rows:
            ws.append([kind, lbl or "(라벨 없음)", short_eid(eid),
                       round(amt, 0) if amt is not None else None])
            r = ws.max_row
            ws.cell(row=r, column=4).number_format = '#,##0;-#,##0'
            for c in range(1, 5):
                ws.cell(row=r, column=c).border = border
                if is_bal: ws.cell(row=r, column=c).fill = bal_fill
            total += amt or 0
        return total

    section(f"[1] 기초 잔액 (2024-12-31)")
    op_rows = fetch_section(opening, op_min)
    op_sum = write_rows("기초", op_rows, is_bal=True)

    section(f"[2] 변동 (FY2025)")
    du_rows = fetch_section(duration, du_min)
    du_sum = write_rows("변동", du_rows)

    section(f"[3] 기말 잔액 (2025-12-31)")
    cl_rows = fetch_section(closing, cl_min)
    cl_sum = write_rows("기말", cl_rows, is_bal=True)

    ws.append([])
    section("[4] 검증: 기초 + Σ변동 = 기말 ?")
    for lbl, v in [
        ("기초 합계", op_sum),
        ("Σ변동", du_sum),
        ("기초 + Σ변동", op_sum + du_sum),
        ("기말 합계", cl_sum),
        ("잔차 (기말 − 계산)", cl_sum - op_sum - du_sum),
    ]:
        ws.append(["", lbl, "", round(v, 0)])
        r = ws.max_row
        ws.cell(row=r, column=4).number_format = '#,##0;-#,##0'
        for c in range(1, 5):
            ws.cell(row=r, column=c).fill = chk_fill
            ws.cell(row=r, column=c).border = border
        ws.cell(row=r, column=2).font = Font(bold=lbl.startswith("잔차"))

    for i, w in enumerate([10, 45, 60, 16], 1):
        ws.column_dimensions[chr(64+i)].width = w
    ws.freeze_panes = f"A{hr+1}"

out = Path("outputs/bel_variation_by_peer.xlsx")
wb.save(out)
print(f"wrote {out}")
print(f"sheets: {wb.sheetnames}")
