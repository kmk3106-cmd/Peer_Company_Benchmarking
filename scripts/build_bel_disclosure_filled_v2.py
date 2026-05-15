"""BEL 변동 disclosure — v2: 회사별 label keyword 매칭으로 element 찾기.

전략 변경:
- 각 13행 라벨에 대해 keyword set 정의
- 회사별 DI817105 element 들의 한국어 라벨 (모든 LABEL_ROLE) 에서 keyword 매칭
- 매칭된 모든 element 의 Sep×Issued (비-Component, 비-위험관리) top-level 합산
- 단, 한 element 가 여러 행에 매칭되면 우선순위 (가장 specific) 적용
- 잔차 검증
"""
from __future__ import annotations
import duckdb
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

PEERS = [
    ("00112332", "미래에셋생명"),
    ("00126256", "삼성생명"),
    ("00113058", "한화생명"),
    ("00117267", "동양생명"),
    ("00139214", "삼성화재"),
    ("00164973", "현대해상"),
    ("00159102", "DB손해보험"),
    ("00135917", "한화손해보험"),
]
SEP = "ifrs-full_SeparateMember"
ISSUED = "ifrs-full_InsuranceContractsIssuedMember"
ROLE = "dart_2024-06-30_role-DI817105"

# 13행 (라벨, 기간, mirae 사용자 값 원 단위)
ROWS = [
    ("부채인 보험계약의 기초 장부금액",            "opening", 26_248_411_835_353),
    ("보험수익",                                  "duration",-1_080_411_547_554),
    ("발생한 보험금 및 기타 보험서비스비용",      "duration",  +592_626_787_607),
    ("보험취득현금흐름 상각",                     "duration",  +218_864_437_494),
    ("발생사고요소의 조정",                       "duration",   +19_584_794_969),
    ("손실부담계약 관련 손실(환입)",              "duration",   +66_710_164,165),  # syntax error fix below
]
# fix tuple syntax
ROWS = [
    ("부채인 보험계약의 기초 장부금액",            "opening", 26_248_411_835_353),
    ("보험수익",                                  "duration",-1_080_411_547_554),
    ("발생한 보험금 및 기타 보험서비스비용",      "duration",  +592_626_787_607),
    ("보험취득현금흐름 상각",                     "duration",  +218_864_437_494),
    ("발생사고요소의 조정",                       "duration",   +19_584_794_969),
    ("손실부담계약 관련 손실(환입)",              "duration",   +66_710_164_165),
    ("수취한 보험료",                             "duration",+4_050_027_788_462),
    ("보험취득현금흐름 지급",                     "duration",  -594_834_366_070),
    ("보험금(투자요소 포함) 및 기타보험서비스비용의 지급", "duration", -4_296_950_166_750),
    ("당기손익인식 보험금융손익",                 "duration",+2_272_778_949_192),
    ("기타포괄손익인식 보험금융손익",             "duration",  -483_472_446_698),
    ("기타증감",                                  "duration",   -16_637_974_642),
    ("부채인 보험계약의 기말 장부금액",           "closing", 26_996_698_255_523),
]

# 13행별 캐노니컬 element 화이트리스트 (taxonomy 표준 + 회사 변형 element 까지)
# 회사별로 분해 단위가 달라도 모두 같은 의미 → 한 행에 합산
# 키워드 기반 element_id 매칭 사용
ELEMENT_PATTERNS = {
    "부채인 보험계약의 기초 장부금액": {
        "period": "opening",
        # 우선순위: net (LiabilityAsset) → liability only
        # "부채인 보험계약" = ThatAreLiabilities. 회사별 사용 element 다름:
        # 손보사 (DB, 한손) — Issued 접두사 사용 (NatureOfInsuranceContracts axis 로 top-level 추출 가능)
        # 생보사 (삼성생명, 한화생명, 동양) — Issued 없는 ThatAreLiabilities 사용 (n_extra=0 또는 n_extra=1 단일axis)
        # LiabilityAsset 은 net 값, 마지막 fallback
        "primary_eids": ["ifrs-full_InsuranceContractsIssuedThatAreLiabilities",
                         "ifrs-full_InsuranceContractsThatAreLiabilities"],
        "fallback_eids": ["ifrs-full_InsuranceContractsLiabilityAsset"],
    },
    "부채인 보험계약의 기말 장부금액": {
        "period": "closing",
        # "부채인 보험계약" = ThatAreLiabilities. 회사별 사용 element 다름:
        # 손보사 (DB, 한손) — Issued 접두사 사용 (NatureOfInsuranceContracts axis 로 top-level 추출 가능)
        # 생보사 (삼성생명, 한화생명, 동양) — Issued 없는 ThatAreLiabilities 사용 (n_extra=0 또는 n_extra=1 단일axis)
        # LiabilityAsset 은 net 값, 마지막 fallback
        "primary_eids": ["ifrs-full_InsuranceContractsIssuedThatAreLiabilities",
                         "ifrs-full_InsuranceContractsThatAreLiabilities"],
        "fallback_eids": ["ifrs-full_InsuranceContractsLiabilityAsset"],
    },
    "보험수익": {
        "period": "duration",
        "primary_eids": ["ifrs-full_IncreaseDecreaseThroughInsuranceRevenueInsuranceContractsLiabilityAsset"],
        "fallback_eids": [],
    },
    "발생한 보험금 및 기타 보험서비스비용": {
        "period": "duration",
        # taxonomy: IncurredClaimsAndOtherIncurredInsuranceServiceExpenses...LiabilityAsset
        "primary_eids": ["ifrs-full_IncreaseDecreaseThroughIncurredClaimsAndOtherIncurredInsuranceServiceExpensesInsuranceContractsLiabilityAsset"],
        # mirae 는 "Adjustment" 포함 element 사용
        "fallback_eids": ["ifrs-full_IncreaseDecreaseThroughIncurredClaimsAndOtherIncurredInsuranceServiceExpensesAdjustmentInsuranceContractsLiabilityAsset"],
    },
    "보험취득현금흐름 상각": {
        "period": "duration",
        "primary_eids": ["ifrs-full_IncreaseDecreaseThroughAmortisationOfInsuranceAcquisitionCashFlowsInsuranceContractsLiabilityAsset"],
        "fallback_eids": [],
    },
    "발생사고요소의 조정": {
        "period": "duration",
        # 과거 서비스 (LIC 추정변경)
        "primary_eids": ["ifrs-full_IncreaseDecreaseThroughChangesThatRelateToPastServiceInsuranceContractsLiabilityAsset"],
        "fallback_eids": [],
    },
    "손실부담계약 관련 손실(환입)": {
        "period": "duration",
        "sum_mode": "sum_all",
        # 21-1 표 = 두 element 합 (최초 인식 onerous + CSM 미조정 추정변경)
        "primary_eids": [
            "ifrs-full_IncreaseDecreaseThroughEffectsOfGroupsOfOnerousContractsInitiallyRecognisedInPeriodInsuranceContractsLiabilityAsset",
            "ifrs-full_IncreaseDecreaseThroughChangesInEstimatesThatDoNotAdjustContractualServiceMarginInsuranceContractsLiabilityAsset",
        ],
        "fallback_eids": [],
    },
    "수취한 보험료": {
        "period": "duration",
        "primary_eids": ["ifrs-full_IncreaseDecreaseThroughPremiumsReceivedForInsuranceContractsIssuedInsuranceContractsLiabilityAsset"],
        "fallback_eids": [],
    },
    "보험취득현금흐름 지급": {
        "period": "duration",
        "primary_eids": ["ifrs-full_IncreaseDecreaseThroughInsuranceAcquisitionCashFlowsInsuranceContractsLiabilityAsset"],
        "fallback_eids": [],
    },
    "보험금(투자요소 포함) 및 기타보험서비스비용의 지급": {
        "period": "duration",
        "primary_eids": ["ifrs-full_IncreaseDecreaseThroughIncurredClaimsPaidAndOtherInsuranceServiceExpensesPaidForInsuranceContractsIssuedExcludingInsuranceAcquisitionCashFlowsInsuranceContractsLiabilityAsset"],
        "fallback_eids": [],
    },
    "당기손익인식 보험금융손익": {
        "period": "duration",
        # 두 element 모두 같은 의미. mirae 는 둘 다 22,728/-22,728 으로 보고 (반대부호). 부채 변동분 (LiabilityAsset) 사용
        "primary_eids": ["ifrs-full_IncreaseDecreaseThroughInsuranceFinanceIncomeOrExpensesInsuranceContractsLiabilityAsset"],
        "fallback_eids": ["ifrs-full_InsuranceFinanceIncomeExpensesFromInsuranceContractsIssuedRecognisedInProfitOrLoss"],
    },
    "기타포괄손익인식 보험금융손익": {
        "period": "duration",
        # taxonomy 정확 명칭 (suffix "ThatWillBeReclassifiedToProfitOrLoss")
        "primary_eids": ["ifrs-full_OtherComprehensiveIncomeBeforeTaxInsuranceFinanceIncomeExpensesFromInsuranceContractsIssuedExcludedFromProfitOrLossThatWillBeReclassifiedToProfitOrLoss"],
        "fallback_eids": ["ifrs-full_OtherComprehensiveIncomeBeforeTaxInsuranceFinanceIncomeExpensesFromInsuranceContractsIssuedExcludedFromProfitOrLoss"],
    },
    "기타증감": {
        "period": "duration",
        # taxonomy 정확 명칭 (회사 사용 short form)
        "primary_eids": [
            "ifrs-full_IncreaseDecreaseThroughAdditionalItemsNecessaryToUnderstandChangeInsuranceContractsLiabilityAsset",
        ],
        # mirae 의 긴 변형도 fallback
        "fallback_eids": [
            "ifrs-full_IncreaseDecreaseThroughAdditionalItemsNecessaryToUnderstandChangeInNetCarryingAmountOfInsuranceContractsInsuranceContractsLiabilityAsset",
            "ifrs-full_IncreaseDecreaseThroughTransfersAndOtherChangesEquity",
        ],
    },
}

# 미래에셋 사용자 값
MIRAE_USER = {(label, pk): v for label, pk, v in ROWS}

con = duckdb.connect("data/db/benchmark.duckdb", read_only=True)

# 1) 회사별 DI817105 에 있는 element_id 집합
def role_eids(cik):
    rows = con.execute("""
        SELECT DISTINCT ELEMENT_ID FROM pre_insurers
        WHERE CIK=? AND REPORT_DATE='20251231' AND ROLE_ID=?
    """, [cik, ROLE]).fetchall()
    return {r[0] for r in rows}

# 2) Sep × Issued, top-level (min n_ax) 값 추출
SQL_VAL = """
-- Sep × Issued, 위험관리 제외, Component axis 제외
-- "extra axis" = Sep, Disagg 외 axis
-- 우선순위:
--   pri 0: extra=0 (Sep×Issued 단독 = 진짜 top-level)
--   pri 1: extra=1 단일 axis, 그 axis 의 모든 member 합 (mutually exclusive decomposition)
--   pri 2: extra>=2 다중 axis — 가장 큰 단일 extra axis 의 합
WITH ctx_ax AS (
  SELECT CONTEXT_ID,
         MAX(PERIOD_INSTANT) AS p_inst,
         MAX(PERIOD_START_DATE) AS p_start,
         MAX(PERIOD_END_DATE) AS p_end,
         COUNT(DISTINCT CASE
           WHEN AXIS_ELEMENT_ID IN ('ifrs-full_ConsolidatedAndSeparateFinancialStatementsAxis',
                                    'ifrs-full_DisaggregationOfInsuranceContractsAxis') THEN NULL
           ELSE AXIS_ELEMENT_ID END) AS n_extra,
         MAX(CASE
           WHEN AXIS_ELEMENT_ID IN ('ifrs-full_ConsolidatedAndSeparateFinancialStatementsAxis',
                                    'ifrs-full_DisaggregationOfInsuranceContractsAxis') THEN NULL
           ELSE AXIS_ELEMENT_ID END) AS extra_axis,
         MIN(CASE
           WHEN AXIS_ELEMENT_ID IN ('ifrs-full_ConsolidatedAndSeparateFinancialStatementsAxis',
                                    'ifrs-full_DisaggregationOfInsuranceContractsAxis') THEN NULL
           ELSE AXIS_ELEMENT_ID END) AS extra_axis_min
  FROM cntxt_insurers WHERE CIK=? AND REPORT_DATE='20251231'
  GROUP BY CONTEXT_ID
  HAVING BOOL_OR(MEMBER_ELEMENT_ID = ?) AND BOOL_OR(MEMBER_ELEMENT_ID = ?)
     AND NOT BOOL_OR(MEMBER_ELEMENT_ID LIKE '%OfDisclosureOfNatureAndExtentOfRisks%')
     AND NOT BOOL_OR(AXIS_ELEMENT_ID = 'ifrs-full_InsuranceContractsByComponentsAxis')
),
fact AS (
  SELECT v.CONTEXT_ID, c.n_extra, c.extra_axis, c.extra_axis_min,
         CASE WHEN c.p_inst='2024-12-31' THEN 'opening'
              WHEN c.p_inst='2025-12-31' THEN 'closing'
              WHEN c.p_start='2025-01-01' AND c.p_end='2025-12-31' THEN 'duration'
              ELSE NULL END AS pk,
         MAX(v.amount_krw) AS amount_krw
  FROM val_insurers v JOIN ctx_ax c USING(CONTEXT_ID)
  WHERE v.CIK=? AND v.REPORT_DATE='20251231' AND v.ELEMENT_ID=? AND v.amount_krw IS NOT NULL
  GROUP BY v.CONTEXT_ID, c.n_extra, c.extra_axis, c.extra_axis_min, c.p_inst, c.p_start, c.p_end
),
-- pri 0: n_extra=0
p0 AS (SELECT SUM(amount_krw) AS amt FROM fact WHERE pk=? AND n_extra=0),
-- pri 1: 각 단일 axis (n_extra=1) 별로 합산. axis 별 합 중 가장 큰 absolute 가 그 행의 값
p1 AS (
  SELECT SUM(amount_krw) AS amt, extra_axis
  FROM fact WHERE pk=? AND n_extra=1
  GROUP BY extra_axis
  ORDER BY ABS(SUM(amount_krw)) DESC LIMIT 1
),
-- pri 2: n_extra>=2 — (n_extra, extra_axis) 조합별 합산. 가장 작은 n_extra (가장 top-level) 우선
p2 AS (
  SELECT SUM(amount_krw) AS amt, n_extra, extra_axis
  FROM fact WHERE pk=? AND n_extra>=2
  GROUP BY n_extra, extra_axis
  ORDER BY n_extra ASC, ABS(SUM(amount_krw)) DESC LIMIT 1
)
SELECT amt FROM (
  SELECT amt, 0 AS pri FROM p0 WHERE amt IS NOT NULL
  UNION ALL SELECT amt, 1 FROM p1 WHERE amt IS NOT NULL
  UNION ALL SELECT amt, 2 FROM p2 WHERE amt IS NOT NULL
) ORDER BY pri LIMIT 1
"""
def get_value(cik, eid, pk):
    r = con.execute(SQL_VAL, [cik, SEP, ISSUED, cik, eid, pk, pk, pk]).fetchone()
    return r[0] if r and r[0] is not None else None


# 회사별 행값 계산
peer_values = {}      # (cik, label) -> value
peer_eids_used = {}   # (cik, label) -> list of eids actually used

for cik, name in PEERS:
    role_set = role_eids(cik) if cik != "00112332" else set()
    for label, pk, _ in ROWS:
        key = (cik, label)
        if cik == "00112332":
            peer_values[key] = MIRAE_USER[(label, pk)]
            peer_eids_used[key] = ["(미래에셋 PDF disclosure)"]
            continue
        spec = ELEMENT_PATTERNS[label]
        period = spec["period"]
        total = None
        used = []
        sum_mode = spec.get("sum_mode", "first_nonnull")  # default: 첫 non-null 사용
        if sum_mode == "sum_all":
            # 모두 합산 (회사가 동일 disclosure 항목을 여러 element 로 분해)
            for eid in spec["primary_eids"]:
                v = get_value(cik, eid, period)
                if v is not None:
                    total = (total or 0) + v
                    used.append(eid.replace("ifrs-full_", ""))
        else:
            # 첫 non-null (대안 element 들 중 한 회사가 사용하는 것)
            for eid in spec["primary_eids"]:
                v = get_value(cik, eid, period)
                if v is not None:
                    total = v
                    used.append(eid.replace("ifrs-full_", ""))
                    break
        # primary 모두 안되면 fallback
        if total is None:
            for eid in spec.get("fallback_eids", []):
                v = get_value(cik, eid, period)
                if v is not None:
                    total = v
                    used.append(eid.replace("ifrs-full_", "") + " (fallback)")
                    break
        peer_values[key] = total
        peer_eids_used[key] = used if used else ["(미공시)"]

# 잔차
def residual(cik):
    op = peer_values.get((cik, "부채인 보험계약의 기초 장부금액")) or 0
    cl = peer_values.get((cik, "부채인 보험계약의 기말 장부금액")) or 0
    dur = 0
    for label, pk, _ in ROWS:
        if pk == "duration":
            v = peer_values.get((cik, label))
            if v is not None: dur += v
    return op + dur - cl

# === Excel 출력 ===
wb = Workbook()
ws = wb.active
ws.title = "BEL 변동 disclosure"

ws.append(["§4-1A 보험계약부채 변동표 — 동업사 cross-tab (별도, 발행, 억원)"])
ws.cell(row=1, column=1).font = Font(bold=True, size=14)
ws.append(["기준: 미래에셋 사업보고서 「21-1. 보험계약부채 변동분의 차이조정 공시」 13행 구조"])
ws.cell(row=2, column=1).font = Font(italic=True, color="666666", size=10)
ws.append(["미래에셋 = 사업보고서 PDF disclosure · 7개사 = XBRL DI817105 ifrs-full 표준 element 매칭 (Sep×Issued, top-level, 비-Component·비-위험관리)"])
ws.cell(row=3, column=1).font = Font(italic=True, color="C62828", size=10)
ws.append([])

hdrs = ["#", "disclosure 행 라벨", "기간"] + [n for _, n in PEERS]
ws.append(hdrs)
hr = ws.max_row
for c in range(1, len(hdrs)+1):
    cell = ws.cell(row=hr, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1E3A5F")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

opening_fill = PatternFill("solid", fgColor="FFF4CE")
closing_fill = PatternFill("solid", fgColor="FFE0B2")
revenue_fill = PatternFill("solid", fgColor="E3F2FD")
expense_fill = PatternFill("solid", fgColor="FFF3E0")
cf_fill      = PatternFill("solid", fgColor="E0F7FA")
fin_fill     = PatternFill("solid", fgColor="F3E5F5")
other_fill   = PatternFill("solid", fgColor="FAFAFA")
border = Border(left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"))

def fill_for(label, pk):
    if pk == "opening": return opening_fill
    if pk == "closing": return closing_fill
    if "보험수익" in label: return revenue_fill
    if any(w in label for w in ["발생한 보험금", "보험취득현금흐름 상각", "발생사고요소", "손실부담"]): return expense_fill
    if any(w in label for w in ["수취한", "지급", "투자요소"]): return cf_fill
    if "보험금융손익" in label: return fin_fill
    return other_fill

PK_LABEL = {"opening": "기초", "duration": "변동", "closing": "기말"}
for i, (label, pk, _) in enumerate(ROWS, 1):
    row = [i, label, PK_LABEL[pk]]
    for cik, _ in PEERS:
        v = peer_values.get((cik, label))
        row.append(round(v / 1e8, 0) if v is not None else None)
    ws.append(row)
    r = ws.max_row
    f = fill_for(label, pk)
    for c in range(1, len(hdrs)+1):
        ws.cell(row=r, column=c).fill = f
        ws.cell(row=r, column=c).border = border
    for c in range(4, len(hdrs)+1):
        cell = ws.cell(row=r, column=c)
        if cell.value is None:
            cell.value = "미공시"
            cell.font = Font(italic=True, color="999999", size=9)
            cell.alignment = Alignment(horizontal="center")
        else:
            cell.number_format = '#,##0;-#,##0;"–"'

# 잔차 행
ws.append([])
ws.append(["검증", "잔차 (기초 + Σ변동 − 기말, 억원)", "", ""] + [""]*len(PEERS))
sr = ws.max_row
for i, (cik, _) in enumerate(PEERS, start=4):
    res = residual(cik)
    ws.cell(row=sr, column=i).value = round(res / 1e8, 0)
    ws.cell(row=sr, column=i).number_format = '#,##0;-#,##0;"– (정합)"'
    op = peer_values.get((cik, "부채인 보험계약의 기초 장부금액")) or 0
    if abs(res) < max(abs(op) * 0.01, 100_000_000_000):
        ws.cell(row=sr, column=i).font = Font(bold=True, color="2E7D32")
    else:
        ws.cell(row=sr, column=i).font = Font(bold=True, color="C62828")
for c in range(1, len(hdrs)+1):
    ws.cell(row=sr, column=c).fill = PatternFill("solid", fgColor="E8F5E9")
    ws.cell(row=sr, column=c).border = border
ws.cell(row=sr, column=1).font = Font(bold=True)
ws.cell(row=sr, column=2).font = Font(bold=True)

# Caveat
ws.append([])
cr = ws.max_row + 1
ws.cell(row=cr, column=1).value = "주석"
ws.cell(row=cr, column=1).font = Font(bold=True, color="C62828")
ws.cell(row=cr, column=2).value = (
    "회사별 disclosure 분해와 ifrs-full element 사용이 달라 매칭 한계 존재. "
    "특히 현대해상: 별도 BS 보험계약부채를 Sep×Issued 단독 컨텍스트로 보고하지 않고 PAA/TypesOfContracts axis 분해로만 보고 → top-level 잔액 추출 불가 (PAA member 단독값만 표시됨, 비-PAA 누락). "
    "삼성생명·동양생명: 일부 행이 InsuranceServiceExpenses 집계 element 만 보고 (sub-decomposition 없음). "
    "기타증감: 한화 일부, 삼성생명, DB·한손 entity 확장 element 사용 → 미공시 표기. "
    "잔차 |%| < 1% 면 정합 OK 로 표시."
)
ws.cell(row=cr, column=2).alignment = Alignment(wrap_text=True, vertical="top")
ws.cell(row=cr, column=2).font = Font(italic=True, color="666666", size=9)
ws.merge_cells(start_row=cr, start_column=2, end_row=cr, end_column=len(hdrs))
ws.row_dimensions[cr].height = 45

widths = [4, 40, 6] + [13]*len(PEERS)
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[hr].height = 28
ws.freeze_panes = f"D{hr+1}"

# 진단 시트
diag = wb.create_sheet(title="element 매핑")
diag.append(["회사별 disclosure 행 → 사용된 element_id (Sep×Issued top-level 값 합산)"])
diag.cell(row=1, column=1).font = Font(bold=True, size=12)
diag.append(["미공시 = primary·fallback 모두 매칭 안됨 (회사가 해당 항목을 entity 확장으로 보고했거나 분해 단위 다름)"])
diag.cell(row=2, column=1).font = Font(italic=True, color="666666", size=9)
diag.append([])
diag.append(["disclosure 행"] + [n for _, n in PEERS])
hr2 = diag.max_row
for c in range(1, 2 + len(PEERS)):
    cell = diag.cell(row=hr2, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1E3A5F")
    cell.alignment = Alignment(horizontal="center", wrap_text=True)
for label, pk, _ in ROWS:
    row = [label]
    for cik, _ in PEERS:
        eids = peer_eids_used.get((cik, label), [])
        row.append("\n".join(eids))
    diag.append(row)
    r = diag.max_row
    for c in range(1, 2 + len(PEERS)):
        diag.cell(row=r, column=c).border = border
        diag.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)
for i, w in enumerate([35] + [40]*len(PEERS), 1):
    diag.column_dimensions[get_column_letter(i)].width = w
diag.freeze_panes = f"B{hr2+1}"

# 매핑 규칙 시트
rule = wb.create_sheet(title="매핑 규칙")
rule.append(["13행 라벨 → 캐노니컬 ifrs-full element_id (primary / fallback)"])
rule.cell(row=1, column=1).font = Font(bold=True, size=12)
rule.append([])
rule.append(["#", "disclosure 행 라벨", "기간", "primary element_ids", "fallback element_ids"])
hr3 = rule.max_row
for c in range(1, 6):
    cell = rule.cell(row=hr3, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1E3A5F")
for i, (label, pk, _) in enumerate(ROWS, 1):
    spec = ELEMENT_PATTERNS.get(label, {})
    prim = "\n".join(e.replace("ifrs-full_", "") for e in spec.get("primary_eids", []))
    fb = "\n".join(e.replace("ifrs-full_", "") for e in spec.get("fallback_eids", []))
    rule.append([i, label, PK_LABEL[pk], prim, fb])
    r = rule.max_row
    rule.cell(row=r, column=4).alignment = Alignment(wrap_text=True, vertical="top")
    rule.cell(row=r, column=5).alignment = Alignment(wrap_text=True, vertical="top")
for i, w in enumerate([4, 38, 6, 70, 55], 1):
    rule.column_dimensions[get_column_letter(i)].width = w

out = Path("outputs/bel_disclosure_filled.xlsx")
wb.save(out)
print(f"wrote {out}")

# 잔차 정합 보고서
print("\n=== 회사별 잔차 (기초 + Σ변동 − 기말, 억원 / 기초 대비 %) ===")
for cik, name in PEERS:
    res = residual(cik)
    op = peer_values.get((cik, "부채인 보험계약의 기초 장부금액")) or 0
    pct = res / op * 100 if op else 0
    flag = "OK" if abs(res) < max(abs(op) * 0.01, 100_000_000_000) else "잔차큼"
    print(f"  {name:<12s} 기초={op/1e8:>+12,.0f}억  잔차={res/1e8:>+10,.0f}억 ({pct:+.2f}%)  [{flag}]")

# 행별 누락 보고
print("\n=== 회사별 미공시 행 ===")
for cik, name in PEERS:
    if cik == "00112332": continue
    miss = [label for label, pk, _ in ROWS
            if peer_eids_used.get((cik, label), [""])[0] == "(미공시)"]
    if miss:
        print(f"  {name}:")
        for m in miss:
            print(f"    - {m}")
