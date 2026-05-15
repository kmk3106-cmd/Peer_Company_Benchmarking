"""§4-1A 보험계약부채 변동표 — 라벨 기반 동업사 cross-tab.

설계:
  - 회사별로 DI817105 role 안의 element 각각에 대해
    PERIOD_KIND × element_id 별 최소 axis depth fact 만 사용 (top-level 합계)
  - 라벨 = element 의 terseLabel (있으면) > default label
  - Cross-tab row key = (depth, label) — 같은 disclosure 라벨끼리 정렬
  - 회사마다 element_id 가 다르게 매핑되어도 라벨이 같으면 같은 row
"""
from __future__ import annotations
import duckdb
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

PEERS = [
    ("00112332", "미래에셋생명"),
    ("00126256", "삼성생명"),
    ("00113058", "한화생명"),
    ("00117267", "동양생명"),
    ("00139214", "삼성화재"),
    ("00164973", "현대해상"),
    ("00159102", "DB손해보험"),
    ("00135917", "한화손해보험"),
]
SEP = "ifrs-full_SeparateMember"
ISSUED = "ifrs-full_InsuranceContractsIssuedMember"

con = duckdb.connect("data/db/benchmark.duckdb", read_only=True)

# 회사별 (element, period_kind) 의 최소 axis depth fact 만 사용 → 라벨 기반 cross-tab
SQL = """
WITH mv_elems AS (
  -- DI817 변동표 role 소속 element + 트리 depth
  SELECT ELEMENT_ID,
         MIN(CASE WHEN PARENT_ELEMENT_ID IS NULL THEN 0 ELSE 1 END) AS is_leaf
  FROM pre_insurers
  WHERE CIK=? AND REPORT_DATE='20251231' AND ROLE_ID LIKE 'dart_2024-06-30_role-DI817%'
  GROUP BY ELEMENT_ID
),
lbl AS (
  -- terseLabel 우선, 없으면 default label
  SELECT ELMT_ID,
         MAX(CASE WHEN LABEL_ROLE_URI LIKE '%terseLabel%' THEN LABEL END) AS terse,
         MAX(CASE WHEN LABEL_ROLE_URI='http://www.xbrl.org/2003/role/label' THEN LABEL END) AS dflt
  FROM lab_insurers WHERE CIK=? AND REPORT_DATE='20251231' AND LANG='ko'
  GROUP BY ELMT_ID
),
ctxs AS (
  SELECT CONTEXT_ID, COUNT(*) AS n_ax,
         MAX(PERIOD_INSTANT) AS p_inst,
         MAX(PERIOD_START_DATE) AS p_start,
         MAX(PERIOD_END_DATE) AS p_end
  FROM cntxt_insurers WHERE CIK=? AND REPORT_DATE='20251231'
  GROUP BY CONTEXT_ID
  HAVING BOOL_OR(MEMBER_ELEMENT_ID = ?) AND BOOL_OR(MEMBER_ELEMENT_ID = ?)
     AND NOT BOOL_OR(MEMBER_ELEMENT_ID LIKE '%OfDisclosureOfNatureAndExtentOfRisks%')
     AND NOT BOOL_OR(AXIS_ELEMENT_ID = 'ifrs-full_InsuranceContractsByComponentsAxis')
),
fact AS (
  SELECT v.ELEMENT_ID,
         CASE WHEN c.p_inst='2024-12-31' THEN 'opening'
              WHEN c.p_inst='2025-12-31' THEN 'closing'
              WHEN c.p_start='2025-01-01' AND c.p_end='2025-12-31' THEN 'duration'
              ELSE NULL END AS pk,
         c.n_ax,
         v.amount_krw
  FROM val_insurers v JOIN ctxs c USING(CONTEXT_ID)
  JOIN mv_elems me ON me.ELEMENT_ID=v.ELEMENT_ID
  WHERE v.CIK=? AND v.REPORT_DATE='20251231' AND v.amount_krw IS NOT NULL
    AND (c.p_inst IN ('2024-12-31','2025-12-31')
         OR (c.p_start='2025-01-01' AND c.p_end='2025-12-31'))
),
top_only AS (
  -- 각 (element, period) 의 최소 axis depth 만
  SELECT f.ELEMENT_ID, f.pk, MIN(f.n_ax) AS min_ax
  FROM fact f WHERE f.pk IS NOT NULL GROUP BY f.ELEMENT_ID, f.pk
)
SELECT
  f.pk, f.ELEMENT_ID,
  COALESCE(lbl.terse, lbl.dflt, '') AS label,
  SUM(f.amount_krw)/1e8 AS amt_eok,
  COUNT(*) AS n
FROM fact f
JOIN top_only t USING(ELEMENT_ID, pk)
LEFT JOIN lbl ON lbl.ELMT_ID = f.ELEMENT_ID
WHERE f.n_ax = t.min_ax AND f.pk IS NOT NULL
GROUP BY f.pk, f.ELEMENT_ID, lbl.terse, lbl.dflt
HAVING SUM(f.amount_krw) IS NOT NULL
"""

# 데이터 수집: peer_facts[(period, label)] = {cik: amt}
from collections import defaultdict
peer_facts = defaultdict(dict)         # (pk, label) -> {cik: amt}
peer_eid = defaultdict(dict)           # (pk, label) -> {cik: element_id} (which element used)
peer_total_lines = {}                  # cik -> count of distinct labels
for cik, _ in PEERS:
    rows = con.execute(SQL, [cik, cik, cik, SEP, ISSUED, cik]).fetchall()
    n = 0
    for pk, eid, label, amt, fact_n in rows:
        if not label:
            label = eid.replace("ifrs-full_", "").replace("dart_", "")[:50] + " (라벨없음)"
        peer_facts[(pk, label)][cik] = amt
        peer_eid[(pk, label)][cik] = eid
        n += 1
    peer_total_lines[cik] = n

# Period order
def pk_order(pk):
    return {"opening": 0, "duration": 1, "closing": 2}.get(pk, 9)

# 행 정렬: period kind → 자사(미래에셋) 절대값 큰 순
all_keys = list(peer_facts.keys())
def sort_key(key):
    pk, lbl = key
    self_amt = peer_facts[key].get("00112332", 0) or 0
    return (pk_order(pk), -abs(self_amt), lbl)
sorted_keys = sorted(all_keys, key=sort_key)

# Excel 작성
wb = Workbook()
ws = wb.active
ws.title = "라벨 기반 cross-tab"

ws.append(["§4-1A 보험계약부채 변동표 — 라벨 기반 동업사 cross-tab (별도, 발행, 억원)"])
ws.cell(row=1, column=1).font = Font(bold=True, size=14)
ws.append(["row 키 = 각 회사가 disclosure 에서 사용한 한글 라벨 (terseLabel 우선) · element_id 회사별 다를 수 있음"])
ws.cell(row=2, column=1).font = Font(italic=True, color="666666", size=10)
ws.append(["각 회사 element 별로 최소 axis depth fact 만 합산 (top-level 값)"])
ws.cell(row=3, column=1).font = Font(italic=True, color="666666", size=10)
ws.append([])

hdrs = ["기간", "row 라벨 (사업보고서 표기)"] + [n for _, n in PEERS]
ws.append(hdrs)
hr = ws.max_row
for c in range(1, len(hdrs)+1):
    cell = ws.cell(row=hr, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1E3A5F")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

PK_FILL = {"opening": "FFF4CE", "duration": "E8F0F8", "closing": "FFE0B2"}
PK_LABEL = {"opening": "[기초]", "duration": "[변동]", "closing": "[기말]"}
border = Border(left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"))

prev_pk = None
for pk, label in sorted_keys:
    row_data = [PK_LABEL[pk], label]
    facts = peer_facts[(pk, label)]
    for cik, _ in PEERS:
        v = facts.get(cik)
        row_data.append(round(v, 0) if v is not None else None)
    ws.append(row_data)
    r = ws.max_row
    fill = PatternFill("solid", fgColor=PK_FILL.get(pk, "FFFFFF"))
    for c in range(1, len(hdrs)+1):
        ws.cell(row=r, column=c).border = border
        ws.cell(row=r, column=c).fill = fill
    for c in range(3, len(hdrs)+1):
        ws.cell(row=r, column=c).number_format = '#,##0;-#,##0;"–"'
    if pk != prev_pk:
        ws.cell(row=r, column=1).font = Font(bold=True, color="1E3A5F")
    prev_pk = pk

widths = [10, 50] + [13]*len(PEERS)
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[hr].height = 32
ws.freeze_panes = f"C{hr+1}"

# 진단 시트: 회사별 element_id 매핑 차이 검사
diag = wb.create_sheet(title="element 매핑 진단")
diag.append(["같은 라벨인데 회사마다 다른 element_id 를 사용한 경우 발견"])
diag.cell(row=1, column=1).font = Font(bold=True, size=12)
diag.append([])
diag.append(["기간", "라벨", "회사", "element_id (회사별)"])
hr2 = diag.max_row
for c in range(1, 5):
    cell = diag.cell(row=hr2, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1E3A5F")
    cell.alignment = Alignment(horizontal="center")

for pk, label in sorted_keys:
    eids = peer_eid[(pk, label)]
    if len(set(eids.values())) <= 1: continue  # 일관 매핑
    for cik, peer in PEERS:
        eid = eids.get(cik)
        if eid:
            diag.append([PK_LABEL[pk], label, peer, eid.replace("ifrs-full_", "")[:90]])

for i, w in enumerate([10, 45, 14, 95], 1):
    diag.column_dimensions[get_column_letter(i)].width = w
diag.freeze_panes = f"A{hr2+1}"

out = Path("outputs/bel_label_crosstab.xlsx")
wb.save(out)
print(f"wrote {out}")
print(f"sheets: {wb.sheetnames}")
print(f"\n총 distinct (기간, 라벨) row 수: {len(sorted_keys)}")
print(f"회사별 fact 수:")
for cik, name in PEERS:
    print(f"  {name:<14s}  {peer_total_lines[cik]} facts")
