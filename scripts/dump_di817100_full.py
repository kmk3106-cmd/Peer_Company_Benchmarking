"""DI817100 role 전체(메인 + 모든 sub-section) 자료 평면 dump.

미래에셋생명(00112332) 기준 — DI817100, DI817100a~z 모든 sub-section 의
presentation tree를 순서대로 펼치고 각 element + 값(있으면) 표시.

출력: report/di817100_full_dump.xlsx
  - 시트 0: 목차 — sub-section 리스트 + 라벨 + element 수
  - 시트 1~12: 각 sub-section의 presentation tree (들여쓰기 + 값)
"""
from __future__ import annotations

from pathlib import Path

import duckdb
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

CIK = "00112332"
COMPANY = "미래에셋생명"

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(italic=True, size=9, color="666666")
ABSTRACT_FILL = PatternFill("solid", fgColor="EFEFEF")
AXIS_FILL = PatternFill("solid", fgColor="DCE6F1")
TABLE_FILL = PatternFill("solid", fgColor="FCE4D6")
ENTITY_FILL = PatternFill("solid", fgColor="FFF4E0")
HAS_VAL_FILL = PatternFill("solid", fgColor="E2EFDA")
THIN = Side(border_style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _short(eid: str) -> str:
    return (
        eid.replace("ifrs-full_", "")
        .replace("dart-gcd_", "")
        .replace("dart_2024-06-30_", "")
        .replace("dart_", "d:")
        .replace(f"entity{CIK}_", "[자사]")
    )


def fetch_sub_roles(con: duckdb.DuckDBPyConnection) -> list[tuple[str, str]]:
    """미래에셋 DI817100 + 모든 sub-section role 리스트."""
    rows = con.execute("""
    SELECT ROLE_ID, ROLE_NM_KO
    FROM role_insurers
    WHERE CIK = ? AND ROLE_ID LIKE '%DI817100%'
    ORDER BY ROLE_ID
    """, [CIK]).fetchall()
    return rows


def fetch_role_tree(con: duckdb.DuckDBPyConnection, role_id: str) -> pd.DataFrame:
    """한 role의 presentation tree + 값 + 라벨."""
    sql = """
    WITH pre AS (
      SELECT ELEMENT_ID, PARENT_ELEMENT_ID,
             TRY_CAST("ORDER" AS DOUBLE) AS ord
      FROM pre_insurers
      WHERE CIK = ? AND ROLE_ID = ?
    ),
    labels AS (
      SELECT ELMT_ID, MAX(LABEL) AS ko_label
      FROM lab_insurers
      WHERE CIK = ? AND LANG = 'ko'
        AND LABEL_ROLE_URI = 'http://www.xbrl.org/2003/role/label'
      GROUP BY ELMT_ID
    ),
    val_summary AS (
      SELECT ELEMENT_ID,
             COUNT(DISTINCT CONTEXT_ID) AS n_ctx,
             MAX(ABS(amount_krw)) AS max_abs,
             MIN(amount_krw) AS min_signed,
             MAX(amount_krw) AS max_signed
      FROM val_insurers
      WHERE CIK = ? AND amount_krw IS NOT NULL
      GROUP BY ELEMENT_ID
    )
    SELECT
      p.ELEMENT_ID, p.PARENT_ELEMENT_ID, p.ord,
      em.ABSTRACT, em.PERIOD_TYPE, em.BALANCE, em.SUBSTITUTION_GROUP,
      l.ko_label,
      vs.n_ctx, vs.max_abs, vs.min_signed, vs.max_signed
    FROM pre p
    LEFT JOIN elmt_raw em ON em.ELEMENT_ID = p.ELEMENT_ID
    LEFT JOIN labels l ON l.ELMT_ID = p.ELEMENT_ID
    LEFT JOIN val_summary vs ON vs.ELEMENT_ID = p.ELEMENT_ID
    """
    df = con.execute(sql, [CIK, role_id, CIK, CIK]).df()
    return df


def order_tree(df: pd.DataFrame) -> pd.DataFrame:
    """presentation tree를 depth-first 순서로 정렬 + depth 컬럼 추가."""
    if df.empty:
        return df
    children_of: dict[str | None, list] = {}
    for _, row in df.iterrows():
        children_of.setdefault(row["PARENT_ELEMENT_ID"], []).append(
            (row["ord"] if pd.notna(row["ord"]) else 0, row["ELEMENT_ID"])
        )
    for k in children_of:
        children_of[k].sort()

    visited: set[str] = set()
    ordered: list[tuple[int, str]] = []

    def walk(parent, depth):
        for _, eid in children_of.get(parent, []):
            if eid in visited:
                continue
            visited.add(eid)
            ordered.append((depth, eid))
            walk(eid, depth + 1)

    # root nodes: those whose parent is None or whose parent is not in df
    df_elem_set = set(df["ELEMENT_ID"])
    roots_keys = [
        p for p in children_of
        if p is None or (isinstance(p, str) and p not in df_elem_set)
    ]
    for k in roots_keys:
        walk(k, 0)
    # Pick up any orphans (parent in df but never traversed)
    for _, row in df.iterrows():
        if row["ELEMENT_ID"] not in visited:
            ordered.append((0, row["ELEMENT_ID"]))
            visited.add(row["ELEMENT_ID"])
            walk(row["ELEMENT_ID"], 1)

    order_df = pd.DataFrame(ordered, columns=["depth", "ELEMENT_ID"])
    order_df["seq"] = range(len(order_df))
    out = order_df.merge(df, on="ELEMENT_ID", how="left")
    # Dedup ELEMENT_ID keeping first occurrence
    out = out.drop_duplicates(subset=["seq"]).sort_values("seq").reset_index(drop=True)
    return out


def classify(row) -> str:
    sub = (row.get("SUBSTITUTION_GROUP") or "")
    if "hypercubeItem" in sub:
        return "Table"
    if "dimensionItem" in sub:
        return "Axis"
    if str(row.get("ABSTRACT")).lower() == "true":
        return "Abstract"
    return "LineItem"


def write_role_sheet(ws, role_id: str, role_label: str, df: pd.DataFrame) -> None:
    sub_code = role_id.split("role-")[-1] if "role-" in role_id else role_id
    ws.cell(row=1, column=1, value=f"{sub_code} — {role_label}").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"{COMPANY} (CIK {CIK}) · presentation 트리 순서 · {len(df)} elements").font = SUBTITLE_FONT

    headers = ["#", "depth", "element", "ko_label", "분류", "period", "bal", "ctx#", "min(억)", "max(억)"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER

    for ri, row in df.iterrows():
        depth = int(row.get("depth") or 0)
        cls = classify(row)
        eid = row["ELEMENT_ID"]
        is_entity = isinstance(eid, str) and eid.startswith("entity")
        indent = "  " * depth
        cells = [
            int(row["seq"]) + 1,
            depth,
            indent + _short(eid),
            row.get("ko_label") or "",
            cls,
            row.get("PERIOD_TYPE") or "",
            row.get("BALANCE") or "",
            int(row["n_ctx"]) if pd.notna(row.get("n_ctx")) else 0,
            round(row["min_signed"] / 1e8, 1) if pd.notna(row.get("min_signed")) else None,
            round(row["max_signed"] / 1e8, 1) if pd.notna(row.get("max_signed")) else None,
        ]
        # Apply fill by classification
        if cls == "Table":
            fill = TABLE_FILL
        elif cls == "Axis":
            fill = AXIS_FILL
        elif cls == "Abstract":
            fill = ABSTRACT_FILL
        else:
            fill = HAS_VAL_FILL if cells[7] > 0 else None

        for ci, v in enumerate(cells, start=1):
            c = ws.cell(row=5 + ri, column=ci, value=v)
            c.border = BORDER
            if is_entity and ci == 3:
                c.fill = ENTITY_FILL
            elif fill:
                c.fill = fill
            if ci >= 8 and isinstance(v, (int, float)):
                c.alignment = Alignment(horizontal="right")
                if isinstance(v, float):
                    c.number_format = "#,##0.0;(#,##0.0);-"

    widths = [5, 6, 75, 42, 10, 10, 6, 6, 12, 12]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "C5"


def write_toc(ws, sub_roles: list[tuple[str, str]], role_dfs: dict[str, pd.DataFrame]) -> None:
    ws.cell(row=1, column=1, value=f"{COMPANY} DI817100 전체 자료 — 목차").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"메인 role + 모든 sub-section ({len(sub_roles)}개)").font = SUBTITLE_FONT

    headers = ["#", "role_id", "한국어 라벨", "element 수", "값보유 수", "axis", "abstract", "lineitem"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER

    for i, (rid, rlabel) in enumerate(sub_roles, start=1):
        df = role_dfs[rid]
        if df.empty:
            n_total = n_val = n_axis = n_abs = n_li = 0
        else:
            n_total = len(df)
            n_val = (df["n_ctx"].fillna(0) > 0).sum()
            n_axis = df.apply(lambda r: "dimensionItem" in (r.get("SUBSTITUTION_GROUP") or ""), axis=1).sum()
            is_abstract = df["ABSTRACT"].astype(str).str.lower() == "true"
            n_abs = is_abstract.sum() - n_axis
            n_li = n_total - is_abstract.sum()
        sub_code = rid.split("role-")[-1]
        cells = [i, sub_code, rlabel or "", n_total, int(n_val), int(n_axis), int(n_abs), int(n_li)]
        for ci, v in enumerate(cells, start=1):
            c = ws.cell(row=4 + i, column=ci, value=v)
            c.border = BORDER
            if ci >= 4:
                c.alignment = Alignment(horizontal="right")
        # main role 강조
        if rid.endswith("DI817100"):
            for ci in range(1, len(headers) + 1):
                ws.cell(row=4 + i, column=ci).font = Font(bold=True)

    widths = [5, 30, 50, 12, 12, 8, 12, 12]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def main() -> int:
    con = duckdb.connect("data/db/benchmark.duckdb", read_only=True)
    sub_roles = fetch_sub_roles(con)
    print(f"DI817100 + sub-section: {len(sub_roles)}개 발견")

    role_dfs: dict[str, pd.DataFrame] = {}
    for rid, rlabel in sub_roles:
        df = fetch_role_tree(con, rid)
        df = order_tree(df)
        role_dfs[rid] = df
        sub_code = rid.split("role-")[-1]
        print(f"  {sub_code:30s} → {len(df)} elements")

    con.close()

    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    ws_toc = wb.create_sheet("목차")
    write_toc(ws_toc, sub_roles, role_dfs)

    # Each role as separate sheet (sheet name limited to 30 chars)
    for rid, rlabel in sub_roles:
        sub_code = rid.split("role-")[-1]
        sheet_name = sub_code[:30]
        ws = wb.create_sheet(sheet_name)
        write_role_sheet(ws, rid, rlabel, role_dfs[rid])

    out = Path("report/di817100_full_dump.xlsx")
    wb.save(out)
    print(f"\n✓ wrote {out} ({out.stat().st_size / 1024:.1f} KB, {len(sub_roles)+1} sheets)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
