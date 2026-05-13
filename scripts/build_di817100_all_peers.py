"""IFRS17 §103 보험계약부채(자산) 변동분 차이조정 — 9개 보험사 전체 추출.

회사별로 한 시트 + 요약 시트 1개. 단위: 억원.

대상: ifrs17_detailed peer group (IFRS17 풍부 9개사).
기준: 별도(separate) — CLAUDE.md §7.

출력: report/IFRS17_보험계약부채변동_9개사.xlsx
"""

from __future__ import annotations

import logging
from pathlib import Path

import duckdb
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from peer_benchmarking.analysis.movement_reconciliation import (
    reconciliation_103_all_peers,
)
from peer_benchmarking.domain import peer_groups

log = logging.getLogger(__name__)

# ─── 표시 스타일 ─────────────────────────────────────────────

TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(italic=True, size=10, color="666666")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
BAL_FILL = PatternFill("solid", fgColor="DCE6F1")  # 시작/종료 잔액
SECTION_FILL = PatternFill("solid", fgColor="F2F2F2")  # 섹션 헤더
THIN = Side(border_style="thin", color="CCCCCC")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 라인 그룹화 (사업보고서 §103 표 형식과 유사)
SECTION_GROUPS = [
    ("[시작]", ["open_balance"]),
    ("미래 서비스 변동", ["initial_recognition", "estimate_change_adjusting_csm",
                  "estimate_change_not_adjusting_csm"]),
    ("현행 서비스 변동", ["csm_recognised_in_revenue", "ra_release", "experience_adjustment"]),
    ("과거 서비스 변동", ["past_service_change"]),
    ("보험계약 현금흐름", ["premium_received", "claims_paid", "acquisition_cashflows"]),
    ("보험금융손익", ["finance_pl"]),
    ("기타", ["other_changes"]),
    ("[종료]", ["close_balance"]),
]


def _eok(val):
    """원 → 억원 변환 (반올림 1자리)."""
    if val is None or pd.isna(val):
        return None
    return round(val / 1e8, 1)


def _write_company_sheet(ws, df: pd.DataFrame, name: str) -> None:
    """한 회사의 §103 표를 한 시트로 쓰기."""
    ws.cell(row=1, column=1, value=f"{name} — IFRS17 §103 보험계약부채(자산) 변동분 차이조정").font = TITLE_FONT
    ws.cell(row=2, column=1, value="별도 기준 · 2025년 사업연도 · 단위: 억원 · 발행보험 합계").font = SUBTITLE_FONT

    # Header (row 4)
    headers = ["구분", "라인 아이템", "BEL", "RA", "CSM", "합계"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = BORDER

    # Rows by section
    df_indexed = df.set_index("line_key")
    r = 5
    for section_name, line_keys in SECTION_GROUPS:
        is_balance = section_name in ("[시작]", "[종료]")
        for i, lkey in enumerate(line_keys):
            if lkey not in df_indexed.index:
                continue
            row = df_indexed.loc[lkey]
            section_display = section_name if i == 0 else ""
            cells = [
                section_display,
                row["ko_label"],
                _eok(row["BEL"]),
                _eok(row["RA"]),
                _eok(row["CSM"]),
                _eok(row["합계"]),
            ]
            fill = BAL_FILL if is_balance else None
            for ci, v in enumerate(cells, start=1):
                c = ws.cell(row=r, column=ci, value=v)
                c.border = BORDER
                if ci >= 3:
                    c.alignment = Alignment(horizontal="right")
                    c.number_format = "#,##0;(#,##0);-"
                if fill:
                    c.fill = fill
                if is_balance or section_display:
                    c.font = Font(bold=True)
            r += 1

    # column widths
    widths = [16, 36, 14, 12, 12, 14]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def _write_summary_sheet(ws, results: dict[str, pd.DataFrame]) -> None:
    """요약: 회사 × (시작잔액 / 종료잔액 / 총 변동) — BEL+RA+CSM 합계 (억원)."""
    ws.cell(row=1, column=1, value="요약: 9개사 보험계약부채(자산) 변동분 (별도, 억원)").font = TITLE_FONT
    ws.cell(row=2, column=1, value="시작잔액·종료잔액은 합계 column 기준. 총변동 = 종료 - 시작.").font = SUBTITLE_FONT

    # Header
    headers = ["회사", "섹터", "시작잔액", "BEL", "RA", "CSM", "종료잔액", "BEL", "RA", "CSM", "총변동"]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER

    companies = peer_groups.load_companies()
    r = 5
    for cik, df in results.items():
        if df.empty:
            continue
        co = companies.get(cik)
        if not co:
            continue
        df_idx = df.set_index("line_key")
        open_row = df_idx.loc["open_balance"] if "open_balance" in df_idx.index else None
        close_row = df_idx.loc["close_balance"] if "close_balance" in df_idx.index else None
        open_tot = _eok(open_row["합계"]) if open_row is not None else None
        close_tot = _eok(close_row["합계"]) if close_row is not None else None
        delta = (close_tot - open_tot) if (open_tot is not None and close_tot is not None) else None

        sector_ko = {"life": "생보", "non_life": "손보"}.get(co.sector, co.sector)
        star = " ★" if co.is_self else ""
        cells = [
            co.name_ko + star, sector_ko,
            open_tot,
            _eok(open_row["BEL"]) if open_row is not None else None,
            _eok(open_row["RA"]) if open_row is not None else None,
            _eok(open_row["CSM"]) if open_row is not None else None,
            close_tot,
            _eok(close_row["BEL"]) if close_row is not None else None,
            _eok(close_row["RA"]) if close_row is not None else None,
            _eok(close_row["CSM"]) if close_row is not None else None,
            delta,
        ]
        for ci, v in enumerate(cells, start=1):
            c = ws.cell(row=r, column=ci, value=v)
            c.border = BORDER
            if ci >= 3:
                c.alignment = Alignment(horizontal="right")
                c.number_format = "#,##0;(#,##0);-"
            if co.is_self:
                c.fill = PatternFill("solid", fgColor="FFF59D")
        r += 1

    widths = [16, 8, 14, 12, 10, 12, 14, 12, 10, 12, 14]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def main() -> int:
    logging.basicConfig(level=logging.INFO, format="%(asctime)s %(levelname)s %(message)s")
    db_path = Path("data/db/benchmark.duckdb")
    out = Path("report/IFRS17_보험계약부채변동_9개사.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)

    con = duckdb.connect(str(db_path), read_only=True)
    log.info("9개사 §103 reconciliation 추출 시작 ...")
    results = reconciliation_103_all_peers(con, peer_group="ifrs17_detailed")
    con.close()

    log.info("결과 회사 수: %d", len(results))
    for cik, df in results.items():
        non_null = df[df["합계"].notna()]
        log.info("  %s : %d/%d lines populated",
                 peer_groups.name_of(cik), len(non_null), len(df))

    log.info("writing Excel ...")
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    # 1. Summary sheet
    ws_sum = wb.create_sheet("요약")
    _write_summary_sheet(ws_sum, results)

    # 2. One sheet per company (ordered: self first, then life, then non_life)
    companies = peer_groups.load_companies()
    def _order_key(item):
        cik, _ = item
        co = companies[cik]
        return (not co.is_self, co.sector != "life", co.name_ko)

    for cik, df in sorted(results.items(), key=_order_key):
        co = companies[cik]
        sheet_name = co.name_ko[:30]
        ws = wb.create_sheet(sheet_name)
        _write_company_sheet(ws, df, co.name_ko)

    wb.save(out)
    log.info("done: %s (%.1f KB)", out, out.stat().st_size / 1024)

    # CLI: 요약 출력
    print()
    print("=" * 95)
    print("9개사 보험계약부채(자산) §103 변동분 차이조정 — 요약 (별도, 억원)")
    print("=" * 95)
    print(f"  {'회사':14s} {'섹터':4s}  {'시작잔액':>12s} {'BEL_시':>10s} {'RA_시':>8s} {'CSM_시':>10s}  {'종료잔액':>12s} {'CSM_종':>10s}  {'총변동':>10s}")
    print("-" * 95)
    for cik, df in sorted(results.items(), key=_order_key):
        co = companies[cik]
        df_idx = df.set_index("line_key")
        if "open_balance" not in df_idx.index or "close_balance" not in df_idx.index:
            continue
        o, cl = df_idx.loc["open_balance"], df_idx.loc["close_balance"]
        o_tot, c_tot = _eok(o["합계"]), _eok(cl["합계"])
        delta = (c_tot - o_tot) if (o_tot is not None and c_tot is not None) else None

        def _fmt(v):
            return f"{v:>10,.0f}" if v is not None else "       N/A"

        sector_ko = {"life": "생보", "non_life": "손보"}.get(co.sector, co.sector)
        star = " ★" if co.is_self else "  "
        print(f"  {co.name_ko:14s} {sector_ko:4s}{star}{_fmt(o_tot):>12s} "
              f"{_fmt(_eok(o['BEL']))} {_fmt(_eok(o['RA']))} {_fmt(_eok(o['CSM']))}  "
              f"{_fmt(c_tot):>12s} {_fmt(_eok(cl['CSM']))}  {_fmt(delta)}")

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
