"""8 role × 9 회사 × element 인벤토리.

목표: 어떤 element가 어느 role 트리에 등장하고, 어느 회사가 보고하는지를 한 화면에서 보기.

대상 role (재보험 DI817200/205 제외):
  DI817100  보험계약부채(자산)의 변동 — 원수 변동표 (메인)
  DI817105  보험계약부채(자산) 잔액 — 원수 잔액표 (보조)
  DI817300  보험계약 정보 (CSM 만기) — 진단
  DI817305  보험계약 정보 잔액 — 진단
  DI818100  위험관리 — 정성 (연결)
  DI818105  위험관리 — 정성 (별도)
  DI818200  위험관리 상세 (연결)
  DI818205  위험관리 상세 (별도)

회사: ifrs17_detailed 9개사 (미래에셋·삼성생명·한화생명·동양생명·삼성화재·현대해상·DB손해·한화손해·흥국화재)

출력: report/role_inventory.xlsx
  - 시트 0: master (모든 role 합쳐서 element_id × role × 회사 사용 매트릭스)
  - 시트 1~8: role별 (element 행, 9 회사 컬럼)
  - 시트 9: summary (role × 회사 element 보유 수)
"""
from __future__ import annotations

from pathlib import Path

import duckdb
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from peer_benchmarking.domain import peer_groups

ROLES = [
    ("DI817100", "보험계약부채(자산) 변동 — 원수 변동표"),
    ("DI817105", "보험계약부채(자산) 잔액 — 원수 잔액표"),
    ("DI817300", "보험계약 정보 (CSM 만기)"),
    ("DI817305", "보험계약 정보 잔액"),
    ("DI818100", "위험관리 정성 (연결)"),
    ("DI818105", "위험관리 정성 (별도)"),
    ("DI818200", "위험관리 상세 (연결)"),
    ("DI818205", "위험관리 상세 (별도)"),
]

CIKS = (
    "00112332", "00126256", "00113058", "00117267",
    "00139214", "00164973", "00159102", "00135917", "00103176",
)

# Styling
HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)
TITLE_FONT = Font(bold=True, size=14, color="1F4E79")
SUBTITLE_FONT = Font(italic=True, size=9, color="666666")
ABSTRACT_FILL = PatternFill("solid", fgColor="EFEFEF")
ENTITY_FILL = PatternFill("solid", fgColor="FFF4E0")
SELF_FILL = PatternFill("solid", fgColor="FFF59D")
THIN = Side(border_style="thin", color="DDDDDD")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)


def _short_element(eid: str) -> str:
    return (
        eid.replace("ifrs-full_", "")
        .replace("dart-gcd_", "")
        .replace("dart_2024-06-30_", "")
        .replace("dart_", "d:")
        .replace("entity00112332_", "[mirae]")
        .replace("entity00126256_", "[samsung]")
        .replace("entity00113058_", "[hanwha]")
        .replace("entity00117267_", "[tongyang]")
        .replace("entity00139214_", "[samsungf]")
        .replace("entity00164973_", "[hyundai]")
        .replace("entity00159102_", "[db]")
        .replace("entity00135917_", "[hanwhag]")
        .replace("entity00103176_", "[heungkuk]")
    )


def fetch_role_inventory(con: duckdb.DuckDBPyConnection, role_code: str) -> pd.DataFrame:
    """모든 sub-section (DI817100, DI817100a, b, c, ...) 포함, 9개사 element 집합."""
    cik_in = ", ".join(f"'{c}'" for c in CIKS)
    role_pattern = f"%{role_code}"
    sql = f"""
    WITH all_elems AS (
      SELECT DISTINCT p.CIK, p.ELEMENT_ID, p.ROLE_ID
      FROM pre_insurers p
      WHERE p.CIK IN ({cik_in})
        AND (
          p.ROLE_ID = 'dart_2024-06-30_role-{role_code}'
          OR p.ROLE_ID LIKE 'dart_2024-06-30_role-{role_code}_' || '%'
          OR REGEXP_MATCHES(p.ROLE_ID, '^dart_.*role-{role_code}[a-z]?$')
        )
    ),
    labels AS (
      SELECT CIK, ELMT_ID,
             MAX(CASE WHEN LANG='ko' THEN LABEL END) AS ko_label,
             MAX(CASE WHEN LANG='en' THEN LABEL END) AS en_label
      FROM lab_insurers
      WHERE LABEL_ROLE_URI = 'http://www.xbrl.org/2003/role/label'
        AND CIK IN ({cik_in})
      GROUP BY CIK, ELMT_ID
    ),
    elem_meta AS (
      SELECT DISTINCT ELEMENT_ID, ABSTRACT, PERIOD_TYPE, BALANCE
      FROM elmt_raw
    )
    SELECT
      ae.ELEMENT_ID                      AS element_id,
      COALESCE(MAX(l.ko_label), '')      AS ko_label,
      COALESCE(MAX(l.en_label), '')      AS en_label,
      MAX(em.ABSTRACT)                   AS abstract,
      MAX(em.PERIOD_TYPE)                AS period_type,
      MAX(em.BALANCE)                    AS balance,
      LIST(DISTINCT ae.CIK ORDER BY ae.CIK) AS using_ciks,
      COUNT(DISTINCT ae.CIK)             AS n_companies
    FROM all_elems ae
    LEFT JOIN labels l ON l.ELMT_ID = ae.ELEMENT_ID AND l.CIK = ae.CIK
    LEFT JOIN elem_meta em ON em.ELEMENT_ID = ae.ELEMENT_ID
    GROUP BY ae.ELEMENT_ID
    """
    df = con.execute(sql).df()
    if df.empty:
        return df

    df["is_entity"] = df["element_id"].str.startswith("entity")
    df["element_short"] = df["element_id"].map(_short_element)

    # Expand using_ciks (numpy array from DuckDB LIST) into 9 company columns
    name_map = peer_groups.load_companies()

    def _in_list(arr, c):
        if arr is None:
            return False
        try:
            return c in list(arr)
        except TypeError:
            return False

    for cik in CIKS:
        col = name_map[cik].name_ko if cik in name_map else cik
        df[col] = df["using_ciks"].map(lambda s, c=cik: "✓" if _in_list(s, c) else "")
    df = df.drop(columns=["using_ciks"])

    # Reorder
    cols = ["element_short", "ko_label", "en_label", "abstract", "period_type", "balance",
            "n_companies", "is_entity", "element_id"] + \
           [name_map[c].name_ko for c in CIKS]
    return df[cols].sort_values(["is_entity", "n_companies"], ascending=[True, False])


def _write_role_sheet(ws, df: pd.DataFrame, role_code: str, role_label: str) -> None:
    ws.cell(row=1, column=1, value=f"{role_code} — {role_label}").font = TITLE_FONT
    ws.cell(row=2, column=1, value=f"총 {len(df)} elements · 9개사 사용 매트릭스 · entity 확장 강조").font = SUBTITLE_FONT

    name_map = peer_groups.load_companies()
    headers = ["element_short", "ko_label", "abstract", "period_type", "balance", "n",
               "is_entity"] + [name_map[c].name_ko for c in CIKS] + ["element_id"]
    display = df.rename(columns={"n_companies": "n"})[headers]

    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDER

    for ri, (_, row) in enumerate(display.iterrows(), start=5):
        is_abstract = str(row["abstract"]).lower() == "true"
        is_entity = bool(row["is_entity"])
        for ci, h in enumerate(headers, start=1):
            v = row[h]
            if isinstance(v, bool):
                v = "Y" if v else ""
            c = ws.cell(row=ri, column=ci, value=v)
            c.border = BORDER
            if h == "미래에셋생명" and v == "✓":
                c.fill = SELF_FILL
            elif is_entity and ci <= 9:
                c.fill = ENTITY_FILL
            elif is_abstract and ci <= 9:
                c.fill = ABSTRACT_FILL
            if h in {"ko_label", "en_label", "element_id", "element_short"}:
                c.alignment = Alignment(vertical="center")

    widths = [40, 50, 10, 12, 10, 5, 8] + [9] * len(CIKS) + [55]
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.freeze_panes = "B5"


def _write_summary(ws, role_dfs: dict[str, pd.DataFrame]) -> None:
    ws.cell(row=1, column=1, value="요약: role × 회사 element 보유 수").font = TITLE_FONT
    ws.cell(row=2, column=1, value="각 cell = 그 회사가 해당 role에 보고한 unique element 수").font = SUBTITLE_FONT

    name_map = peer_groups.load_companies()
    headers = ["role", "설명", "총 element"] + [name_map[c].name_ko for c in CIKS]
    for ci, h in enumerate(headers, start=1):
        c = ws.cell(row=4, column=ci, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="center")
        c.border = BORDER

    r = 5
    for (role_code, role_label), df in zip(ROLES, [role_dfs[r[0]] for r in ROLES], strict=True):
        per_cik = []
        for cik in CIKS:
            col = name_map[cik].name_ko
            per_cik.append(int((df[col] == "✓").sum()))
        ws.cell(row=r, column=1, value=role_code).font = Font(bold=True)
        ws.cell(row=r, column=2, value=role_label)
        ws.cell(row=r, column=3, value=len(df))
        for i, n in enumerate(per_cik):
            cell = ws.cell(row=r, column=4 + i, value=n)
            cell.alignment = Alignment(horizontal="right")
            if cik := CIKS[i]:
                if cik == "00112332":
                    cell.fill = SELF_FILL
        r += 1

    widths = [12, 32, 10] + [11] * len(CIKS)
    for ci, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def main() -> int:
    db = Path("data/db/benchmark.duckdb")
    out = Path("report/role_inventory.xlsx")
    out.parent.mkdir(parents=True, exist_ok=True)

    con = duckdb.connect(str(db), read_only=True)
    print(f"target: {out}")
    print(f"roles: {len(ROLES)}, companies: {len(CIKS)}")

    role_dfs = {}
    for code, label in ROLES:
        print(f"  fetching {code} {label} ...")
        df = fetch_role_inventory(con, code)
        role_dfs[code] = df
        print(f"    {len(df)} unique elements")

    con.close()

    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)

    ws_sum = wb.create_sheet("요약")
    _write_summary(ws_sum, role_dfs)

    for code, label in ROLES:
        ws = wb.create_sheet(code)
        _write_role_sheet(ws, role_dfs[code], code, label)

    wb.save(out)
    print(f"\n✓ wrote {out} ({out.stat().st_size / 1024:.1f} KB)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
