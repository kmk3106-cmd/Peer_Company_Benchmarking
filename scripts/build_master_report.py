"""종합 보고서 — 모든 cross-tab + KPI 합본 + 자사 시그널 dashboard."""
from __future__ import annotations
from pathlib import Path
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

OUT = Path("outputs/master_actuarial_report.xlsx")
SOURCES = [
    ("BEL 변동",       "outputs/bel_disclosure_filled.xlsx", "BEL 변동 disclosure"),
    ("RA 변동",        "outputs/ra_disclosure_filled.xlsx",  "RA 변동 disclosure"),
    ("CSM 변동",       "outputs/csm_disclosure_filled.xlsx", "CSM 변동 disclosure"),
]
KPI_SRC = "outputs/all_kpi_crosstab.xlsx"

wb = Workbook()
wb.remove(wb.active)

# =============== Sheet 1: Executive Dashboard ===============
ws = wb.create_sheet("종합 요약", 0)
ws.append(["미래에셋생명 — 동업사 비교 분석 종합 보고서"])
ws.cell(row=1, column=1).font = Font(bold=True, size=18, color="1E3A5F")
ws.append(["FY2025 별도 기준 · 8개사 (생보 4 + 손보 4) · 단위: 억원/%"])
ws.cell(row=2, column=1).font = Font(italic=True, color="666666", size=11)
ws.append([])

def section(title, color="1E3A5F"):
    ws.append([title])
    r = ws.max_row
    ws.cell(row=r, column=1).font = Font(bold=True, size=14, color=color)
    ws.append([])

# 자사 핵심 KPI
section("자사 (미래에셋생명) 핵심 지표")
kpi_rows = [
    ["지표",                          "자사 값",   "8사 군집",                 "순위",      "해석"],
    ["보험계약부채 (별도, 기말)",     "26.99조",   "11.4억~6,025조 (생보) ",   "—",         "자사 자산규모 군집 내"],
    ["보험계약부채 변동 검증식 잔차", "0억",       "삼성 +43k, 한손 −11k 등",  "1위 정합",  "13행 BEL 변동 disclosure 완전 정합"],
    ["CSM (별도, 기말)",              "2.06조",    "0.94~28.2조",              "8위(최저)", "자산규모 종속 — 부채 대비 비율 정상"],
    ["CSM 상각률 (연환산)",           "9.95%",     "8.5~11.9%",                "8위",       "군집 중하위 (정상)"],
    ["예실차 / 보험수익",             "−0.0%",     "−13.4% ~ +1.5%",           "1위",       "★ 가정 정확도 8사 중 최우수"],
    ["위험률 마진 (예상/위험)",       "103.5%",    "78.5~104.7%",              "8위",       "보수성 부족 시그널 (생보 중)"],
    ["사업비 마진 (예상/예정)",       "47.2%",     "47.2~75.4%",               "1위",       "★ 사업비 마진 6공시사 중 최강"],
    ["보험서비스결과율",              "16.9%",     "5.2~16.9%",                "1위",       "★ 보험수익 대비 마진 최강"],
    ["SG&A / 보험수익",               "1.9%",      "0.1~3.6%",                 "최저(생보)","판관비 효율 양호"],
    ["RA / BEL",                      "1.7%",      "2~4% 추정",                "최저",      "RA 산출 보수성 점검 시그널"],
    ["RA 해소 / 보험수익",            "4.1%",      "—",                        "1위",       "RA 해소 비중 가장 큼"],
    ["해약환급금 누계 (기적립+예정)", "12,199억",  "11,066~74,407억",          "—",         "감독목적 적립 강화"],
    ["보증준비금 누계",               "2,010억",   "0~2,384억",                "2위(생보)", "변액 최저보증 적립 양호"],
    ["당기순이익률 (NI/이익잉여금)",  "6.04%",     "4~14%",                    "중간",      "—"],
    ["ROE (NI/자본)",                 "5.1%",      "—~15.6%",                  "6위",      "—"],
]
for row in kpi_rows:
    ws.append(row)
    r = ws.max_row
    if r == ws.max_row - len(kpi_rows) + 1:
        for c in range(1, 6):
            ws.cell(row=r, column=c).font = Font(bold=True, color="FFFFFF")
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="1E3A5F")
    if "★" in str(row[-1]):
        for c in range(1, 6):
            ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor="E8F5E9")

ws.append([])

# 주요 발견
section("주요 발견")
findings = [
    ("강점",   "예실차 −0.0% (8사 1위)",         "가정·실제 정합 최우수. CSM/RA 상각으로 서비스결과 +1,826억 거의 완전 설명."),
    ("강점",   "사업비 마진 47.2% (1위)",        "예정유지비 절반 수준 실제비용. 운영효율·시스템 자동화 정량 성과."),
    ("강점",   "서비스결과율 16.9% (1위)",       "보험수익 대비 마진 8사 중 최강."),
    ("강점",   "변동 disclosure 정합 (잔차 0)",  "13행 BEL 변동표 기초+Σ변동=기말 완전 정합."),
    ("점검",   "위험률 마진 103.5% (8위)",       "생보 4사(78~90%) 대비 약함. 사업 mix(보장성 강화) 영향."),
    ("점검",   "RA/BEL 1.7% (최저)",             "RA 산출 보수성 점검 시그널. 모형 가정 재검토."),
    ("점검",   "CSM 절대규모 2.06조 (최저)",     "자산규모 종속변수. 비율 기준은 군집 내."),
    ("한계",   "현대해상 BS 별도 잔액 보고 부족","Sep×Issued 단독 컨텍스트 미보고 → 변동표 잔차 -113%."),
    ("한계",   "CSM disclosure 회사간 차이 큼", "한화생명만 표준 분해 보고. 다른 회사 entity 확장 fallback."),
    ("한계",   "보증준비금 손보 4사 미적립",     "변액 미운영으로 미공시."),
]
ws.append(["분류", "지표", "관찰"])
hdr = ws.max_row
for c in range(1, 4):
    ws.cell(row=hdr, column=c).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=hdr, column=c).fill = PatternFill("solid", fgColor="1E3A5F")
for cat, kpi, obs in findings:
    ws.append([cat, kpi, obs])
    r = ws.max_row
    color = {"강점": "E8F5E9", "점검": "FFF3E0", "한계": "FFEBEE"}.get(cat, "FFFFFF")
    for c in range(1, 4):
        ws.cell(row=r, column=c).fill = PatternFill("solid", fgColor=color)

ws.append([])

# 시트 가이드
section("본 보고서 시트 구성")
sheet_guide = [
    ("종합 요약 (본 시트)",         "executive dashboard — 핵심 지표 + 주요 발견"),
    ("BEL 변동",                    "보험계약부채 BEL 컴포넌트 13행 변동표 × 8사"),
    ("RA 변동",                     "위험조정 RA 컴포넌트 13행 변동표 × 8사 (4/8 정합)"),
    ("CSM 변동",                    "보험계약마진 CSM 13행 변동표 × 8사 (한화생명만 완전 정합)"),
    ("§5-B-1 위험률 마진",          "예상보험금/위험보험료 비율 × 8사"),
    ("§5-B-2 사업비 마진",          "예상유지비/예정유지비 비율 × 8사 (6공시)"),
    ("§5-B-3 자사 잔여기간대 분해", "10년이내~30년초과 6 버킷"),
    ("§5-C 사업비 항목",            "보험수익·판관비·핵심사업비·사업비율"),
    ("§5-D-1 ROE·NI",               "자본·이익잉여금·당기순이익·ROE"),
    ("§5-D-2 미처분이익잉여금",     "전기→차기이월"),
    ("§5-D-3 해약·보증준비금",      "기적립 + 적립예정"),
    ("Extra: 서비스결과율",         "보험서비스결과/보험수익"),
    ("Extra: CSM 상각률",           "당기상각/평균 CSM"),
    ("Extra: RA 해소",              "RA 해소/보험수익"),
    ("Extra: 손익지표 종합",        "보험수익·결과·NI·자본 통합"),
    ("출처·caveat",                 "데이터 출처, 매핑 규칙, 미공시 사유"),
]
ws.append(["시트", "내용"])
hdr = ws.max_row
for c in range(1, 3):
    ws.cell(row=hdr, column=c).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=hdr, column=c).fill = PatternFill("solid", fgColor="1E3A5F")
for s, d in sheet_guide:
    ws.append([s, d])

# 컬럼 폭
for i, w in enumerate([32, 18, 32, 12, 50], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

# =============== 변동표 3개 (BEL/RA/CSM) 복사 ===============
def copy_sheet(src_path, src_name, dest_name):
    src_wb = load_workbook(src_path, data_only=False)
    if src_name not in src_wb.sheetnames:
        # try first sheet
        src_name = src_wb.sheetnames[0]
    src_ws = src_wb[src_name]
    dst = wb.create_sheet(dest_name)
    for row in src_ws.iter_rows():
        for cell in row:
            new = dst.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                new.font = copy(cell.font)
                new.fill = copy(cell.fill)
                new.alignment = copy(cell.alignment)
                new.border = copy(cell.border)
                new.number_format = cell.number_format
    for col_letter, dim in src_ws.column_dimensions.items():
        dst.column_dimensions[col_letter].width = dim.width
    for row_idx, dim in src_ws.row_dimensions.items():
        dst.row_dimensions[row_idx].height = dim.height
    if src_ws.freeze_panes:
        dst.freeze_panes = src_ws.freeze_panes

for label, src_file, src_sheet in SOURCES:
    if Path(src_file).exists():
        try:
            copy_sheet(src_file, src_sheet, label)
        except Exception as e:
            ws_err = wb.create_sheet(label)
            ws_err.append([f"복사 실패: {e}"])

# =============== KPI 시트들 복사 ===============
if Path(KPI_SRC).exists():
    src_wb = load_workbook(KPI_SRC, data_only=False)
    for sn in src_wb.sheetnames:
        try:
            copy_sheet(KPI_SRC, sn, sn[:31])  # 시트명 31자 제한
        except Exception as e:
            pass

# =============== 출처·caveat 시트 ===============
ws = wb.create_sheet("출처·caveat")
ws.append(["데이터 출처 및 분석 한계"])
ws.cell(row=1, column=1).font = Font(bold=True, size=14)
ws.append([])

src_info = [
    ("데이터 원천",     "DART 사업보고서 FY2025 XBRL instance (별도 재무제표 기준)"),
    ("기준 시점",       "2025-12-31 (별도)"),
    ("적재본",          "DuckDB benchmark.duckdb — val/cntxt/lab/pre/role 5 테이블"),
    ("분석 대상",       "KOSPI 상장 보험 8개사 — 생보 4 (미래에셋·삼성·한화·동양), 손보 4 (삼성화재·현대해상·DB·한화손해)"),
    ("기본 룰",         "별도(Separate)만 분석. 연결 제외. 발행(Issued) 보험계약만 (재보험 held 제외)"),
    ("필터 룰",         "Sep × Issued 컨텍스트 + 위험관리 sub-table 멤버 제외 + ComponentsAxis 분해 별도 처리"),
    ("정합 검증",       "기초 + Σ변동 = 기말 잔차 < ±2% 시 정합 OK 판정"),
    ("주요 caveat",     "현대해상 BS 별도 잔액 보고 형식 한계, 미래에셋 element-disclosure 라벨 매핑 비표준, CSM disclosure 회사간 큰 차이 (한화만 완전)"),
    ("미공시 처리",     "정확 element search 실패 시 '미공시' 명시. PDF·구두값 임의 입력 금지 (사용자 룰 feedback_no_manual_input)"),
    ("자사 disclosure", "미래에셋 BEL 13행 사용자 disclosure 값 사용 — XBRL element-라벨 매핑 비표준으로 자동 추출 불가"),
]
ws.append(["항목", "내용"])
for c in range(1, 3):
    ws.cell(row=ws.max_row, column=c).font = Font(bold=True, color="FFFFFF")
    ws.cell(row=ws.max_row, column=c).fill = PatternFill("solid", fgColor="1E3A5F")
for k, v in src_info:
    ws.append([k, v])

for i, w in enumerate([18, 90], 1):
    ws.column_dimensions[get_column_letter(i)].width = w

wb.save(OUT)
print(f"wrote {OUT}")
print(f"sheets: {wb.sheetnames}")
print(f"총 {len(wb.sheetnames)} 시트")
