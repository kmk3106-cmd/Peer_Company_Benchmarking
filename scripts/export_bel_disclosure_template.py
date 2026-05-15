"""§4-1A BEL 변동 — 사용자 disclosure 표 13행 동업사 cross-tab.

방식:
  - 자사(미래에셋): 사용자 제공 disclosure 값 직접 사용 (XBRL 자동매핑 어려움 caveat)
  - 7개 동업사: terseLabel = '<row label>' 매칭으로 element_id 찾고 값 추출
"""
from __future__ import annotations
import duckdb
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

PEERS = [
    ("00112332", "미래에셋생명"),
    ("00126256", "삼성생명"),
    ("00113058", "한화생명"),
    ("00117267", "동양생명"),
    ("00139214", "삼성화재"),
    ("00164973", "현대해상"),
    ("00159102", "DB손해보험"),
    ("00135917", "한화손해보험"),
]
SEP = "ifrs-full_SeparateMember"
ISSUED = "ifrs-full_InsuranceContractsIssuedMember"

con = duckdb.connect("data/db/benchmark.duckdb", read_only=True)

# disclosure 표 13행 (라벨, 기간, 자사 사용자 값 원 단위)
ROWS = [
    ("부채인 보험계약의 기초 장부금액",            "opening", 26_248_411_835_353),
    ("보험수익",                                  "duration",-1_080_411_547_554),
    ("발생한 보험금 및 기타 보험서비스비용",      "duration",  +592_626_787_607),
    ("보험취득현금흐름 상각",                     "duration",  +218_864_437_494),
    ("발생사고요소의 조정",                       "duration",   +19_584_794_969),
    ("손실부담계약 관련 손실(환입)",              "duration",   +66_710_164_165),
    ("수취한 보험료",                             "duration",+4_050_027_788_462),
    ("보험취득현금흐름 지급",                     "duration",  -594_834_366_070),
    ("보험금(투자요소 포함) 및 기타보험서비스비용의 지급", "duration", -4_296_950_166_750),
    ("당기손익인식 보험금융손익",                 "duration",+2_272_778_949_192),
    ("기타포괄손익인식 보험금융손익",             "duration",  -483_472_446_698),
    ("기타증감",                                  "duration",   -16_637_974_642),
    ("부채인 보험계약의 기말 장부금액",           "closing", 26_996_698_255_523),
]

# 라벨 simplified (terse 검색용)
LBL_SHORT = {
    "부채인 보험계약의 기초 장부금액": "부채인 보험계약",
    "부채인 보험계약의 기말 장부금액": "부채인 보험계약",
}

# 라벨로 element 찾기 (terseLabel 우선)
def find_eids(cik, label_text):
    rows = con.execute("""
        SELECT DISTINCT ELMT_ID, LABEL_ROLE_URI FROM lab_insurers
        WHERE CIK=? AND REPORT_DATE='20251231' AND LANG='ko' AND TRIM(LABEL) = ?
    """, [cik, label_text]).fetchall()
    terse = [r[0] for r in rows if "terseLabel" in (r[1] or "")]
    if terse: return terse
    return [r[0] for r in rows]

SQL_VAL = """
WITH ctxs AS (
  SELECT CONTEXT_ID, COUNT(*) AS n_ax,
         MAX(PERIOD_INSTANT) AS p_inst,
         MAX(PERIOD_START_DATE) AS p_start,
         MAX(PERIOD_END_DATE) AS p_end
  FROM cntxt_insurers WHERE CIK=? AND REPORT_DATE='20251231'
  GROUP BY CONTEXT_ID
  HAVING BOOL_OR(MEMBER_ELEMENT_ID = ?) AND BOOL_OR(MEMBER_ELEMENT_ID = ?)
     AND NOT BOOL_OR(MEMBER_ELEMENT_ID LIKE '%OfDisclosureOfNatureAndExtentOfRisks%')
     AND NOT BOOL_OR(AXIS_ELEMENT_ID = 'ifrs-full_InsuranceContractsByComponentsAxis')
),
fact AS (
  SELECT v.amount_krw, c.n_ax,
         CASE WHEN c.p_inst='2024-12-31' THEN 'opening'
              WHEN c.p_inst='2025-12-31' THEN 'closing'
              WHEN c.p_start='2025-01-01' AND c.p_end='2025-12-31' THEN 'duration'
              ELSE NULL END AS pk
  FROM val_insurers v JOIN ctxs c USING(CONTEXT_ID)
  WHERE v.CIK=? AND v.REPORT_DATE='20251231' AND v.ELEMENT_ID=? AND v.amount_krw IS NOT NULL
),
top_only AS (
  SELECT MIN(n_ax) AS min_ax FROM fact WHERE pk=?
)
SELECT SUM(f.amount_krw) FROM fact f, top_only t
WHERE f.n_ax = t.min_ax AND f.pk = ?
"""

def get_value(cik, eid, pk):
    r = con.execute(SQL_VAL, [cik, SEP, ISSUED, cik, eid, pk, pk]).fetchone()
    return r[0] if r and r[0] is not None else None

# 자사 disclosure 값 직접 사용
USER_VALUES = {(label, pk): v for label, pk, v in ROWS}

# 각 회사 × 행 값 계산
peer_values = {}  # (cik, label, pk) -> value (원 단위)
peer_eids = {}    # (cik, label) -> element_id used
for cik, name in PEERS:
    for label, pk, _ in ROWS:
        if cik == "00112332":
            # 자사: 사용자 값 직접
            peer_values[(cik, label, pk)] = USER_VALUES[(label, pk)]
            peer_eids[(cik, label)] = "(사용자 disclosure 직접 사용)"
            continue
        # 7개사: terseLabel 매칭
        search = LBL_SHORT.get(label, label)
        eids = find_eids(cik, search)
        total = None
        used = []
        for eid in eids:
            v = get_value(cik, eid, pk)
            if v is not None:
                total = (total or 0) + v
                used.append(eid)
        peer_values[(cik, label, pk)] = total
        peer_eids[(cik, label)] = ", ".join(e.replace("ifrs-full_", "")[:35] for e in used) if used else "(매칭 실패)"

# Excel
wb = Workbook()
ws = wb.active
ws.title = "BEL 변동 disclosure"

ws.append(["§4-1A 보험계약부채 변동표 — 동업사 cross-tab (별도, 발행, 억원)"])
ws.cell(row=1, column=1).font = Font(bold=True, size=14)
ws.append(["기준: 미래에셋 사업보고서 「21-1. 보험계약부채 변동분의 차이조정 공시」 13행 구조"])
ws.cell(row=2, column=1).font = Font(italic=True, color="666666", size=10)
ws.append(["미래에셋(자사) = 사업보고서 disclosure 값 직접 사용 · 7개사 = XBRL terseLabel 매칭 추출"])
ws.cell(row=3, column=1).font = Font(italic=True, color="C62828", size=10)
ws.append([])

hdrs = ["#", "disclosure 행 라벨", "기간"] + [n for _, n in PEERS]
ws.append(hdrs)
hr = ws.max_row
for c in range(1, len(hdrs)+1):
    cell = ws.cell(row=hr, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1E3A5F")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(left=Side(style="thin"), right=Side(style="thin"),
                         top=Side(style="thin"), bottom=Side(style="thin"))

opening_fill = PatternFill("solid", fgColor="FFF4CE")
closing_fill = PatternFill("solid", fgColor="FFE0B2")
revenue_fill = PatternFill("solid", fgColor="E3F2FD")
expense_fill = PatternFill("solid", fgColor="FFF3E0")
cf_fill      = PatternFill("solid", fgColor="E0F7FA")
fin_fill     = PatternFill("solid", fgColor="F3E5F5")
other_fill   = PatternFill("solid", fgColor="FAFAFA")

border = Border(left=Side(style="thin", color="CCCCCC"),
                right=Side(style="thin", color="CCCCCC"),
                top=Side(style="thin", color="CCCCCC"),
                bottom=Side(style="thin", color="CCCCCC"))

def fill_for(label, pk):
    if pk == "opening": return opening_fill
    if pk == "closing": return closing_fill
    if "보험수익" in label: return revenue_fill
    if any(w in label for w in ["발생한 보험금", "보험취득현금흐름 상각", "발생사고요소", "손실부담"]): return expense_fill
    if any(w in label for w in ["수취한", "지급", "투자요소"]): return cf_fill
    if "보험금융손익" in label: return fin_fill
    return other_fill

PK_LABEL = {"opening": "기초", "duration": "변동", "closing": "기말"}
for i, (label, pk, _) in enumerate(ROWS, 1):
    row = [i, label, PK_LABEL[pk]]
    for cik, _ in PEERS:
        v = peer_values.get((cik, label, pk))
        # 억원 변환
        row.append(round(v / 1e8, 0) if v is not None else None)
    ws.append(row)
    r = ws.max_row
    fill = fill_for(label, pk)
    for c in range(1, len(hdrs)+1):
        ws.cell(row=r, column=c).fill = fill
        ws.cell(row=r, column=c).border = border
    for c in range(4, len(hdrs)+1):
        ws.cell(row=r, column=c).number_format = '#,##0;-#,##0;"–"'

# 검증식
ws.append([])
ws.append(["검증", "기초 + Σ변동 = 기말 ?", "", ""] + [""]*len(PEERS))
sum_row_idx = ws.max_row
for i, (cik, _) in enumerate(PEERS, start=4):
    col = get_column_letter(i)
    # rows: hr+1 (opening) to hr+12 (last duration) to hr+13 (closing)
    op_row = hr + 1
    cl_row = hr + 13
    dur_start = hr + 2
    dur_end = hr + 12
    formula = f"={col}{op_row}+SUM({col}{dur_start}:{col}{dur_end})-{col}{cl_row}"
    ws.cell(row=sum_row_idx, column=i).value = formula
    ws.cell(row=sum_row_idx, column=i).number_format = '#,##0;-#,##0;"– (정합)"'
for c in range(1, len(hdrs)+1):
    ws.cell(row=sum_row_idx, column=c).fill = PatternFill("solid", fgColor="E8F5E9")
    ws.cell(row=sum_row_idx, column=c).border = border
ws.cell(row=sum_row_idx, column=1).font = Font(bold=True)
ws.cell(row=sum_row_idx, column=2).value = "잔차 (기초 + Σ변동 − 기말)"
ws.cell(row=sum_row_idx, column=2).font = Font(bold=True)

# 컬럼 폭
widths = [4, 40, 6] + [13]*len(PEERS)
for i, w in enumerate(widths, 1):
    ws.column_dimensions[get_column_letter(i)].width = w
ws.row_dimensions[hr].height = 28
ws.freeze_panes = f"D{hr+1}"

# 진단 시트: 회사별 element_id 매핑
diag = wb.create_sheet(title="element 매핑")
diag.append(["회사별 disclosure 행 라벨 → 사용된 element_id"])
diag.cell(row=1, column=1).font = Font(bold=True, size=12)
diag.append([])
diag.append(["disclosure 행"] + [n for _, n in PEERS])
hr2 = diag.max_row
for c in range(1, 2 + len(PEERS)):
    cell = diag.cell(row=hr2, column=c)
    cell.font = Font(bold=True, color="FFFFFF")
    cell.fill = PatternFill("solid", fgColor="1E3A5F")
    cell.alignment = Alignment(horizontal="center", wrap_text=True)

for label, pk, _ in ROWS:
    row = [label]
    for cik, _ in PEERS:
        row.append(peer_eids.get((cik, label), ""))
    diag.append(row)
    r = diag.max_row
    for c in range(1, 2 + len(PEERS)):
        diag.cell(row=r, column=c).border = border
        diag.cell(row=r, column=c).alignment = Alignment(vertical="top", wrap_text=True)

for i, w in enumerate([35] + [26]*len(PEERS), 1):
    diag.column_dimensions[get_column_letter(i)].width = w
diag.freeze_panes = f"B{hr2+1}"

out = Path("outputs/bel_disclosure_template_crosstab.xlsx")
wb.save(out)
print(f"wrote {out}")
print(f"sheets: {wb.sheetnames}")
