"""Render layer smoke tests (Excel + HTML)."""

from __future__ import annotations

from pathlib import Path

import duckdb
import openpyxl
import pytest

from peer_benchmarking.render import excel, plotly_html

DB_PATH = Path(__file__).parent.parent / "data" / "db" / "benchmark.duckdb"
HAS_DB = DB_PATH.exists() and DB_PATH.stat().st_size > 1_000_000


@pytest.fixture(scope="module")
def con():
    if not HAS_DB:
        pytest.skip("benchmark.duckdb not present — run Phase 2 ingest first")
    c = duckdb.connect(str(DB_PATH), read_only=True)
    yield c
    c.close()


def test_excel_report_creates_all_sheets(con, tmp_path):
    out = tmp_path / "report.xlsx"
    excel.build_report(con, out)
    assert out.exists()
    wb = openpyxl.load_workbook(out)
    sheets = wb.sheetnames
    # core sheets must exist
    assert "개요" in sheets
    assert any("보험계약부채" in s for s in sheets)
    assert any("회계모형" in s for s in sheets)
    assert any("구성요소" in s for s in sheets)


def test_excel_overview_contains_self_summary(con, tmp_path):
    out = tmp_path / "report.xlsx"
    excel.build_report(con, out)
    wb = openpyxl.load_workbook(out)
    overview_text = "\n".join(
        " ".join(str(c) for c in row if c is not None)
        for row in wb["개요"].iter_rows(values_only=True)
    )
    assert "미래에셋생명" in overview_text
    assert "조원" in overview_text


def test_html_report_has_plotly_charts(con, tmp_path):
    out = tmp_path / "report.html"
    plotly_html.build_html_report(con, out)
    assert out.exists()
    html = out.read_text(encoding="utf-8")
    # 5 charts expected
    assert html.count("Plotly.newPlot") == 5
    assert "미래에셋생명" in html
    # CDN-loaded plotly.js, embedded once
    assert html.count("cdn.plot.ly/plotly") + html.count("cdn.plotly.com/plotly") >= 1
